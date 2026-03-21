from typing import List, Any, Optional, Dict
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, status, Query, BackgroundTasks, Body, Request, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date
import io
import json

from app.database import get_db
from app.core.security import get_current_corporate_admin_context, get_current_approver_context, get_current_treasury_context, get_issuance_read_context, check_subscription_status, TokenData
from app.core.document_generator import generate_pdf_from_html
from app.core.encryption import encrypt_data 

# Models
from app.models.models import Bank, Currency 
from app.models.models_issuance import IssuedLGRecord, IssuanceRequest, IssuanceFacilitySubLimit, IssuanceFacility, IssuanceWorkflowPolicy, CustomerFormConfiguration, IssuanceRequestSnapshot, IssuanceRequestVersion, AdminChangeRequest, BankFormIssueReport
# NOTE: Ensure you created app/models/models_reconciliation.py first!
from app.models.models_reconciliation import BankPositionBatch, BankPositionRow 
# Schemas
from app.schemas.all_schemas import BankOut, CurrencyOut 
from app.schemas.schemas_issuance import (
    IssuanceRequestCreate, IssuanceRequestOut, IssuanceRequestUpdate, IssuanceRequestDraftCreate,
    CustomerFormConfigurationCreateUpdate, IssuanceRequestVersionOut,
    IssuanceFacilityCreate, IssuanceFacilityOut, SuitableFacilityOut, 
    IssuanceRequestContentUpdate, IssuanceFacilityUpdate,
    IssuedLGRecordOut, IssuedLGRecordDetailOut,
    IssuanceExecuteRequest, IssuanceCancelRequest,
    ReconciliationRequest, ReconciliationResult,
    IssuanceWorkflowPolicyCreate, IssuanceWorkflowPolicyOut,
    BankIssuanceOptionOut, BankIssuanceOptionCreateUpdate,
    AdminChangeRequestCreate, AdminChangeRequestOut, AdminChangeRequestAction,
    BankFormIssueReportCreate, BankFormIssueReportOut, BankFormIssueReportUpdate
)
from app.services.issuance_service import issuance_service

# CRUD
from app.crud.crud_issuance import crud_issuance_request
from app.crud.crud_facility import crud_facility
from app.crud.crud_bank_methods import crud_bank_methods
from fastapi.responses import StreamingResponse

router = APIRouter()

# ==============================================================================
# 1. DICTIONARIES (Banks & Currencies)
# ==============================================================================

@router.get("/banks", response_model=List[BankOut])
def get_issuance_banks(db: Session = Depends(get_db)):
    """Fetch all banks for dropdowns."""
    return db.query(Bank).all()

@router.get("/currencies", response_model=List[CurrencyOut])
def get_issuance_currencies(db: Session = Depends(get_db)):
    """Fetch all currencies for dropdowns."""
    return db.query(Currency).all()

# ==============================================================================
# BANK ISSUANCE METHODS (LIBRARY)
# ==============================================================================

@router.get("/bank-methods", response_model=List[BankIssuanceOptionOut])
def list_bank_methods(
    bank_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """List all issuance methods for a specific bank."""
    return crud_bank_methods.get_by_bank(db, bank_id=bank_id, only_active=False)

@router.post("/bank-methods", response_model=BankIssuanceOptionOut)
def create_bank_method(
    bank_id: int,
    method_in: BankIssuanceOptionCreateUpdate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Create a new issuance method for a bank."""
    return crud_bank_methods.create_method(db, bank_id=bank_id, obj_in=method_in.model_dump(), user_id=current_user.user_id)

@router.put("/bank-methods/{method_id}", response_model=BankIssuanceOptionOut)
def update_bank_method(
    method_id: int,
    method_in: BankIssuanceOptionCreateUpdate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Update an existing bank issuance method."""
    return crud_bank_methods.update_method(db, method_id=method_id, obj_in=method_in.model_dump(), user_id=current_user.user_id)

@router.delete("/bank-methods/{method_id}", response_model=BankIssuanceOptionOut)
def deactivate_bank_method(
    method_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Deactivate a bank issuance method."""
    return crud_bank_methods.update_method(db, method_id=method_id, obj_in={"is_active": False}, user_id=current_user.user_id)

# ==============================================================================
# LG TYPES (read-only for all authenticated users)
# ==============================================================================

@router.get("/lg-types")
def list_lg_types(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Return all active LG types for dropdowns."""
    from app.models import LgType
    types = db.query(LgType).filter(LgType.is_deleted == False).order_by(LgType.name).all()
    return [{"id": t.id, "name": t.name} for t in types]

# ==============================================================================
# BANK FORM TEMPLATES (Upload, AI Analysis, Fill)
# ==============================================================================

from fastapi import File, UploadFile
from app.models.models_issuance import BankFormTemplate

@router.post("/bank-forms/upload")
async def upload_bank_form(
    bank_id: int = Query(...),
    form_name: str = Query(...),
    form_type: str = Query("FILLABLE_PDF"),
    form_language: str = Query("BILINGUAL", description="AR / EN / BILINGUAL"),
    lg_type_ids: str = Query(None, description="Optional: comma-separated LG type IDs this form covers. NULL = universal."),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    # Allow System Owner and Corporate Admin
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges: Requires System Owner or Corporate Admin role.")
    """
    Upload a bank's PDF form template. Creates a BankFormTemplate record
    and stores the file locally. Does NOT trigger AI analysis automatically.
    """
    import os
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(400, "Only PDF files are accepted.")
    
    pdf_bytes = await file.read()
    
    # Store the file locally
    upload_dir = os.path.join("uploads", "bank_forms", str(bank_id))
    os.makedirs(upload_dir, exist_ok=True)
    
    # Check for existing active forms for this bank to determine version
    existing = db.query(BankFormTemplate).filter(
        BankFormTemplate.bank_id == bank_id,
        BankFormTemplate.is_active == True,
        BankFormTemplate.is_deleted == False,
    ).order_by(BankFormTemplate.version.desc()).first()
    
    new_version = (existing.version + 1) if existing else 1
    
    # Save file
    safe_filename = f"v{new_version}_{file.filename}"
    file_path = os.path.join(upload_dir, safe_filename)
    with open(file_path, "wb") as f:
        f.write(pdf_bytes)
    
    # Try to extract interactive form fields
    try:
        from app.core.pdf_form_filler import get_pdf_form_fields
        detected_fields = get_pdf_form_fields(pdf_bytes)
    except Exception:
        detected_fields = []
    
    # Create DB record
    form_template = BankFormTemplate(
        bank_id=bank_id,
        name=form_name,
        version=new_version,
        form_type=form_type,
        form_language=form_language if form_language in ('AR', 'EN', 'BILINGUAL') else 'BILINGUAL',
        lg_type_ids=[int(x.strip()) for x in lg_type_ids.split(',') if x.strip()] if lg_type_ids else None,
        file_path=file_path,
        original_filename=file.filename,
        ai_analysis_status="PENDING",
        is_active=True,
        uploaded_by=current_user.user_id if hasattr(current_user, 'user_id') else None,
    )
    
    # If we detected interactive fields, store them in the AI analysis
    if detected_fields:
        form_template.ai_analysis = {"detected_interactive_fields": detected_fields}
    
    db.add(form_template)
    db.commit()
    db.refresh(form_template)
    
    return {
        "id": form_template.id,
        "name": form_template.name,
        "version": form_template.version,
        "form_type": form_template.form_type,
        "bank_id": form_template.bank_id,
        "original_filename": form_template.original_filename,
        "ai_analysis_status": form_template.ai_analysis_status,
        "detected_fields_count": len(detected_fields),
        "message": f"Form uploaded successfully (v{new_version}). Run AI analysis to auto-map fields."
    }


@router.post("/bank-forms/{form_id}/analyze")
async def analyze_bank_form(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    # Allow System Owner and Corporate Admin
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges: Requires System Owner or Corporate Admin role.")
    """
    Triggers AI analysis on an uploaded bank form to auto-map fields.
    This is called ONCE per form upload. The result is cached and reused.
    """
    form_template = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    
    if not form_template:
        raise HTTPException(404, "Bank form template not found.")
    
    if not form_template.file_path:
        raise HTTPException(400, "No PDF file associated with this form template.")
    
    # Read the PDF file
    import os
    if not os.path.exists(form_template.file_path):
        raise HTTPException(404, "PDF file not found on disk.")
    
    with open(form_template.file_path, "rb") as f:
        pdf_bytes = f.read()
    
    # Update status
    form_template.ai_analysis_status = "ANALYZING"
    db.commit()
    
    try:
        from app.core.ai_integration import analyze_bank_form_pdf
        from app.core.pdf_form_filler import get_pdf_form_fields
        
        # Get detected fields
        try:
            detected_fields = get_pdf_form_fields(pdf_bytes)
        except Exception:
            detected_fields = []
        
        # Run AI analysis
        result = await analyze_bank_form_pdf(
            pdf_bytes=pdf_bytes,
            filename=form_template.original_filename or "bank_form.pdf",
            detected_fields=detected_fields,
            form_type=form_template.form_type or "FILLABLE_PDF",
        )
        
        # Cache the results
        form_template.field_mapping = result.get("field_mapping", [])
        form_template.ai_analysis = result
        form_template.ai_analysis_status = "COMPLETED"
        
        # Update name if AI detected a better one
        if result.get("form_title") and not form_template.name:
            form_template.name = result["form_title"]
        
        db.commit()
        db.refresh(form_template)
        
        return {
            "id": form_template.id,
            "status": "COMPLETED",
            "form_title": result.get("form_title", ""),
            "bank_name_detected": result.get("bank_name_detected", ""),
            "total_fields": result.get("total_fields", 0),
            "mapped_fields": result.get("mapped_fields", 0),
            "unmapped_fields": result.get("unmapped_fields", []),
            "field_mapping": result.get("field_mapping", []),
            "form_notes": result.get("form_notes", ""),
        }
        
    except Exception as e:
        form_template.ai_analysis_status = "FAILED"
        form_template.ai_analysis = {"error": str(e)}
        db.commit()
        raise HTTPException(500, f"AI analysis failed: {str(e)}")


@router.get("/bank-forms")
def list_bank_forms(
    bank_id: int = Query(None),
    include_archived: bool = Query(False, description="Include suspended and deleted forms"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """List bank form templates, optionally filtered by bank. Use include_archived=true to see deleted/suspended forms."""
    if include_archived:
        query = db.query(BankFormTemplate)  # Show all including deleted/suspended
    else:
        query = db.query(BankFormTemplate).filter(
            BankFormTemplate.is_active == True,
            BankFormTemplate.is_deleted == False,
        )
    if bank_id:
        query = query.filter(BankFormTemplate.bank_id == bank_id)
    
    forms = query.order_by(BankFormTemplate.bank_id, BankFormTemplate.priority.desc(), BankFormTemplate.version.desc()).all()
    
    return [
        {
            "id": f.id,
            "bank_id": f.bank_id,
            "bank_name": f.bank.name if f.bank else "Unknown",
            "name": f.name,
            "version": f.version,
            "form_type": f.form_type,
            "form_language": getattr(f, 'form_language', 'BILINGUAL') or 'BILINGUAL',
            "original_filename": f.original_filename,
            "ai_analysis_status": f.ai_analysis_status,
            "lg_type_ids": f.lg_type_ids or [],
            "mapped_fields_count": len(f.field_mapping) if f.field_mapping else 0,
            "is_active": f.is_active,
            "is_deleted": f.is_deleted,
            "priority": f.priority or 0,
            "created_at": f.created_at.isoformat() if f.created_at else None,
        }
        for f in forms
    ]


@router.get("/bank-forms/{form_id}")
def get_bank_form(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Get full details of a bank form template including field mapping."""
    form_template = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    
    if not form_template:
        raise HTTPException(404, "Bank form template not found.")
    
    return {
        "id": form_template.id,
        "bank_id": form_template.bank_id,
        "bank_name": form_template.bank.name if form_template.bank else "Unknown",
        "name": form_template.name,
        "version": form_template.version,
        "form_type": form_template.form_type,
        "form_language": getattr(form_template, 'form_language', 'BILINGUAL') or 'BILINGUAL',
        "original_filename": form_template.original_filename,
        "file_path": form_template.file_path,
        "ai_analysis_status": form_template.ai_analysis_status,
        "lg_type_ids": form_template.lg_type_ids or [],
        "field_mapping": form_template.field_mapping,
        "ai_analysis": form_template.ai_analysis,
        "is_active": form_template.is_active,
        "priority": form_template.priority or 0,
        "created_at": form_template.created_at.isoformat() if form_template.created_at else None,
        "updated_at": form_template.updated_at.isoformat() if form_template.updated_at else None,
    }


@router.put("/bank-forms/{form_id}/mapping")
def update_bank_form_mapping(
    form_id: int,
    mapping: List[dict] = Body(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges.")
    """Manually override/fine-tune the AI-generated field mapping."""
    form_template = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    
    if not form_template:
        raise HTTPException(404, "Bank form template not found.")
    
    form_template.field_mapping = mapping
    db.commit()
    
    return {"message": "Field mapping updated", "id": form_template.id}


@router.delete("/bank-forms/{form_id}")
def delete_bank_form(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Soft-deletes a bank form template."""
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges.")
    
    form = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    if not form:
        raise HTTPException(404, "Bank form template not found.")
    
    form.is_deleted = True
    form.is_active = False
    db.commit()
    return {"message": f"Form '{form.name}' deleted.", "id": form_id}


@router.patch("/bank-forms/{form_id}/restore")
def restore_bank_form(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Restores a soft-deleted bank form template."""
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges.")
    
    form = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
    ).first()
    if not form:
        raise HTTPException(404, "Bank form template not found.")
    
    form.is_deleted = False
    form.is_active = True
    db.commit()
    return {"message": f"Form '{form.name}' restored.", "id": form_id, "is_active": True, "is_deleted": False}


@router.patch("/bank-forms/{form_id}/toggle-active")
def toggle_bank_form_active(
    form_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Toggles a bank form between active and suspended."""
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges.")
    
    form = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    if not form:
        raise HTTPException(404, "Bank form template not found.")
    
    form.is_active = not form.is_active
    db.commit()
    status = "active" if form.is_active else "suspended"
    return {"message": f"Form '{form.name}' is now {status}.", "id": form_id, "is_active": form.is_active}


@router.patch("/bank-forms/{form_id}/priority")
def set_bank_form_priority(
    form_id: int,
    priority: int = Query(..., ge=0, le=100),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Sets the priority ranking of a bank form (0-100, higher = preferred)."""
    from app.constants import UserRole
    if current_user.role not in [UserRole.SYSTEM_OWNER, UserRole.CORPORATE_ADMIN]:
        raise HTTPException(403, "Not enough privileges.")
    
    form = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    if not form:
        raise HTTPException(404, "Bank form template not found.")
    
    form.priority = priority
    db.commit()
    return {"message": f"Form '{form.name}' priority set to {priority}.", "id": form_id, "priority": priority}


@router.post("/bank-forms/{form_id}/fill/{request_id}")
async def fill_bank_form(
    form_id: int,
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    Fill a bank's PDF form with data from an issuance request.
    Uses the cached field mapping (no AI call). Returns the filled PDF.
    """
    from sqlalchemy.orm import selectinload
    
    form_template = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    
    if not form_template:
        raise HTTPException(404, "Bank form template not found.")
    
    if not form_template.field_mapping:
        raise HTTPException(400, "Form has no field mapping. Run AI analysis first.")
    
    if not form_template.file_path:
        raise HTTPException(400, "No PDF file associated with this form.")
    
    import os
    if not os.path.exists(form_template.file_path):
        raise HTTPException(404, "PDF file not found on disk.")
    
    # Load the request with relationships
    request = db.query(IssuanceRequest).options(
        selectinload(IssuanceRequest.currency),
        selectinload(IssuanceRequest.lg_type),
        selectinload(IssuanceRequest.issuing_entity),
        selectinload(IssuanceRequest.customer),
        selectinload(IssuanceRequest.project),
    ).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id,
    ).first()
    
    if not request:
        raise HTTPException(404, "Issuance request not found.")
    
    # Build data dict
    from app.core.pdf_form_filler import fill_pdf_form, build_request_data_dict
    request_data = build_request_data_dict(request, db, bank_id=form_template.bank_id)
    
    # Read blank form
    with open(form_template.file_path, "rb") as f:
        template_pdf_bytes = f.read()
    
    # Fill the form
    filled_pdf = fill_pdf_form(
        template_pdf_bytes=template_pdf_bytes,
        field_mapping=form_template.field_mapping,
        request_data=request_data,
    )
    
    filename = f"Filled_{form_template.name}_{request.serial_number}.pdf"
    
    return StreamingResponse(
        io.BytesIO(filled_pdf),
        media_type="application/pdf",
        headers={'Content-Disposition': f'inline; filename="{filename}"'}
    )

# ==============================================================================
# CUSTOMER BANK ACCOUNTS
# ==============================================================================

from app.models.models_issuance import CustomerBankAccount

@router.get("/bank-accounts")
def list_bank_accounts(
    bank_id: int = Query(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """List all bank accounts for this customer, optionally filtered by bank."""
    query = db.query(CustomerBankAccount).filter(
        CustomerBankAccount.customer_id == current_user.customer_id,
        CustomerBankAccount.is_deleted == False,
    )
    if bank_id:
        query = query.filter(CustomerBankAccount.bank_id == bank_id)
    
    accounts = query.order_by(CustomerBankAccount.bank_id, CustomerBankAccount.is_default.desc()).all()
    
    return [
        {
            "id": a.id,
            "bank_id": a.bank_id,
            "bank_name": a.bank.name if a.bank else "Unknown",
            "entity_id": a.entity_id,
            "entity_name": a.entity.entity_name if a.entity else None,
            "account_name": a.account_name,
            "account_number": a.account_number,
            "customer_number": a.customer_number,
            "branch_name": a.branch_name,
            "iban": a.iban,
            "is_default": a.is_default,
            "is_active": a.is_active,
        }
        for a in accounts
    ]


@router.post("/bank-accounts")
def create_bank_account(
    body: dict,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context),
):
    """Create a new bank account for the customer."""
    # If setting as default, unset any existing default for same customer+bank
    if body.get("is_default"):
        existing_defaults = db.query(CustomerBankAccount).filter(
            CustomerBankAccount.customer_id == current_user.customer_id,
            CustomerBankAccount.bank_id == body["bank_id"],
            CustomerBankAccount.is_default == True,
            CustomerBankAccount.is_deleted == False,
        ).all()
        for d in existing_defaults:
            d.is_default = False
    
    account = CustomerBankAccount(
        customer_id=current_user.customer_id,
        bank_id=body["bank_id"],
        entity_id=body.get("entity_id"),
        account_name=body["account_name"],
        account_number=body["account_number"],
        customer_number=body.get("customer_number"),
        branch_name=body.get("branch_name"),
        iban=body.get("iban"),
        is_default=body.get("is_default", False),
    )
    db.add(account)
    db.commit()
    db.refresh(account)
    
    return {
        "id": account.id,
        "message": "Bank account created successfully",
    }


@router.put("/bank-accounts/{account_id}")
def update_bank_account(
    account_id: int,
    body: dict,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context),
):
    """Update a bank account."""
    account = db.query(CustomerBankAccount).filter(
        CustomerBankAccount.id == account_id,
        CustomerBankAccount.customer_id == current_user.customer_id,
        CustomerBankAccount.is_deleted == False,
    ).first()
    
    if not account:
        raise HTTPException(404, "Bank account not found.")
    
    # If setting as default, unset existing
    if body.get("is_default") and not account.is_default:
        existing_defaults = db.query(CustomerBankAccount).filter(
            CustomerBankAccount.customer_id == current_user.customer_id,
            CustomerBankAccount.bank_id == account.bank_id,
            CustomerBankAccount.is_default == True,
            CustomerBankAccount.is_deleted == False,
            CustomerBankAccount.id != account_id,
        ).all()
        for d in existing_defaults:
            d.is_default = False
    
    for field in ["account_name", "account_number", "customer_number", "branch_name", "iban", "is_default", "entity_id", "is_active"]:
        if field in body:
            setattr(account, field, body[field])
    
    db.commit()
    return {"message": "Bank account updated"}


@router.delete("/bank-accounts/{account_id}")
def delete_bank_account(
    account_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context),
):
    """Soft-delete a bank account."""
    account = db.query(CustomerBankAccount).filter(
        CustomerBankAccount.id == account_id,
        CustomerBankAccount.customer_id == current_user.customer_id,
        CustomerBankAccount.is_deleted == False,
    ).first()
    
    if not account:
        raise HTTPException(404, "Bank account not found.")
    
    account.is_deleted = True
    account.is_active = False
    db.commit()
    return {"message": "Bank account deleted"}


# ==============================================================================
# ADMIN FORM CONFIGURATION
# ==============================================================================

@router.get("/form-config", response_model=CustomerFormConfigurationCreateUpdate)
def get_form_configuration(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Fetch the custom form layout for the customer. Returns defaults if none exist."""
    config = db.query(CustomerFormConfiguration).filter(
        CustomerFormConfiguration.customer_id == current_user.customer_id
    ).first()
    
    if not config:
        # Return default empty config
        return CustomerFormConfigurationCreateUpdate()
        
    return CustomerFormConfigurationCreateUpdate(
        field_configurations=config.field_configurations,
        custom_field_1_config=config.custom_field_1_config,
        custom_field_2_config=config.custom_field_2_config,
        mandatory_document_types=config.mandatory_document_types,
        reference_types=config.reference_types,
        document_config=config.document_config
    )

@router.put("/form-config", response_model=CustomerFormConfigurationCreateUpdate)
def update_form_configuration(
    config_in: CustomerFormConfigurationCreateUpdate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Updates the form layout. Validates against hiding critical fields.
    Multi-admin: requires dual-control approval via AdminChangeRequest."""
    
    # Capture new value as serializable dict
    new_val = {
        "field_configurations": {k: v.model_dump() for k, v in config_in.field_configurations.items()},
        "mandatory_document_types": config_in.mandatory_document_types,
        "reference_types": config_in.reference_types,
        "document_config": config_in.document_config,
    }
    if config_in.custom_field_1_config:
        new_val["custom_field_1_config"] = config_in.custom_field_1_config.model_dump()
    if config_in.custom_field_2_config:
        new_val["custom_field_2_config"] = config_in.custom_field_2_config.model_dump()

    # Capture old value for audit
    existing = db.query(CustomerFormConfiguration).filter(
        CustomerFormConfiguration.customer_id == current_user.customer_id
    ).first()
    old_val = {}
    if existing:
        old_val = {
            "field_configurations": existing.field_configurations,
            "mandatory_document_types": existing.mandatory_document_types,
        }

    change_req, auto_approved = _create_governed_change(
        db, current_user.customer_id, current_user.user_id,
        "FORM_CONFIG_UPDATE", {"old_value": old_val, "new_value": new_val}
    )

    if auto_approved:
        # Single-admin: already applied by _apply_admin_change inside _create_governed_change
        return config_in

    # Multi-admin: return 202 — change is pending approval
    from fastapi.responses import JSONResponse
    return JSONResponse(
        status_code=202,
        content={
            "message": "Configuration change submitted for approval by a second administrator.",
            "change_request_id": change_req.id,
            "status": "PENDING"
        }
    )


# ==============================================================================
# DUPLICATE & SIMILARITY CHECKS
# ==============================================================================

@router.get("/check-duplicate-reference")
def check_duplicate_reference(
    reference_type: str = Query(...),
    reference_number: str = Query(...),
    exclude_request_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Check if a request with the same reference type + number exists for this customer."""
    from app.models.models_issuance import IssuanceRequest, IssuedLGRecord
    from sqlalchemy import func
    
    query = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == current_user.customer_id,
        func.lower(IssuanceRequest.reference_type) == reference_type.lower(),
        func.lower(IssuanceRequest.reference_number) == reference_number.strip().lower(),
        IssuanceRequest.is_deleted == False
    )
    if exclude_request_id:
        query = query.filter(IssuanceRequest.id != exclude_request_id)
    
    request_matches = query.order_by(IssuanceRequest.created_at.desc()).limit(3).all()
    
    # Also check against issued LGs (via their linked request's reference)
    lg_query = db.query(IssuedLGRecord).join(
        IssuanceRequest, IssuedLGRecord.request_id == IssuanceRequest.id
    ).filter(
        IssuedLGRecord.customer_id == current_user.customer_id,
        func.lower(IssuanceRequest.reference_type) == reference_type.lower(),
        func.lower(IssuanceRequest.reference_number) == reference_number.strip().lower(),
    ).limit(3).all()

    all_matches = []
    for m in request_matches:
        all_matches.append({
            "id": m.id,
            "serial_number": m.serial_number,
            "status": m.status,
            "amount": str(m.amount) if m.amount else None,
            "beneficiary_name": m.beneficiary_name,
            "created_at": str(m.created_at) if m.created_at else None,
            "type": "request"
        })
    for lg in lg_query:
        all_matches.append({
            "id": lg.id,
            "serial_number": lg.lg_ref_number,
            "status": f"ISSUED ({lg.status})",
            "amount": str(lg.current_amount) if lg.current_amount else None,
            "beneficiary_name": lg.beneficiary_name,
            "created_at": str(lg.created_at) if lg.created_at else None,
            "type": "issued_lg"
        })
    
    if not all_matches:
        return {"found": False, "matches": []}
    
    return {"found": True, "matches": all_matches}


@router.get("/requests/{request_id}/similarity-check")
def check_similarity_against_issued_lgs(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    Scores the given request against all issued LGs for this customer.
    Weights: ref_type+number 30%, beneficiary_name 25%, amount 20%, lg_type 15%, expiry 10%.
    Returns matches ≥70%.
    """
    from app.models.models_issuance import IssuanceRequest, IssuedLGRecord
    from difflib import SequenceMatcher
    
    request = db.query(IssuanceRequest).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id
    ).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Fetch all issued LGs for this customer (active or expired, within last 3 years)
    # Exclude LGs that were issued from THIS request (self-match prevention)
    from datetime import datetime, timedelta
    cutoff = datetime.now() - timedelta(days=3*365)
    issued_lgs = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == current_user.customer_id,
        IssuedLGRecord.created_at >= cutoff,
        IssuedLGRecord.request_id != request_id,  # Exclude self
    ).all()
    
    results = []
    
    for lg in issued_lgs:
        score = 0.0
        breakdown = {}
        
        # Get the originating request for reference details
        linked_req = None
        if lg.request_id:
            linked_req = db.query(IssuanceRequest).filter(
                IssuanceRequest.id == lg.request_id
            ).first()
        
        # 1. Reference Type + Number (30%)
        if linked_req and request.reference_type and request.reference_number:
            if (linked_req.reference_type == request.reference_type and
                    linked_req.reference_number == request.reference_number):
                score += 30
                breakdown["reference"] = {"matched": True, "score": 30}
            else:
                breakdown["reference"] = {"matched": False, "score": 0}
        else:
            breakdown["reference"] = {"matched": False, "score": 0}
        
        # 2. Beneficiary Name (25%) — fuzzy
        if request.beneficiary_name and lg.beneficiary_name:
            ratio = SequenceMatcher(
                None, 
                request.beneficiary_name.lower().strip(),
                lg.beneficiary_name.lower().strip()
            ).ratio()
            if ratio >= 0.8:
                name_score = round(ratio * 25, 1)
                score += name_score
                breakdown["beneficiary"] = {"matched": True, "score": name_score, "similarity": round(ratio * 100)}
            else:
                breakdown["beneficiary"] = {"matched": False, "score": 0, "similarity": round(ratio * 100)}
        else:
            breakdown["beneficiary"] = {"matched": False, "score": 0}
        
        # 3. Amount (20%) — within ±5%
        if request.amount and lg.current_amount:
            req_amt = float(request.amount)
            lg_amt = float(lg.current_amount)
            if req_amt > 0 and lg_amt > 0:
                diff_pct = abs(req_amt - lg_amt) / max(req_amt, lg_amt)
                if diff_pct <= 0.05:
                    amount_score = round((1 - diff_pct / 0.05) * 20, 1)
                    score += amount_score
                    breakdown["amount"] = {"matched": True, "score": amount_score, "lg_amount": str(lg_amt)}
                else:
                    breakdown["amount"] = {"matched": False, "score": 0, "lg_amount": str(lg_amt)}
            else:
                breakdown["amount"] = {"matched": False, "score": 0}
        else:
            breakdown["amount"] = {"matched": False, "score": 0}
        
        # 4. LG Type (15%) — exact
        if linked_req and request.lg_type_id and linked_req.lg_type_id:
            if request.lg_type_id == linked_req.lg_type_id:
                score += 15
                breakdown["lg_type"] = {"matched": True, "score": 15}
            else:
                breakdown["lg_type"] = {"matched": False, "score": 0}
        else:
            breakdown["lg_type"] = {"matched": False, "score": 0}
        
        # 5. Expiry Date (10%) — within ±30 days
        if request.requested_expiry_date and lg.expiry_date:
            from datetime import date as date_type
            delta_days = abs((request.requested_expiry_date - lg.expiry_date).days)
            if delta_days <= 30:
                expiry_score = round((1 - delta_days / 30) * 10, 1)
                score += expiry_score
                breakdown["expiry"] = {"matched": True, "score": expiry_score, "days_diff": delta_days}
            else:
                breakdown["expiry"] = {"matched": False, "score": 0, "days_diff": delta_days}
        else:
            breakdown["expiry"] = {"matched": False, "score": 0}
        
        total_score = round(score, 1)
        if total_score >= 70:
            results.append({
                "lg_ref_number": lg.lg_ref_number,
                "lg_id": lg.id,
                "beneficiary_name": lg.beneficiary_name,
                "amount": str(lg.current_amount),
                "currency": lg.currency.iso_code if lg.currency else "",
                "issue_date": str(lg.issue_date) if lg.issue_date else None,
                "expiry_date": str(lg.expiry_date) if lg.expiry_date else None,
                "status": lg.status,
                "score": total_score,
                "breakdown": breakdown
            })
    
    results.sort(key=lambda x: x["score"], reverse=True)
    
    return {
        "request_id": request_id,
        "total_issued_compared": len(issued_lgs),
        "matches": results[:10]  # Top 10 at most
    }


# ==============================================================================
# EDIT NOTIFICATION HELPER
# ==============================================================================

def _send_edit_notifications(db, request, editor: TokenData, metadata: dict):
    """
    Sends FYI notifications after a post-submission edit.
    - Safe edits: log audit entry tagging requestor + prior approvers.
    - Risky edits (re-approval): same, plus a warning that approval chain was reset.
    Uses the existing audit log system which feeds the notification banner.
    """
    from app.crud.crud import log_action
    
    # Gather recipients: requestor + anyone who already approved
    notify_user_ids = set()
    if request.requestor_user_id:
        notify_user_ids.add(request.requestor_user_id)
    
    # Prior approvers from approval_chain_audit
    if request.approval_chain_audit:
        for step in request.approval_chain_audit:
            if step.get('user_id'):
                notify_user_ids.add(step['user_id'])
    
    # Don't notify the editor themselves
    notify_user_ids.discard(editor.user_id)
    
    if not notify_user_ids:
        return  # No one to notify
    
    re_approval = metadata.get('re_approval_triggered', False)
    changed_fields = metadata.get('risky_fields_changed', []) + metadata.get('safe_fields_changed', [])
    reason = metadata.get('change_reason', '')
    
    action_type = "EDIT_RE_APPROVAL_NOTICE" if re_approval else "EDIT_FYI_NOTICE"
    
    log_action(
        db,
        user_id=editor.user_id,
        action_type=action_type,
        entity_type="IssuanceRequest",
        entity_id=request.id,
        details={
            "serial_number": request.serial_number,
            "editor_role": editor.role if hasattr(editor, 'role') else "unknown",
            "changed_fields": changed_fields,
            "reason": reason,
            "re_approval_triggered": re_approval,
            "notify_user_ids": list(notify_user_ids),
        },
        customer_id=editor.customer_id
    )


# ==============================================================================
# REQUESTS MANAGEMENT (CORE)
# ==============================================================================

@router.post("/requests/", response_model=IssuanceRequestOut)
def create_issuance_request_internal(
    request_in: IssuanceRequestCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Internal: Creates a DRAFT request (all required fields must be present)."""
    return crud_issuance_request.create_request(
        db, 
        obj_in=request_in, 
        customer_id=current_user.customer_id, 
        user_id=current_user.user_id 
    )

@router.post("/requests/draft", response_model=IssuanceRequestOut)
def save_draft_request_internal(
    request_in: IssuanceRequestDraftCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """Internal: Saves an incomplete draft — all fields are optional."""
    return crud_issuance_request.create_request(
        db, 
        obj_in=request_in, 
        customer_id=current_user.customer_id, 
        user_id=current_user.user_id 
    )

@router.get("/requests/", response_model=List[IssuanceRequestOut])
def get_issuance_requests(
    skip: int = 0, limit: int = 100,
    db: Session = Depends(get_db), 
    current_user: TokenData = Depends(get_issuance_read_context)
):
    return crud_issuance_request.get_by_customer(db, customer_id=current_user.customer_id, skip=skip, limit=limit)

@router.get("/requests/{request_id}", response_model=IssuanceRequestOut)
def get_single_issuance_request(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """Get a single issuance request by ID."""
    return crud_issuance_request.get_single(db, request_id, current_user.customer_id)

@router.delete("/requests/{request_id}")
def delete_draft_request(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """Delete a DRAFT request. Only drafts can be deleted."""
    return crud_issuance_request.delete_draft(db, request_id, current_user.customer_id, current_user.user_id)

@router.put("/requests/{request_id}", response_model=IssuanceRequestOut)
def edit_issuance_request(
    request_id: int,
    request_in: IssuanceRequestUpdate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    Edit a request (corporate admin or end user).
    - If blacklisted fields changed on a post-submission request → re-approval triggered.
    - If only safe fields changed → notify requestor + approvers, no re-approval.
    - DRAFT edits have no governance restrictions.
    """
    updated = crud_issuance_request.update_request(
        db, request_id, request_in, current_user.customer_id, current_user.user_id
    )

    # Send notifications for post-submission edits
    metadata = getattr(updated, '_edit_metadata', None)
    if metadata and (metadata.get('re_approval_triggered') or metadata.get('safe_fields_changed')):
        try:
            _send_edit_notifications(db, updated, current_user, metadata)
        except Exception:
            import logging
            logging.getLogger(__name__).warning("Failed to send edit notifications", exc_info=True)

    return updated


# ==============================================================================
# GOVERNANCE: VERSIONS & SNAPSHOTS
# ==============================================================================

@router.get("/requests/{request_id}/snapshot")
def get_request_snapshot(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """Retrieves the Immutable V1 Submission Snapshot."""
    snapshot = db.query(IssuanceRequestSnapshot).join(IssuanceRequestSnapshot.request).filter(
        IssuanceRequestSnapshot.request_id == request_id,
        crud_issuance_request.model.customer_id == current_user.customer_id
    ).first()
    
    if not snapshot:
        raise HTTPException(404, "Snapshot not found. Has this request been submitted?")
    return snapshot.snapshot_data

@router.get("/requests/{request_id}/versions", response_model=List[IssuanceRequestVersionOut])
def get_request_versions(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """Retrieves linear edit history and diffs."""
    # Ensure ownership
    req = crud_issuance_request.get(db, id=request_id)
    if not req or req.customer_id != current_user.customer_id:
        raise HTTPException(404, "Request not found")
        
    versions = db.query(IssuanceRequestVersion).filter(
        IssuanceRequestVersion.request_id == request_id
    ).order_by(IssuanceRequestVersion.version_number.desc()).all()
    
    return versions

# ==============================================================================
# 4. INTELLIGENT DECISION SUPPORT
# ==============================================================================

@router.get("/requests/{request_id}/suitable-facilities", response_model=List[SuitableFacilityOut])
def get_suitable_facilities(
    request_id: int, 
    db: Session = Depends(get_db), 
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    Returns facilities with utilization, cost calculations, and recommendation tags.
    Delegates to the service's Smart Engine v2.
    """
    return issuance_service.get_suitable_facilities(db, request_id)

@router.post("/requests/{request_id}/reserve", response_model=IssuanceRequestOut)
def reserve_facility_for_request(
    request_id: int,
    sub_limit_id: int = Query(..., description="The sub-limit to reserve capacity on"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """Reserve facility capacity without issuing the LG. APPROVED_INTERNAL → FACILITY_RESERVED."""
    return issuance_service.reserve_facility(db, request_id, current_user.user_id, sub_limit_id)

@router.post("/requests/{request_id}/release-reservation", response_model=IssuanceRequestOut)
def release_facility_reservation(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """Release a facility reservation, freeing capacity. FACILITY_RESERVED → APPROVED_INTERNAL."""
    return issuance_service.release_reservation(db, request_id, current_user.user_id)

# C5: Pre-execution checks (FX drift warning)
@router.get("/requests/{request_id}/pre-execution-check")
def pre_execution_check(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """
    Runs pre-execution checks before issuing an LG.
    Returns warnings about FX rate drift since reservation.
    Frontend should call this before showing the execute dialog.
    """
    return issuance_service.pre_execution_check(db, request_id)

@router.post("/requests/{request_id}/issue", response_model=IssuedLGRecordDetailOut)
async def issue_lg(
    request_id: int,
    body: IssuanceExecuteRequest,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """
    Unified Issuance Endpoint.
    Validates approval, acquires atomic lock, checks facility,
    creates exposure entry, creates LG record, and transitions status.
    """
    new_lg = await issuance_service.issue_lg(
        db=db,
        request_id=request_id,
        user_id=current_user.user_id,
        sub_limit_id=body.sub_limit_id,
        issued_ref_number=body.issued_ref_number,
        issue_date=body.issue_date,
        expiry_date=body.expiry_date,
        issuance_method=body.issuance_method,
        bank_method_id=body.bank_method_id,
        bank_id=body.bank_id,
        manual_pricing=body.manual_pricing
    )
    return new_lg


@router.post("/requests/{request_id}/cancel", response_model=IssuanceRequestOut)
def cancel_issuance_request(
    request_id: int,
    body: IssuanceCancelRequest,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """
    Cancel a request before bank confirmation.
    Releases any facility reservation and unlocks the request.
    """
    return issuance_service.cancel_request(
        db=db,
        request_id=request_id,
        user_id=current_user.user_id,
        reason=body.reason
    )

# ==============================================================================
# 5. UTILITIES (PDF, SECURITY)
# ==============================================================================

@router.get("/requests/{request_id}/print-form")
async def print_issuance_application_form(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    request = crud_issuance_request.get(db, id=request_id)
    if not request: raise HTTPException(404, "Request not found")

    html_content = f"""
    <html><body>
        <h1>LG Issuance Application</h1>
        <p><strong>Ref:</strong> {request.id}</p>
        <p><strong>Beneficiary:</strong> {request.beneficiary_name}</p>
        <p><strong>Amount:</strong> {request.amount}</p>
    </body></html>
    """
    pdf_bytes = await generate_pdf_from_html(html_content, f"application_{request.id}")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=application_{request.id}.pdf"}
    )

@router.get("/requests/{request_id}/generate-letter")
async def generate_issuance_letter(
    request_id: int,
    additional_text: str = Query("", description="Extra free text instructions to include in the letter"),
    use_special_wording: bool = Query(False, description="Override to use special wording instead of bank standard"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    Generates a signed company letter PDF for an issuance request using the template system.
    Uses customer-specific template if available, otherwise falls back to the global default.
    Available placeholders are defined under action_type 'LG_ISSUANCE_REQUEST'.
    """
    result = await issuance_service.generate_issuance_letter(
        db=db,
        request_id=request_id,
        customer_id=current_user.customer_id,
        additional_text=additional_text,
        use_special_wording=use_special_wording,
    )
    headers = {
        'Content-Disposition': f'inline; filename="{result["filename"]}"'
    }
    return StreamingResponse(
        io.BytesIO(result["pdf_bytes"]),
        media_type="application/pdf",
        headers=headers,
    )

@router.get("/generate-portal-link")
def generate_portal_link(
    recipient_email: str = Query(...),
    department: str = Query(...),
    hours_valid: int = Query(24),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """
    Revised: Generates an encrypted link for a specific guest/employee.
    Format: customer_id|department|email|expiry_timestamp
    """
    from datetime import datetime, timedelta, timezone
    from app.crud.crud import log_action # Ensure this is imported

    # 1. Calculate Expiry
    expiry = (datetime.now(timezone.utc) + timedelta(hours=hours_valid)).isoformat()
    
    # 2. Build Payload (using | as a separator for clarity)
    payload = f"{current_user.customer_id}|{department}|{recipient_email}|{expiry}"
    
    # 3. Encrypt using your core encryption
    token = encrypt_data(payload)

    # 4. Record the action in the System Log
    log_action(
        db,
        user_id=current_user.user_id,
        action_type="EXTERNAL_INVITE_GENERATED",
        entity_type="IssuanceRequest",
        entity_id=None,
        details={"recipient": recipient_email, "dept": department, "expiry": expiry},
        customer_id=current_user.customer_id
    )

    return {"token": token, "link": f"/public/request?token={token}"}

# ==============================================================================
# 6. RECONCILIATION ENGINE (NEW)
# ==============================================================================

@router.post("/reconciliation/run", response_model=ReconciliationResult)
def run_bank_position_reconciliation(
    recon_data: ReconciliationRequest,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """
    Matches Excel rows against DB and SAVES a historical batch.
    """
    discrepancies = []
    matched = 0
    mismatched_amount = 0
    missing_system = 0
    
    # A. FETCH SYSTEM RECORDS
    bank_facilities = db.query(IssuanceFacility).filter(
        IssuanceFacility.bank_id == recon_data.bank_id, 
        IssuanceFacility.customer_id == current_user.customer_id
    ).all()
    
    fac_ids = [f.id for f in bank_facilities]
    
    system_records = db.query(IssuedLGRecord)\
        .join(IssuanceFacilitySubLimit)\
        .filter(IssuanceFacilitySubLimit.facility_id.in_(fac_ids))\
        .filter(IssuedLGRecord.status == 'ACTIVE')\
        .all()
    
    system_map = {rec.lg_ref_number: rec for rec in system_records}
    bank_refs_processed = set()

    # B. CREATE DB BATCH (History)
    # Using 'models_reconciliation.py' models
    batch = BankPositionBatch(
        bank_id=recon_data.bank_id,
        as_of_date=recon_data.as_of_date,
        uploaded_by_user_id=current_user.user_id
    )
    db.add(batch)
    db.flush() # Get ID
    
    # C. ITERATE & COMPARE
    for row in recon_data.rows:
        bank_refs_processed.add(row.ref_number)
        
        # Save raw row to DB
        db_row = BankPositionRow(
            batch_id=batch.id,
            ref_number=row.ref_number,
            amount=row.amount,
            currency_code=row.currency,
            status_in_bank=row.status
        )
        
        if row.ref_number in system_map:
            sys_rec = system_map[row.ref_number]
            if abs(float(sys_rec.current_amount) - row.amount) > 1.0:
                mismatched_amount += 1
                db_row.recon_status = "MISMATCH"
                db_row.recon_note = f"Amount mismatch: Bank={row.amount}, Sys={sys_rec.current_amount}"
                discrepancies.append({
                    "type": "AMOUNT_MISMATCH",
                    "ref": row.ref_number,
                    "bank_amount": row.amount,
                    "system_amount": float(sys_rec.current_amount),
                    "diff": row.amount - float(sys_rec.current_amount)
                })
            else:
                matched += 1
                db_row.recon_status = "MATCHED"
        else:
            missing_system += 1
            db_row.recon_status = "MISSING_IN_SYSTEM"
            db_row.recon_note = "Found in bank, missing in system"
            discrepancies.append({
                "type": "MISSING_IN_SYSTEM",
                "ref": row.ref_number,
                "bank_amount": row.amount,
                "note": "Bank has this LG, but we do not."
            })
            
        db.add(db_row)
            
    # D. REVERSE CHECK
    for sys_ref, sys_rec in system_map.items():
        if sys_ref not in bank_refs_processed:
            discrepancies.append({
                "type": "MISSING_IN_BANK",
                "ref": sys_ref,
                "system_amount": float(sys_rec.current_amount),
                "note": "We show this as ACTIVE, but it is not in Bank Position."
            })

    # E. COMMIT
    batch.total_records = len(recon_data.rows)
    batch.matched_records = matched
    db.commit()

    return {
        "total_bank_records": len(recon_data.rows),
        "matched_count": matched,
        "mismatched_amount_count": mismatched_amount,
        "missing_in_system_count": missing_system,
        "discrepancies": discrepancies
    }

# ==============================================================================
# 5. WORKFLOW CONFIGURATION (The Matrix Engine)
# ==============================================================================

def _detect_coverage_gaps(policies_in) -> list:
    """
    Analyzes a set of workflow policies for coverage gaps.
    Returns a list of warning strings for the admin.
    """
    from decimal import Decimal, InvalidOperation

    warnings = []
    has_always = False
    amount_ranges = []  # [(min, max)] — max=None means open-ended
    amount_over_thresholds = []

    for p in policies_in:
        ct = p.condition_type
        if ct in ("ALWAYS", "ANY_DEPARTMENT"):
            has_always = True
        elif ct == "AMOUNT_RANGE" and p.condition_value:
            try:
                raw = str(p.condition_value).strip().strip("()")
                if "," in raw:
                    parts = raw.split(",")
                else:
                    parts = raw.split("-")
                min_val = Decimal(parts[0].strip()) if parts[0].strip() else Decimal("0")
                max_val = Decimal(parts[1].strip()) if len(parts) > 1 and parts[1].strip() else None
                amount_ranges.append((min_val, max_val))
            except (InvalidOperation, IndexError):
                pass
        elif ct == "AMOUNT_OVER" and p.condition_value:
            try:
                threshold = Decimal(str(p.condition_value))
                amount_over_thresholds.append(threshold)
            except (InvalidOperation, ValueError):
                pass

    # Only analyze if there are amount-based policies but no catch-all
    if (amount_ranges or amount_over_thresholds) and not has_always:
        # Check if there's an open-ended upper range
        has_open_upper = any(max_val is None for _, max_val in amount_ranges)
        has_amount_over = len(amount_over_thresholds) > 0

        if amount_ranges and not has_open_upper and not has_amount_over:
            # Find the highest upper bound
            max_upper = max(
                (max_val for _, max_val in amount_ranges if max_val is not None),
                default=Decimal("0")
            )
            warnings.append(
                f"Coverage gap detected: Requests with amounts above {max_upper:,.0f} "
                f"do not match any approval rule and will be blocked. "
                f"Consider adding an 'Amount Greater Than {max_upper:,.0f}' rule "
                f"or an 'Always' rule to cover all amounts."
            )

        # Check for gaps between ranges (e.g., 0-50K and 100K-200K → gap at 50K-100K)
        if len(amount_ranges) > 1:
            sorted_ranges = sorted(amount_ranges, key=lambda x: x[0])
            for i in range(len(sorted_ranges) - 1):
                _, curr_max = sorted_ranges[i]
                next_min, _ = sorted_ranges[i + 1]
                if curr_max is not None and next_min > curr_max:
                    warnings.append(
                        f"Coverage gap detected: Requests with amounts between "
                        f"{curr_max:,.0f} and {next_min:,.0f} do not match any rule "
                        f"and will be blocked."
                    )

    # Edge case: policies exist but ALL are conditional (no ALWAYS rule)
    if policies_in and not has_always:
        conditional_only = all(
            p.condition_type not in ("ALWAYS", "ANY_DEPARTMENT")
            for p in policies_in
        )
        if conditional_only:
            warnings.append(
                "All approval rules have conditions. Requests that don't match "
                "any condition will be blocked. Consider adding an 'Always' rule "
                "as a catch-all to ensure all requests have an approval path."
            )

    return warnings

@router.get("/workflow-policies", response_model=List[IssuanceWorkflowPolicyOut])
def list_workflow_policies(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """ List all active approval rules for this customer in sequential order. """
    return db.query(IssuanceWorkflowPolicy).filter(
        IssuanceWorkflowPolicy.customer_id == current_user.customer_id
    ).order_by(IssuanceWorkflowPolicy.step_sequence.asc()).all()

@router.put("/workflow-policies")
def update_workflow_policies(
    policies_in: List[IssuanceWorkflowPolicyCreate],
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """ Bulk replace all approval rules. Wipes old rules and sets the new sequence.
    Multi-admin: requires dual-control approval via AdminChangeRequest.
    Returns saved policies + coverage gap warnings. """
    
    # 0. Validate: all amount-based steps must use the same currency
    amount_based_types = {"AMOUNT_OVER", "AMOUNT_RANGE"}
    amount_currencies = set()
    for p in policies_in:
        if p.condition_type in amount_based_types and p.currency_id:
            amount_currencies.add(p.currency_id)
    if len(amount_currencies) > 1:
        raise HTTPException(
            status_code=400,
            detail="All amount-based approval steps must use the same currency. "
                   "Please set the same currency for all Amount Over and Amount Range conditions."
        )

    # Serialize policies for the change payload
    new_val = [p.model_dump() for p in policies_in]

    change_req, auto_approved = _create_governed_change(
        db, current_user.customer_id, current_user.user_id,
        "APPROVAL_MATRIX_UPDATE", {"new_value": new_val}
    )

    if auto_approved:
        # _apply_admin_change already applied the policies — reload them
        new_policies = db.query(IssuanceWorkflowPolicy).filter(
            IssuanceWorkflowPolicy.customer_id == current_user.customer_id
        ).order_by(IssuanceWorkflowPolicy.step_sequence.asc()).all()
        warnings = _detect_coverage_gaps(policies_in)
        return {"policies": new_policies, "warnings": warnings}

    # Multi-admin: return 202 — change is pending
    from fastapi.responses import JSONResponse
    return JSONResponse(
        status_code=202,
        content={
            "message": "Approval matrix change submitted for approval by a second administrator.",
            "change_request_id": change_req.id,
            "status": "PENDING"
        }
    )

# ==============================================================================
# 6. APPROVAL ACTIONS
# ==============================================================================

@router.get("/requests/{request_id}/approval-roadmap")
def get_approval_roadmap(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """Returns the full approval lifecycle roadmap for a request."""
    return issuance_service.get_approval_roadmap(db, request_id)

@router.get("/my-pending-approvals", response_model=List[IssuanceRequestOut])
def get_my_pending_approvals(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context)
):
    """
    Returns issuance requests where the current user is a designated approver.
    Queries pending_approver_users JSONB field using proper containment.
    """
    from app.models.models_issuance import IssuanceRequest
    from sqlalchemy.dialects.postgresql import JSONB
    from sqlalchemy import cast, text
    
    # Use PostgreSQL JSONB @> operator to check if array contains user_id
    # This avoids substring false positives (e.g., user 2 matching "[42]")
    requests = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == current_user.customer_id,
        IssuanceRequest.status == "PENDING_APPROVAL",
        IssuanceRequest.is_deleted == False,
        IssuanceRequest.pending_approver_users.cast(JSONB).contains([current_user.user_id])
    ).order_by(IssuanceRequest.created_at.desc()).all()
    
    print(f"[DEBUG APPROVAL] my-pending-approvals: user_id={current_user.user_id}, found {len(requests)} requests: {[(r.id, r.pending_approver_users) for r in requests]}")
    
    return requests


@router.get("/my-approval-history", response_model=List[IssuanceRequestOut])
def get_my_approval_history(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context)
):
    """
    Returns all non-draft issuance requests for the customer.
    pending_approver_users gets cleared after approval/rejection,
    so we show all requests that have ever entered the approval flow.
    """
    from app.models.models_issuance import IssuanceRequest

    requests = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == current_user.customer_id,
        IssuanceRequest.is_deleted == False,
        IssuanceRequest.status.notin_(["DRAFT"]),
    ).order_by(IssuanceRequest.created_at.desc()).all()

    return requests

@router.post("/requests/{request_id}/submit", response_model=IssuanceRequestOut)
async def submit_request_for_approval(
    request_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """ User submits DRAFT -> PENDING_APPROVAL """
    result = issuance_service.submit_for_approval(db, request_id, current_user.user_id)
    
    print(f"[DEBUG EMAIL] internal_submit: status={result.status}, approvers={result.pending_approver_users}")
    
    # --- Notification (matching corporate_admin.py create_user pattern exactly) ---
    if result.status == "PENDING_APPROVAL" and result.pending_approver_users:
        from app.core.email_service import send_email, get_global_email_settings
        from app.services.issuance_notifications import _get_user_emails
        from app.models import User
        import os
        
        email_settings = get_global_email_settings()
        submitter = db.query(User).filter(User.id == current_user.user_id).first()
        currency = result.currency.iso_code if result.currency else "N/A"
        approver_ids = [int(uid) for uid in result.pending_approver_users]
        approver_emails = _get_user_emails(db, approver_ids)
        frontend_url = os.getenv("FRONTEND_URL", "http://localhost:3000")
        
        print(f"[DEBUG EMAIL] internal_submit: approver_ids={approver_ids}, emails={approver_emails}, host={email_settings.smtp_host}")
        
        if approver_emails:
            submitter_email = submitter.email if submitter else result.requestor_email
            subject = f"ACTION REQUIRED: LG Request {result.serial_number} Awaiting Approval"
            body = f"""
            <html>
            <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
                <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                    <h2 style="color: #1a56db; margin-top: 0;">🔔 New Approval Request</h2>
                    <p>A new LG issuance request has been submitted and requires your approval.</p>
                    <div style="background: #f8fafc; border-left: 4px solid #1a56db; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr><td style="padding: 4px 0; color: #666;">Reference:</td><td style="padding: 4px 0; font-weight: bold;">{result.serial_number}</td></tr>
                            <tr><td style="padding: 4px 0; color: #666;">Amount:</td><td style="padding: 4px 0; font-weight: bold;">{currency} {result.amount}</td></tr>
                            <tr><td style="padding: 4px 0; color: #666;">Beneficiary:</td><td style="padding: 4px 0;">{result.beneficiary_name}</td></tr>
                            <tr><td style="padding: 4px 0; color: #666;">Submitted by:</td><td style="padding: 4px 0;">{submitter_email or 'External Requestor'}</td></tr>
                        </table>
                    </div>
                    <div style="text-align: center; margin: 25px 0;">
                        <a href="{frontend_url}/corporate-admin/approval-inbox" style="padding: 12px 30px; background: #1a56db; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">Review Request</a>
                    </div>
                    <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
                    <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
                </div>
            </body>
            </html>
            """
            background_tasks.add_task(
                send_email,
                db,
                approver_emails,
                subject,
                body,
                {},
                email_settings,
            )
            print(f"[DEBUG EMAIL] internal_submit: background_task added for {approver_emails}")
    
    # --- In-App Notification (additive — does not replace email) ---
    if result.status == "PENDING_APPROVAL" and result.pending_approver_users:
        from app.services.notification_service import notify_request_submitted
        from app.models import User
        submitter = db.query(User).filter(User.id == current_user.user_id).first()
        submitter_name = submitter.email if submitter else "External Requestor"
        approver_ids = [int(uid) for uid in result.pending_approver_users]
        notify_request_submitted(
            db, approver_ids, result.serial_number,
            submitter_name, current_user.user_id, request_id
        )
    
    return result

@router.post("/requests/{request_id}/approve", response_model=IssuanceRequestOut)
def approve_request_action(
    request_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context)
):
    """ 
    Approver (Manager) approves the request. 
    Moves to Next Step OR 'APPROVED_INTERNAL'.
    """
    result = issuance_service.approve_request(db, request_id, current_user.user_id)
    currency = result.currency.iso_code if result.currency else "N/A"
    
    # --- Resolve email data NOW while DB session is active (not in background task) ---
    from app.core.email_service import send_email, get_customer_email_settings
    from app.services.issuance_notifications import _get_user_email, _get_user_emails, _base_url
    
    email_settings, _ = get_customer_email_settings(db, result.customer_id)
    
    # --- Notification 1: Tell submitter their request was approved ---
    to_emails = []
    submitter_email = _get_user_email(db, result.requestor_user_id)
    if submitter_email:
        to_emails.append(submitter_email)
    if result.requestor_email and result.requestor_email not in to_emails:
        to_emails.append(result.requestor_email)
    
    if to_emails:
        is_final = result.status == "APPROVED_INTERNAL"
        status_label = "Fully Approved ✅" if is_final else "Step Approved — Proceeding to Next Approver"
        link = f"{_base_url()}/corporate-admin/issuance/requests"
        subject = f"LG Request {result.serial_number} — {status_label}"
        body = f"""
        <html>
        <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #16a34a; margin-top: 0;">✅ Request Approved</h2>
                <p>Your LG issuance request has been <strong>{status_label.lower()}</strong>.</p>
                <div style="background: #f0fdf4; border-left: 4px solid #16a34a; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr><td style="padding: 4px 0; color: #666;">Reference:</td><td style="padding: 4px 0; font-weight: bold;">{result.serial_number}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Amount:</td><td style="padding: 4px 0; font-weight: bold;">{currency} {result.amount}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Beneficiary:</td><td style="padding: 4px 0;">{result.beneficiary_name}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Status:</td><td style="padding: 4px 0; font-weight: bold; color: #16a34a;">{result.status}</td></tr>
                    </table>
                </div>
                {"<p>Your request is now ready for issuance execution.</p>" if is_final else "<p>The request is moving to the next approval step.</p>"}
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{link}" style="padding: 12px 30px; background: #16a34a; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">View Status</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
                <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
            </div>
        </body>
        </html>
        """
        background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings)
    
    # --- Notification 2: If still pending, notify next set of approvers ---
    if result.status == "PENDING_APPROVAL" and result.pending_approver_users:
        approver_ids = [int(uid) for uid in result.pending_approver_users]
        approver_emails = _get_user_emails(db, approver_ids)
        if approver_emails:
            link2 = f"{_base_url()}/corporate-admin/approval-inbox"
            subject2 = f"ACTION REQUIRED: LG Issuance Request {result.serial_number} Awaiting Your Approval"
            body2 = f"""
            <html>
            <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
                <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                    <h2 style="color: #1a56db; margin-top: 0;">🔔 New Approval Request</h2>
                    <p>An LG issuance request requires your approval.</p>
                    <div style="background: #f8fafc; border-left: 4px solid #1a56db; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr><td style="padding: 4px 0; color: #666;">Reference:</td><td style="padding: 4px 0; font-weight: bold;">{result.serial_number}</td></tr>
                            <tr><td style="padding: 4px 0; color: #666;">Amount:</td><td style="padding: 4px 0; font-weight: bold;">{currency} {result.amount}</td></tr>
                            <tr><td style="padding: 4px 0; color: #666;">Beneficiary:</td><td style="padding: 4px 0;">{result.beneficiary_name}</td></tr>
                        </table>
                    </div>
                    <div style="text-align: center; margin: 25px 0;">
                        <a href="{link2}" style="padding: 12px 30px; background: #1a56db; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">Review Request</a>
                    </div>
                    <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
                    <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
                </div>
            </body>
            </html>
            """
            background_tasks.add_task(send_email, db, approver_emails, subject2, body2, {}, email_settings)
    
    # --- In-App Notification (additive — does not replace email) ---
    from app.services.notification_service import notify_request_approved
    from app.models import User
    is_final = result.status == "APPROVED_INTERNAL"
    approver = db.query(User).filter(User.id == current_user.user_id).first()
    approver_name = approver.email if approver else "Approver"
    
    # Notify the submitter
    notify_recipients = []
    if result.requestor_user_id:
        notify_recipients.append(result.requestor_user_id)
    # If still pending, also notify next approvers
    if not is_final and result.pending_approver_users:
        notify_recipients.extend([int(uid) for uid in result.pending_approver_users])
    
    notify_request_approved(
        db, notify_recipients, result.serial_number,
        step_number=0, approver_name=approver_name,
        actor_user_id=current_user.user_id, request_id=request_id,
        is_fully_approved=is_final
    )
    
    # --- Notify external requestor ---
    if result.requestor_email:
        event = "APPROVED_INTERNAL" if is_final else "APPROVED_STEP"
        _send_requestor_status_notification(db, background_tasks, result, event)
    
    return result

@router.post("/requests/{request_id}/reject", response_model=IssuanceRequestOut)
def reject_request_action(
    request_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context)
):
    """ Approver rejects the request. """
    result = issuance_service.reject_request(db, request_id, current_user.user_id)
    
    # --- Resolve email data NOW while DB session is active ---
    from app.core.email_service import send_email, get_customer_email_settings
    from app.services.issuance_notifications import _get_user_email, _base_url
    from app.models import User
    
    email_settings, _ = get_customer_email_settings(db, result.customer_id)
    rejector = db.query(User).filter(User.id == current_user.user_id).first()
    rejector_email = rejector.email if rejector else "Admin"
    currency = result.currency.iso_code if result.currency else "N/A"
    
    to_emails = []
    submitter_email = _get_user_email(db, result.requestor_user_id)
    if submitter_email:
        to_emails.append(submitter_email)
    if result.requestor_email and result.requestor_email not in to_emails:
        to_emails.append(result.requestor_email)
    
    if to_emails:
        link = f"{_base_url()}/corporate-admin/issuance/requests"
        subject = f"LG Request {result.serial_number} — Rejected ❌"
        body = f"""
        <html>
        <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #dc2626; margin-top: 0;">❌ Request Rejected</h2>
                <p>Your LG issuance request has been rejected by an approver.</p>
                <div style="background: #fef2f2; border-left: 4px solid #dc2626; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr><td style="padding: 4px 0; color: #666;">Reference:</td><td style="padding: 4px 0; font-weight: bold;">{result.serial_number}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Amount:</td><td style="padding: 4px 0; font-weight: bold;">{currency} {result.amount}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Beneficiary:</td><td style="padding: 4px 0;">{result.beneficiary_name}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Rejected by:</td><td style="padding: 4px 0;">{rejector_email}</td></tr>
                    </table>
                </div>
                <p>Please review the request details and contact the approver if clarification is needed.</p>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{link}" style="padding: 12px 30px; background: #dc2626; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">View Request</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
                <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
            </div>
        </body>
        </html>
        """
        background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings)
    
    # --- In-App Notification (additive — does not replace email) ---
    from app.services.notification_service import notify_request_rejected
    notify_recipients = []
    if result.requestor_user_id:
        notify_recipients.append(result.requestor_user_id)
    
    notify_request_rejected(
        db, notify_recipients, result.serial_number,
        rejector_name=rejector_email, reason="See request details",
        actor_user_id=current_user.user_id, request_id=request_id
    )
    
    # --- Notify external requestor ---
    if result.requestor_email:
        _send_requestor_status_notification(db, background_tasks, result, "REQUEST_REJECTED")
    
    return result

class ReturnForRevisionPayload(BaseModel):
    revision_notes: Optional[str] = None

@router.post("/requests/{request_id}/return-for-revision", response_model=IssuanceRequestOut)
def return_for_revision_action(
    request_id: int,
    payload: ReturnForRevisionPayload,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context)
):
    """
    Approver returns the request for revision instead of rejecting.
    Requestor can edit and re-submit — approval resumes from the same step.
    """
    result = issuance_service.return_for_revision(
        db, request_id, current_user.user_id, payload.revision_notes
    )
    
    # --- Email notification to requestor ---
    from app.core.email_service import send_email, get_customer_email_settings
    from app.services.issuance_notifications import _get_user_email, _base_url
    from app.models import User
    
    email_settings, _ = get_customer_email_settings(db, result.customer_id)
    returner = db.query(User).filter(User.id == current_user.user_id).first()
    returner_email = returner.email if returner else "Approver"
    currency = result.currency.iso_code if result.currency else "N/A"
    
    to_emails = []
    submitter_email = _get_user_email(db, result.requestor_user_id)
    if submitter_email:
        to_emails.append(submitter_email)
    if result.requestor_email and result.requestor_email not in to_emails:
        to_emails.append(result.requestor_email)
    
    if to_emails:
        link = f"{_base_url()}/corporate-admin/issuance/requests"
        notes_html = f"<p><strong>Revision Notes:</strong> {payload.revision_notes}</p>" if payload.revision_notes else ""
        subject = f"LG Request {result.serial_number} — Returned for Revision 🔄"
        body = f"""
        <html>
        <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #f59e0b; margin-top: 0;">🔄 Returned for Revision</h2>
                <p>Your LG issuance request has been returned for revision by an approver. Please review the notes, make corrections, and re-submit.</p>
                <div style="background: #fffbeb; border-left: 4px solid #f59e0b; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr><td style="padding: 4px 0; color: #666;">Reference:</td><td style="padding: 4px 0; font-weight: bold;">{result.serial_number}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Amount:</td><td style="padding: 4px 0; font-weight: bold;">{currency} {result.amount}</td></tr>
                        <tr><td style="padding: 4px 0; color: #666;">Returned by:</td><td style="padding: 4px 0;">{returner_email}</td></tr>
                    </table>
                </div>
                {notes_html}
                <p>Once you have made the required changes, please re-submit the request. <strong>Approval will resume from the step that returned it.</strong></p>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{link}" style="padding: 12px 30px; background: #f59e0b; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">Edit & Resubmit</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
                <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
            </div>
        </body>
        </html>
        """
        background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings)
    
    # --- Notify external requestor ---
    if result.requestor_email:
        _send_requestor_status_notification(db, background_tasks, result, "RETURNED_FOR_REVISION")
    
    return result

from app.schemas.schemas_issuance import SuitableFacilityOut

@router.get("/requests/{request_id}/recommendations", response_model=List[SuitableFacilityOut])
def get_facility_recommendations(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """
    Returns a list of facilities that CAN issue this LG, sorted by SLA.
    """
    return issuance_service.get_suitable_facilities(db, request_id)


# ==============================================================================
# C3. TREASURY ENRICHMENT (Technical Input by Treasury)
# ==============================================================================

@router.patch("/requests/{request_id}/enrich")
def enrich_request(
    request_id: int,
    enrichment_data: Dict[str, Any] = Body(..., description="Treasury enrichment fields to add/update"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """
    Treasury team adds technical details to a request (applicable rules,
    advising bank, margin instructions, internal notes, etc.).
    Available at any time after submission — before, during, or after approval.
    Merges into the existing treasury_enrichment JSONB field.
    """
    from app.models.models_issuance import IssuanceRequest
    from datetime import datetime as dt

    request = db.query(IssuanceRequest).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id,
        IssuanceRequest.is_deleted == False,
    ).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")

    if request.status == "DRAFT":
        raise HTTPException(status_code=400, detail="Cannot enrich a draft — request must be submitted first")

    # Merge enrichment data
    existing = request.treasury_enrichment or {}
    existing.update(enrichment_data)
    existing["enriched_by_user_id"] = current_user.user_id
    existing["enriched_at"] = dt.utcnow().isoformat()

    # Use flag_modified to ensure JSONB change is detected
    from sqlalchemy.orm.attributes import flag_modified
    request.treasury_enrichment = existing
    flag_modified(request, "treasury_enrichment")

    # Also update applicable_rules on the request if provided in enrichment
    if "applicable_rules" in enrichment_data:
        request.applicable_rules = enrichment_data["applicable_rules"]

    # Update cross_border_details if provided
    if "cross_border_details" in enrichment_data and isinstance(enrichment_data["cross_border_details"], dict):
        cbd = request.cross_border_details or {}
        cbd.update(enrichment_data["cross_border_details"])
        request.cross_border_details = cbd
        flag_modified(request, "cross_border_details")

    db.flush()
    db.refresh(request)

    return {
        "message": "Request enriched successfully",
        "request_id": request.id,
        "treasury_enrichment": request.treasury_enrichment,
        "applicable_rules": request.applicable_rules,
        "cross_border_details": request.cross_border_details,
    }


# ==============================================================================
# 8. BENEFICIARY LOOKUP (Smart Auto-Fill)
# ==============================================================================

@router.get("/beneficiary-lookup")
def beneficiary_lookup(
    id_number: str = Query(..., description="Beneficiary ID/Number to look up"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """
    Returns the most recent beneficiary data matching the given ID/number.
    Searches across all issuance requests for this customer.
    """
    from app.models.models_issuance import IssuanceRequest
    
    request = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == current_user.customer_id,
        IssuanceRequest.beneficiary_id_number == id_number,
        IssuanceRequest.is_deleted == False
    ).order_by(IssuanceRequest.created_at.desc()).first()
    
    if not request:
        return {"found": False}
    
    return {
        "found": True,
        "beneficiary_name": request.beneficiary_name,
        "beneficiary_country": request.beneficiary_country,
        "beneficiary_address": request.beneficiary_address,
        "beneficiary_contact_person": request.beneficiary_contact_person,
        "beneficiary_phone": request.beneficiary_phone,
        "beneficiary_email": request.beneficiary_email,
    }


@router.get("/beneficiary-suggest")
def beneficiary_suggest(
    name: str = Query(..., min_length=3, description="Partial beneficiary name to search"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """
    Fuzzy match: returns previous beneficiary names similar to the input.
    Searches across all issuance requests for this customer.
    """
    from app.models.models_issuance import IssuanceRequest
    from sqlalchemy import func as sa_func
    
    # Use ILIKE for case-insensitive partial match
    matches = db.query(
        IssuanceRequest.beneficiary_name,
        IssuanceRequest.beneficiary_id_number,
        IssuanceRequest.beneficiary_country,
        IssuanceRequest.beneficiary_address,
        IssuanceRequest.beneficiary_contact_person,
        IssuanceRequest.beneficiary_phone,
        IssuanceRequest.beneficiary_email,
        sa_func.max(IssuanceRequest.created_at).label('latest')
    ).filter(
        IssuanceRequest.customer_id == current_user.customer_id,
        IssuanceRequest.beneficiary_name.ilike(f"%{name}%"),
        IssuanceRequest.is_deleted == False
    ).group_by(
        IssuanceRequest.beneficiary_name,
        IssuanceRequest.beneficiary_id_number,
        IssuanceRequest.beneficiary_country,
        IssuanceRequest.beneficiary_address,
        IssuanceRequest.beneficiary_contact_person,
        IssuanceRequest.beneficiary_phone,
        IssuanceRequest.beneficiary_email,
    ).order_by(sa_func.max(IssuanceRequest.created_at).desc()).limit(5).all()
    
    return [
        {
            "beneficiary_name": m.beneficiary_name,
            "beneficiary_id_number": m.beneficiary_id_number,
            "beneficiary_country": m.beneficiary_country,
            "beneficiary_address": m.beneficiary_address,
            "beneficiary_contact_person": m.beneficiary_contact_person,
            "beneficiary_phone": m.beneficiary_phone,
            "beneficiary_email": m.beneficiary_email,
        }
        for m in matches
    ]


@router.get("/beneficiary-nearmatch")
def beneficiary_nearmatch(
    name: str = Query(..., min_length=2, description="Beneficiary name to check for near matches"),
    threshold: float = Query(0.85, ge=0.5, le=1.0, description="Minimum similarity ratio (0.85 = 85%)"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """
    Checks if a beneficiary name closely matches any existing beneficiary
    in the customer's history. Returns matches ≥ threshold similarity.
    Used for the side-note warning in the approval/execution modal.
    """
    from app.models.models_issuance import IssuanceRequest
    from difflib import SequenceMatcher

    # Get unique beneficiary names for this customer
    beneficiaries = db.query(
        IssuanceRequest.beneficiary_name,
        IssuanceRequest.serial_number,
        IssuanceRequest.beneficiary_id_number,
    ).filter(
        IssuanceRequest.customer_id == current_user.customer_id,
        IssuanceRequest.beneficiary_name != None,
        IssuanceRequest.is_deleted == False,
        IssuanceRequest.status != "DRAFT",
    ).order_by(IssuanceRequest.created_at.desc()).all()

    # Deduplicate by name (keep most recent)
    seen = {}
    for b in beneficiaries:
        normalized = b.beneficiary_name.strip().lower()
        if normalized not in seen:
            seen[normalized] = b

    # Compare
    input_normalized = name.strip().lower()
    near_matches = []
    for normalized_name, record in seen.items():
        if normalized_name == input_normalized:
            continue  # Skip exact match
        ratio = SequenceMatcher(None, input_normalized, normalized_name).ratio()
        if ratio >= threshold:
            near_matches.append({
                "beneficiary_name": record.beneficiary_name,
                "beneficiary_id_number": record.beneficiary_id_number,
                "similarity": round(ratio * 100, 1),
                "last_seen_request": record.serial_number,
            })

    # Sort by similarity descending
    near_matches.sort(key=lambda x: x["similarity"], reverse=True)
    return near_matches[:5]


# ==============================================================================
# 9. DOCUMENT MANAGEMENT (Issuance Request Documents)
# ==============================================================================

from fastapi import UploadFile, File

@router.post("/requests/{request_id}/documents")
async def upload_request_document(
    request_id: int,
    document_type: str = Query(..., description="CONTRACT, PURCHASE_ORDER, THIRD_PARTY, SPECIAL_WORDING, OTHER"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """Upload a document to an issuance request. Uses customer-specific GCS bucket."""
    from app.models.models_issuance import IssuanceRequest, IssuanceRequestDocument
    from app.core.ai_integration import _upload_to_gcs, GCS_BUCKET_NAME
    from app.crud.crud_lg_document import _slugify_doc_type
    import uuid
    from datetime import datetime as dt
    
    request = db.query(IssuanceRequest).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id
    ).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Read file
    file_content = await file.read()
    file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'bin'
    
    # Generate unique filename
    unique_filename = f"REQ-{request_id}_{document_type}_{dt.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}.{file_extension}"
    
    # Construct GCS path
    doc_type_slug = _slugify_doc_type(document_type)
    blob_path = f"customer_{current_user.customer_id}/issuance_req_{request_id}/{doc_type_slug}/{unique_filename}"
    
    # Get customer-specific bucket or fallback
    from app.crud import crud_customer_configuration
    bucket_name = GCS_BUCKET_NAME
    bucket_config = crud_customer_configuration.get_customer_config_or_global_fallback(
        db, current_user.customer_id, "STORAGE_BUCKET_NAME"
    )
    if bucket_config and bucket_config.get('effective_value'):
        bucket_name = bucket_config['effective_value']
    
    # Upload
    stored_uri = await _upload_to_gcs(bucket_name, blob_path, file_content, file.content_type)
    if not stored_uri:
        raise HTTPException(status_code=500, detail="Failed to upload document")
    
    # Save metadata
    doc = IssuanceRequestDocument(
        request_id=request_id,
        document_type=document_type,
        file_name=file.filename,
        file_path=stored_uri,
        uploaded_by=current_user.user_id
    )
    db.add(doc)
    db.flush()
    db.refresh(doc)
    
    return {
        "id": doc.id,
        "document_type": doc.document_type,
        "file_name": doc.file_name,
        "created_at": str(doc.created_at) if doc.created_at else None
    }


@router.get("/requests/{request_id}/documents")
def list_request_documents(
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """List all documents attached to an issuance request."""
    from app.models.models_issuance import IssuanceRequestDocument
    
    docs = db.query(IssuanceRequestDocument).filter(
        IssuanceRequestDocument.request_id == request_id,
        IssuanceRequestDocument.is_deleted == False
    ).all()
    
    return [
        {
            "id": d.id,
            "document_type": d.document_type,
            "file_name": d.file_name,
            "created_at": str(d.created_at) if d.created_at else None,
            "ai_verification_result": d.ai_verification_result
        }
        for d in docs
    ]


@router.get("/requests/{request_id}/documents/{document_id}/download")
async def download_request_document(
    request_id: int,
    document_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """Returns a tokenized signed URL for downloading the document."""
    from app.models.models_issuance import IssuanceRequestDocument
    from app.core.ai_integration import generate_signed_gcs_url
    
    doc = db.query(IssuanceRequestDocument).filter(
        IssuanceRequestDocument.id == document_id,
        IssuanceRequestDocument.request_id == request_id,
        IssuanceRequestDocument.is_deleted == False
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Generate signed URL with limited validity (1 hour)
    signed_url = await generate_signed_gcs_url(doc.file_path, expiration=3600)
    if not signed_url:
        raise HTTPException(status_code=500, detail="Could not generate download link. File may not be accessible.")
    
    return {
        "file_name": doc.file_name,
        "document_type": doc.document_type,
        "download_url": signed_url
    }


# ==============================================================================
# ISSUED LGs — List, Reprint, Bank Options
# ==============================================================================


def _get_lg_copy_docs(db, request_id):
    """Get BANK_LG_COPY documents for a given request."""
    from app.models.models_issuance import IssuanceRequestDocument
    docs = db.query(IssuanceRequestDocument).filter(
        IssuanceRequestDocument.request_id == request_id,
        IssuanceRequestDocument.document_type == "BANK_LG_COPY"
    ).order_by(IssuanceRequestDocument.created_at.desc()).limit(5).all()
    return [{
        "id": d.id,
        "file_name": d.file_name,
        "file_path": d.file_path,
        "created_at": d.created_at.isoformat() if d.created_at else None,
    } for d in docs]


@router.get("/issued-lgs")
def list_issued_lgs(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """List all Issued LG records for the customer, with comprehensive details."""
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.models import Bank, Currency, User

    records = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == current_user.customer_id
    ).order_by(IssuedLGRecord.created_at.desc()).all()

    result = []
    for r in records:
        # Resolve bank: direct relationship first, then via facility chain
        bank_name = "N/A"
        resolved_bank_id = None
        facility_name = None
        facility_ref = None
        sub_limit_name = None
        
        # Priority 1: Direct bank_id on the record
        if r.bank_id and r.bank:
            bank_name = r.bank.name
            resolved_bank_id = r.bank_id
        
        # Priority 2: Through sub_limit → facility → bank
        if r.sub_limit and r.sub_limit.facility:
            f = r.sub_limit.facility
            if resolved_bank_id is None and f.bank:
                bank_name = f.bank.name
                resolved_bank_id = f.bank_id
            facility_name = f.facility_name
            facility_ref = f.reference_number
            sub_limit_name = r.sub_limit.limit_name

        currency_code = r.currency.iso_code if r.currency else "N/A"

        # Get linked request details
        request_data = None
        if r.request_id:
            req = db.query(IssuanceRequest).filter(IssuanceRequest.id == r.request_id).first()
            if req:
                request_data = {
                    "id": req.id,
                    "serial_number": req.serial_number,
                    "requestor_name": req.requestor_name,
                    "requestor_email": req.requestor_email,
                    "lg_type": req.lg_type.name if req.lg_type else None,
                    "lg_purpose": req.lg_purpose,
                    "beneficiary_name": req.beneficiary_name,
                    "beneficiary_address": req.beneficiary_address,
                    "reference_type": req.reference_type,
                    "reference_number": req.reference_number,
                    "project_name": req.project.name if req.project else None,
                    "department": req.department,
                    "amount": float(req.amount) if req.amount else 0,
                    "status": req.status,
                    "requested_expiry_date": str(req.requested_expiry_date) if req.requested_expiry_date else None,
                    "requires_special_wording": req.requires_special_wording,
                    "other_conditions": req.other_conditions,
                    "is_cross_border": req.is_cross_border or False,
                    "is_third_party": req.is_third_party or False,
                    "submitted_at": req.submitted_at.isoformat() if getattr(req, 'submitted_at', None) else None,
                    "submitted_by_user_id": getattr(req, 'submitted_by_user_id', None),
                    "applicable_rules": req.applicable_rules,
                    "cross_border_details": req.cross_border_details,
                    "treasury_enrichment": req.treasury_enrichment,
                    "reference_end_date": str(req.reference_end_date) if req.reference_end_date else None,
                    "created_at": req.created_at.isoformat() if req.created_at else None,
                    "approval_chain_audit": req.approval_chain_audit or [],
                }

        # Issued-by user
        issued_by_name = None
        if r.issued_by_user_id:
            user = db.query(User).filter(User.id == r.issued_by_user_id).first()
            if user:
                issued_by_name = f"{user.first_name} {user.last_name}" if hasattr(user, 'first_name') else user.email

        # Current owner name
        current_owner_name = None
        if getattr(r, 'current_owner_user_id', None):
            owner = db.query(User).filter(User.id == r.current_owner_user_id).first()
            if owner:
                current_owner_name = owner.email

        result.append({
            "id": r.id,
            "lg_ref_number": r.lg_ref_number,
            "internal_serial": r.internal_serial,
            "beneficiary_name": r.beneficiary_name,
            "current_amount": float(r.current_amount),
            "currency_code": currency_code,
            "currency_id": r.currency_id,
            "issue_date": str(r.issue_date) if r.issue_date else None,
            "expiry_date": str(r.expiry_date) if r.expiry_date else None,
            "status": r.status,
            "issuance_method": r.issuance_method,
            # Bank & Facility
            "bank_name": bank_name,
            "bank_id": resolved_bank_id,
            "facility_name": facility_name,
            "facility_ref": facility_ref,
            "sub_limit_name": sub_limit_name,
            # Bank Confirmation
            "bank_confirmation_ref": r.bank_confirmation_ref,
            "bank_confirmation_date": str(r.bank_confirmation_date) if r.bank_confirmation_date else None,
            # Delivery Tracking
            "delivery_date": str(r.delivery_date) if r.delivery_date else None,
            "delivery_method": r.delivery_method,
            "delivery_notes": r.delivery_notes,
            # Bank Reply Tracking
            "bank_reply_type": r.bank_reply_type,
            "bank_reply_date": str(r.bank_reply_date) if r.bank_reply_date else None,
            "bank_reply_notes": r.bank_reply_notes,
            "bank_lg_number": r.bank_lg_number,
            # Verification
            "verification_status": r.verification_status,
            "verification_notes": r.verification_notes,
            "verified_at": r.verified_at.isoformat() if r.verified_at else None,
            "verified_by_user_id": r.verified_by_user_id,
            # Handover
            "handover_date": str(r.handover_date) if r.handover_date else None,
            "handover_notes": r.handover_notes,
            "handover_by_user_id": r.handover_by_user_id,
            "recipient_name": r.recipient_name,
            # Custody
            "original_copy_collected_by": r.original_copy_collected_by,
            "original_copy_collected_date": str(r.original_copy_collected_date) if r.original_copy_collected_date else None,
            "soft_copy_path": r.soft_copy_path,
            "custody_holder": r.custody_holder,
            "custody_transfer_log": r.custody_transfer_log or [],
            "action_history": r.action_history or [],
            # Accountability
            "issued_by_user_id": r.issued_by_user_id,
            "issued_by_name": issued_by_name,
            # Phase A new fields
            "reference_validity_flag": getattr(r, 'reference_validity_flag', None),
            "current_owner_user_id": getattr(r, 'current_owner_user_id', None),
            "current_owner_name": current_owner_name,
            # Bank LG fields for comparison
            "bank_lg_amount": float(r.bank_lg_amount) if r.bank_lg_amount else None,
            "bank_lg_issue_date": str(r.bank_lg_issue_date) if r.bank_lg_issue_date else None,
            "bank_lg_expiry_date": str(r.bank_lg_expiry_date) if r.bank_lg_expiry_date else None,
            # LG Copy Documents (for admin review)
            "lg_copy_documents": _get_lg_copy_docs(db, r.request_id) if r.request_id else [],
            # Linked Request
            "request": request_data,
            # Timestamps
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        })

    return result


@router.get("/issued-lgs/export")
def export_issued_lgs(
    export_type: str = Query("summary", description="summary | detailed | full_audit"),
    status_filter: Optional[str] = Query(None, description="Filter by status, e.g. ACTIVE"),
    search: Optional[str] = Query(None, description="Search LG ref or beneficiary"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """Export Issued LGs to XLSX. Types: summary, detailed, full_audit."""
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.models import Bank, Currency, User
    from fastapi.responses import StreamingResponse
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed. Run: pip install openpyxl")

    query = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == current_user.customer_id
    )
    if status_filter:
        query = query.filter(IssuedLGRecord.status == status_filter)
    records = query.order_by(IssuedLGRecord.created_at.desc()).all()

    # Optional search filter (in-memory for simplicity)
    if search:
        s = search.lower()
        records = [r for r in records if
                   (r.lg_ref_number and s in r.lg_ref_number.lower()) or
                   (r.beneficiary_name and s in r.beneficiary_name.lower())]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issued LGs"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc'),
    )

    # Define columns based on export type
    # Status label mapping for export
    status_labels = {
        "PENDING_CONFIRMATION": "Instruction Issued",
        "DELIVERED_TO_BANK": "At Bank",
        "BANK_INQUIRY": "Bank Inquiry",
        "BANK_REJECTED": "Rejected by Bank",
        "PENDING_VERIFICATION": "Needs Review",
        "ISSUED": "Issued",
        "CONFIRMED": "Verified",
        "HANDED_OVER": "Active",
        "EXPIRED": "Expired",
        "CANCELLED": "Cancelled",
        "PENDING_CLOSE": "Closing",
        "CLOSED": "Closed",
        "LIQUIDATED": "Liquidated",
        "SLA_EXCEEDED": "SLA Breach",
    }

    if export_type == "summary":
        headers = ["Serial", "LG Ref", "Status", "Amount", "Currency", "Expiry Date", "Bank", "Beneficiary"]
    elif export_type == "detailed":
        headers = [
            "Serial", "LG Ref", "Status", "Amount", "Currency",
            "Issue Date", "Expiry Date", "Bank", "Beneficiary",
            "Bank LG Number", "Facility", "Sub-Limit", "Method",
            "Requestor", "Department", "LG Type", "LG Purpose",
            "Reference Type", "Reference #", "Validity Flag",
            "Delivery Date", "Delivery Method",
            "Bank Reply Type", "Bank Reply Date",
            "Verification Status", "Verified At",
            "Handover Date", "Recipient",
            "Custody Holder", "Issued By", "Created At",
        ]
    else:  # full_audit
        headers = [
            "Serial", "LG Ref", "Status", "Amount", "Currency",
            "Issue Date", "Expiry Date", "Bank", "Beneficiary",
            "Bank LG Number", "Facility", "Sub-Limit", "Method",
            "Requestor", "Department", "LG Type", "LG Purpose",
            "Reference Type", "Reference #", "Validity Flag",
            "Delivery Date", "Delivery Method",
            "Bank Reply Type", "Bank Reply Date",
            "Verification Status", "Verified At",
            "Handover Date", "Recipient",
            "Custody Holder", "Issued By", "Created At",
            "Action History", "Approval Chain",
        ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_idx, r in enumerate(records, 2):
        bank_name = r.bank.name if r.bank else "N/A"
        if not r.bank_id and r.sub_limit and r.sub_limit.facility and r.sub_limit.facility.bank:
            bank_name = r.sub_limit.facility.bank.name
        currency_code = r.currency.iso_code if r.currency else "N/A"
        facility_name = r.sub_limit.facility.facility_name if r.sub_limit and r.sub_limit.facility else ""
        sub_limit_name = r.sub_limit.limit_name if r.sub_limit else ""
        owner_name = ""
        if getattr(r, 'current_owner_user_id', None):
            owner = db.query(User).filter(User.id == r.current_owner_user_id).first()
            owner_name = owner.email if owner else ""

        # Friendly status label
        display_status = status_labels.get(r.status, r.status or "")

        # Build row data from request
        ref_type = ""
        ref_number = ""
        requestor_name = ""
        department = ""
        lg_type = ""
        lg_purpose = ""
        issued_by_name = ""
        approval_chain_str = ""
        if r.request_id:
            req = db.query(IssuanceRequest).filter(IssuanceRequest.id == r.request_id).first()
            if req:
                ref_type = req.reference_type or ""
                ref_number = req.reference_number or ""
                lg_type = (req.lg_type.name if req.lg_type else "") or ""
                lg_purpose = req.lg_purpose or ""
                department = req.department or ""
                # Requestor — use stored name, fall back to user email
                requestor_name = req.requestor_name or ""
                if not requestor_name and req.requestor_user_id:
                    requestor = db.query(User).filter(User.id == req.requestor_user_id).first()
                    requestor_name = requestor.email if requestor else ""
                # Issued by
                if req.issued_by_user_id:
                    issuer = db.query(User).filter(User.id == req.issued_by_user_id).first()
                    issued_by_name = issuer.email if issuer else ""
                # Approval chain
                if export_type == "full_audit" and req.approval_chain_audit:
                    import json as json_mod
                    approval_chain_str = json_mod.dumps(req.approval_chain_audit, default=str)

        # Tracking fields
        delivery_date = str(getattr(r, 'delivery_date', '') or '')
        delivery_method = getattr(r, 'delivery_method', '') or ''
        bank_reply_type = getattr(r, 'bank_reply_type', '') or ''
        bank_reply_date = str(getattr(r, 'bank_reply_date', '') or '')
        verification_status = getattr(r, 'verification_status', '') or ''
        verified_at = str(getattr(r, 'verified_at', '') or '')
        handover_date = str(getattr(r, 'handover_date', '') or '')
        recipient_name = getattr(r, 'recipient_name', '') or ''
        internal_serial = getattr(r, 'internal_serial', '') or ''
        bank_lg_number = getattr(r, 'bank_lg_number', '') or ''

        if export_type == "summary":
            row_data = [
                internal_serial, r.lg_ref_number, display_status,
                float(r.current_amount), currency_code,
                str(r.expiry_date) if r.expiry_date else "", bank_name, r.beneficiary_name,
            ]
        elif export_type == "detailed":
            row_data = [
                internal_serial, r.lg_ref_number, display_status,
                float(r.current_amount), currency_code,
                str(r.issue_date) if r.issue_date else "", str(r.expiry_date) if r.expiry_date else "",
                bank_name, r.beneficiary_name, bank_lg_number,
                facility_name, sub_limit_name, r.issuance_method or "",
                requestor_name, department, lg_type, lg_purpose,
                ref_type, ref_number, getattr(r, 'reference_validity_flag', "") or "",
                delivery_date, delivery_method,
                bank_reply_type, bank_reply_date,
                verification_status, verified_at,
                handover_date, recipient_name,
                r.custody_holder or "", issued_by_name,
                r.created_at.isoformat() if r.created_at else "",
            ]
        else:  # full_audit
            import json
            history = json.dumps(r.action_history or r.custody_transfer_log or [], default=str)
            row_data = [
                internal_serial, r.lg_ref_number, display_status,
                float(r.current_amount), currency_code,
                str(r.issue_date) if r.issue_date else "", str(r.expiry_date) if r.expiry_date else "",
                bank_name, r.beneficiary_name, bank_lg_number,
                facility_name, sub_limit_name, r.issuance_method or "",
                requestor_name, department, lg_type, lg_purpose,
                ref_type, ref_number, getattr(r, 'reference_validity_flag', "") or "",
                delivery_date, delivery_method,
                bank_reply_type, bank_reply_date,
                verification_status, verified_at,
                handover_date, recipient_name,
                r.custody_holder or "", issued_by_name,
                r.created_at.isoformat() if r.created_at else "",
                history, approval_chain_str,
            ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"issued_lgs_{export_type}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/issued-lgs/{lg_id}/reprint")
async def reprint_issued_lg(
    lg_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    Reprint/view a previously issued LG. Primarily serves the uploaded LG
    document (soft_copy_path or BANK_LG_COPY attachment). Falls back to
    re-generation only if no uploaded document exists.
    """
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest, IssuanceRequestDocument
    from fastapi.responses import StreamingResponse, RedirectResponse
    import io, os

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="Issued LG not found.")

    # ---- PRIORITY 1: Serve uploaded soft copy (GCS path) ----
    if lg.soft_copy_path:
        # If it's a GCS URI, generate a signed URL and redirect
        if lg.soft_copy_path.startswith("gs://"):
            try:
                from app.core.ai_integration import storage_client
                import datetime as dt
                parts = lg.soft_copy_path.replace("gs://", "").split("/", 1)
                bucket_name = parts[0]
                blob_path = parts[1] if len(parts) > 1 else ""
                bucket = storage_client.bucket(bucket_name)
                blob = bucket.blob(blob_path)
                signed_url = blob.generate_signed_url(
                    version="v4",
                    expiration=dt.timedelta(minutes=15),
                    method="GET"
                )
                return RedirectResponse(url=signed_url)
            except Exception as e:
                # Fall through to other methods
                import logging
                logging.getLogger(__name__).warning(f"Failed to serve soft_copy_path: {e}")
        # If it's a local path
        elif os.path.exists(lg.soft_copy_path):
            with open(lg.soft_copy_path, "rb") as f:
                content = f.read()
            ext = lg.soft_copy_path.rsplit(".", 1)[-1].lower()
            media_types = {"pdf": "application/pdf", "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "tiff": "image/tiff", "webp": "image/webp"}
            media_type = media_types.get(ext, "application/octet-stream")
            return StreamingResponse(
                io.BytesIO(content),
                media_type=media_type,
                headers={"Content-Disposition": f'inline; filename="lg_{lg.lg_ref_number}.{ext}"'}
            )

    # ---- PRIORITY 2: Serve BANK_LG_COPY document attached to the request ----
    if lg.request_id:
        doc = db.query(IssuanceRequestDocument).filter(
            IssuanceRequestDocument.request_id == lg.request_id,
            IssuanceRequestDocument.document_type == "BANK_LG_COPY"
        ).order_by(IssuanceRequestDocument.created_at.desc()).first()

        if doc and doc.file_path:
            file_path = doc.file_path
            if file_path.startswith("gs://"):
                try:
                    from app.core.ai_integration import storage_client
                    import datetime as dt
                    parts = file_path.replace("gs://", "").split("/", 1)
                    bucket_name = parts[0]
                    blob_path = parts[1] if len(parts) > 1 else ""
                    bucket = storage_client.bucket(bucket_name)
                    blob = bucket.blob(blob_path)
                    signed_url = blob.generate_signed_url(
                        version="v4",
                        expiration=dt.timedelta(minutes=15),
                        method="GET"
                    )
                    return RedirectResponse(url=signed_url)
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).warning(f"Failed to serve BANK_LG_COPY: {e}")
            elif os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    content = f.read()
                ext = file_path.rsplit(".", 1)[-1].lower()
                media_types = {"pdf": "application/pdf", "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png"}
                media_type = media_types.get(ext, "application/octet-stream")
                return StreamingResponse(
                    io.BytesIO(content),
                    media_type=media_type,
                    headers={"Content-Disposition": f'inline; filename="lg_{lg.lg_ref_number}.{ext}"'}
                )

    # ---- FALLBACK: No uploaded document found ----
    raise HTTPException(
        status_code=404,
        detail="No uploaded LG document found for this record. Please upload the LG copy first."
    )


@router.get("/banks/{bank_id}/issuance-options")
def get_bank_issuance_options(
    bank_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """Return available issuance methods for a bank (company letter, bank form, API)."""
    from app.models.models_issuance import BankIssuanceOption, BankFormTemplate

    options = db.query(BankIssuanceOption).filter(
        BankIssuanceOption.bank_id == bank_id,
        BankIssuanceOption.is_active == True
    ).all()

    # Also check if this bank has any active form templates
    has_form_templates = db.query(BankFormTemplate).filter(
        BankFormTemplate.bank_id == bank_id,
        BankFormTemplate.is_active == True,
        BankFormTemplate.is_deleted == False
    ).first() is not None

    # Build response — always include "Company Letter" as a base option
    methods = []

    # Company Letter is always available
    methods.append({
        "id": "COMPANY_LETTER",
        "strategy_code": "COMPANY_LETTER",
        "display_name": "Company Letter",
        "description": "Generate a signed company letter to the bank requesting LG issuance",
        "available": True,
    })

    # Bank Form if templates exist
    methods.append({
        "id": "BANK_FORM",
        "strategy_code": "BANK_FORM",
        "display_name": "Fill Bank Form",
        "description": "Auto-fill the bank's official PDF application form",
        "available": has_form_templates,
    })

    # Add any custom options from BankIssuanceOption table
    for opt in options:
        if opt.strategy_code not in ["COMPANY_LETTER", "BANK_FORM"]:
            methods.append({
                "id": str(opt.id),
                "strategy_code": opt.strategy_code,
                "display_name": opt.display_name,
                "description": opt.configuration.get("description", ""),
                "available": True,
            })

    # API placeholder
    methods.append({
        "id": "BANK_API",
        "strategy_code": "BANK_API",
        "display_name": "Bank API",
        "description": "Submit directly via bank's API integration (coming soon)",
        "available": False,
    })

    return methods


@router.post("/bank-forms/auto-fill/{request_id}")
async def auto_fill_bank_form(
    request_id: int,
    bank_id: int = Query(..., description="The bank to find forms for"),
    user_values: Optional[Dict[str, str]] = Body(None, description="User-provided values for missing fields"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    Two-phase bank form auto-fill:
    
    Phase 1 (user_values=null): Build data, detect missing fields.
      - If ALL fields have values → return filled PDF immediately.
      - If some fields are empty → return JSON with missing_fields + saved_values.
    
    Phase 2 (user_values={...}): Merge user values into data, generate PDF, save values for reuse.
    """
    from sqlalchemy.orm import selectinload
    from app.models.models_issuance import FormFieldUserValue
    
    # Load request
    request = db.query(IssuanceRequest).options(
        selectinload(IssuanceRequest.currency),
        selectinload(IssuanceRequest.lg_type),
        selectinload(IssuanceRequest.issuing_entity),
        selectinload(IssuanceRequest.customer),
        selectinload(IssuanceRequest.project),
    ).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id,
    ).first()
    
    if not request:
        raise HTTPException(404, "Issuance request not found.")
    
    # Find best matching form: type-specific first, then universal, language-aware
    req_lang = getattr(request, 'lg_language', 'AR') or 'AR'  # AR or EN
    
    all_candidates = db.query(BankFormTemplate).filter(
        BankFormTemplate.bank_id == bank_id,
        BankFormTemplate.is_active == True,
        BankFormTemplate.is_deleted == False,
        BankFormTemplate.ai_analysis_status == "COMPLETED",
    ).order_by(BankFormTemplate.version.desc()).all()
    
    # 4-tier priority selection:
    # 1. Type-specific + matching language (AR/EN)
    # 2. Type-specific + BILINGUAL
    # 3. Universal + matching language
    # 4. Universal + BILINGUAL
    form_template = None
    
    # P1: type match + exact language
    for f in all_candidates:
        if f.lg_type_ids and request.lg_type_id in f.lg_type_ids:
            if getattr(f, 'form_language', 'BILINGUAL') == req_lang:
                form_template = f
                break
    
    # P2: type match + bilingual
    if not form_template:
        for f in all_candidates:
            if f.lg_type_ids and request.lg_type_id in f.lg_type_ids:
                if getattr(f, 'form_language', 'BILINGUAL') == 'BILINGUAL':
                    form_template = f
                    break
    
    # P3: universal + exact language
    if not form_template:
        for f in all_candidates:
            if not f.lg_type_ids:
                if getattr(f, 'form_language', 'BILINGUAL') == req_lang:
                    form_template = f
                    break
    
    # P4: universal + bilingual
    if not form_template:
        for f in all_candidates:
            if not f.lg_type_ids:
                if getattr(f, 'form_language', 'BILINGUAL') == 'BILINGUAL':
                    form_template = f
                    break
    
    if not form_template and all_candidates:
        # Last resort: any active analyzed form for this bank
        form_template = all_candidates[0]
    
    if not form_template:
        raise HTTPException(404, f"No analyzed bank form template found for this bank. Please upload and analyze a form first.")
    
    if not form_template.field_mapping:
        raise HTTPException(400, "Form has no field mapping. Run AI analysis first.")
    
    if not form_template.file_path:
        raise HTTPException(400, "No PDF file associated with this form.")
    
    import os
    if not os.path.exists(form_template.file_path):
        raise HTTPException(404, "PDF file not found on disk.")
    
    # Build data dict (auto-fills from system data + bank account)
    from app.core.pdf_form_filler import fill_pdf_form, build_request_data_dict
    request_data = build_request_data_dict(request, db, bank_id=bank_id)
    
    # Look up special wording attachment (for auto-open after form download)
    special_wording_doc_id = None
    if request.requires_special_wording:
        from app.models.models_issuance import IssuanceRequestDocument
        sw_doc = db.query(IssuanceRequestDocument).filter(
            IssuanceRequestDocument.request_id == request.id,
            IssuanceRequestDocument.document_type == "SPECIAL_WORDING",
            IssuanceRequestDocument.is_deleted == False,
        ).first()
        if sw_doc:
            special_wording_doc_id = sw_doc.id
    
    # Load saved user values for this customer + form
    saved_rows = db.query(FormFieldUserValue).filter(
        FormFieldUserValue.customer_id == current_user.customer_id,
        FormFieldUserValue.form_template_id == form_template.id,
    ).all()
    saved_values = {row.pdf_field_name: row.saved_value for row in saved_rows}
    
    # ── PHASE 1: Detect missing fields ──
    if user_values is None:
        missing_fields = []
        for mapping_entry in form_template.field_mapping:
            mapped_to = mapping_entry.get("mapped_to", "")
            pdf_field = mapping_entry.get("pdf_field_name", "")
            field_type = mapping_entry.get("field_type", "text").lower()
            
            # Skip checkboxes — they always have a boolean value
            if field_type == "checkbox":
                continue
            
            # Check if the system data has a value
            val = request_data.get(mapped_to, "")
            if isinstance(val, (bool, int, float)):
                continue  # These always have values
            
            is_empty = not val or str(val).strip() == ""
            
            if is_empty:
                missing_fields.append({
                    "pdf_field_name": pdf_field,
                    "label": mapping_entry.get("label", pdf_field),
                    "mapped_to": mapped_to,
                    "field_type": field_type,
                    "saved_value": saved_values.get(pdf_field, ""),
                })
        
        # If nothing is missing, generate PDF directly
        if not missing_fields:
            # Merge any saved values that might map to unmapped fields
            for pdf_field, sv in saved_values.items():
                if sv:
                    request_data[pdf_field] = sv
            
            with open(form_template.file_path, "rb") as f:
                template_pdf_bytes = f.read()
            
            if form_template.form_type == "PHYSICAL_OVERLAY":
                from app.core.pdf_form_filler import generate_overlay_pdf
                # Only filter by language on BILINGUAL forms; single-language forms get all fields
                fill_lang = req_lang if getattr(form_template, 'form_language', 'BILINGUAL') == 'BILINGUAL' else None
                filled_pdf = generate_overlay_pdf(
                    template_pdf_bytes=template_pdf_bytes,
                    field_mapping=form_template.field_mapping,
                    request_data=request_data,
                    lg_language=fill_lang,
                )
            elif form_template.form_type == "SCANNED_FILL":
                from app.core.pdf_form_filler import generate_scanned_fill_pdf
                fill_lang = req_lang if getattr(form_template, 'form_language', 'BILINGUAL') == 'BILINGUAL' else None
                filled_pdf = generate_scanned_fill_pdf(
                    template_pdf_bytes=template_pdf_bytes,
                    field_mapping=form_template.field_mapping,
                    request_data=request_data,
                    lg_language=fill_lang,
                )
            else:
                fill_lang = req_lang if getattr(form_template, 'form_language', 'BILINGUAL') == 'BILINGUAL' else None
                filled_pdf = fill_pdf_form(
                    template_pdf_bytes=template_pdf_bytes,
                    field_mapping=form_template.field_mapping,
                    request_data=request_data,
                    lg_language=fill_lang,
                )
            
            filename = f"Filled_{form_template.name}_{request.serial_number}.pdf"
            return StreamingResponse(
                io.BytesIO(filled_pdf),
                media_type="application/pdf",
                headers={
                    'Content-Disposition': f'inline; filename="{filename}"',
                    'X-Form-Template-Id': str(form_template.id),
                    'X-Form-Template-Name': form_template.name,
                    'X-Form-Type': form_template.form_type or 'FILLABLE_PDF',
                    'X-Special-Wording-Doc-Id': str(special_wording_doc_id) if special_wording_doc_id else '',
                    'Access-Control-Expose-Headers': 'X-Form-Type, X-Special-Wording-Doc-Id',
                }
            )
        
        # Some fields are missing — return them for user input
        return {
            "status": "missing_fields",
            "form_template_id": form_template.id,
            "form_template_name": form_template.name,
            "form_type": form_template.form_type or "FILLABLE_PDF",
            "missing_fields": missing_fields,
            "total_fields": len(form_template.field_mapping),
            "auto_filled_fields": len(form_template.field_mapping) - len(missing_fields),
            "special_wording_doc_id": special_wording_doc_id,
        }
    
    # ── PHASE 2: Merge user values and generate PDF ──
    # Merge user-provided values into request_data
    for mapping_entry in form_template.field_mapping:
        pdf_field = mapping_entry.get("pdf_field_name", "")
        mapped_to = mapping_entry.get("mapped_to", "")
        
        if pdf_field in user_values and user_values[pdf_field]:
            request_data[mapped_to] = user_values[pdf_field]
    
    # Also inject saved values for any still-empty fields
    for pdf_field, sv in saved_values.items():
        if sv and pdf_field not in user_values:
            # Find the mapped_to for this pdf_field
            for m in form_template.field_mapping:
                if m.get("pdf_field_name") == pdf_field:
                    mapped_key = m.get("mapped_to", "")
                    if mapped_key and not request_data.get(mapped_key):
                        request_data[mapped_key] = sv
                    break

    # Save user values for future use
    for pdf_field, value in user_values.items():
        if value is not None:
            existing = db.query(FormFieldUserValue).filter(
                FormFieldUserValue.customer_id == current_user.customer_id,
                FormFieldUserValue.form_template_id == form_template.id,
                FormFieldUserValue.pdf_field_name == pdf_field,
            ).first()
            
            if existing:
                existing.saved_value = value
            else:
                db.add(FormFieldUserValue(
                    customer_id=current_user.customer_id,
                    form_template_id=form_template.id,
                    pdf_field_name=pdf_field,
                    saved_value=value,
                ))
    db.commit()

    # Generate filled PDF
    with open(form_template.file_path, "rb") as f:
        template_pdf_bytes = f.read()
    
    if form_template.form_type == "PHYSICAL_OVERLAY":
        from app.core.pdf_form_filler import generate_overlay_pdf
        fill_lang = req_lang if getattr(form_template, 'form_language', 'BILINGUAL') == 'BILINGUAL' else None
        filled_pdf = generate_overlay_pdf(
            template_pdf_bytes=template_pdf_bytes,
            field_mapping=form_template.field_mapping,
            request_data=request_data,
            lg_language=fill_lang,
        )
    elif form_template.form_type == "SCANNED_FILL":
        from app.core.pdf_form_filler import generate_scanned_fill_pdf
        fill_lang = req_lang if getattr(form_template, 'form_language', 'BILINGUAL') == 'BILINGUAL' else None
        filled_pdf = generate_scanned_fill_pdf(
            template_pdf_bytes=template_pdf_bytes,
            field_mapping=form_template.field_mapping,
            request_data=request_data,
            lg_language=fill_lang,
        )
    else:
        fill_lang = req_lang if getattr(form_template, 'form_language', 'BILINGUAL') == 'BILINGUAL' else None
        filled_pdf = fill_pdf_form(
            template_pdf_bytes=template_pdf_bytes,
            field_mapping=form_template.field_mapping,
            request_data=request_data,
            lg_language=fill_lang,
        )
    
    filename = f"Filled_{form_template.name}_{request.serial_number}.pdf"
    
    return StreamingResponse(
        io.BytesIO(filled_pdf),
        media_type="application/pdf",
        headers={
            'Content-Disposition': f'inline; filename="{filename}"',
            'X-Form-Template-Id': str(form_template.id),
            'X-Form-Template-Name': form_template.name,
        }
    )


@router.post("/issued-lgs/{lg_id}/reprint")
async def reprint_issued_lg(
    lg_id: int,
    additional_text: str = Query("", description="Extra instructions for the letter"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """Reprint an issued LG's letter/form. Logs the reprint action."""
    from app.models.models_issuance import IssuedLGRecord

    lg_record = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg_record:
        raise HTTPException(status_code=404, detail="Issued LG record not found.")

    if not lg_record.request_id:
        raise HTTPException(status_code=400, detail="No linked request found for this LG record.")

    # Log the reprint action in the custody_transfer_log (reusing the JSONB field)
    from datetime import datetime as dt
    reprint_log = lg_record.custody_transfer_log or []
    reprint_log.append({
        "action": "REPRINT",
        "user_id": current_user.user_id,
        "timestamp": dt.now().isoformat(),
        "method": lg_record.issuance_method,
    })
    lg_record.custody_transfer_log = reprint_log
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(lg_record, "custody_transfer_log")
    db.commit()

    # Re-generate the letter
    result = await issuance_service.generate_issuance_letter(
        db=db,
        request_id=lg_record.request_id,
        customer_id=current_user.customer_id,
        additional_text=additional_text,
    )
    headers = {
        'Content-Disposition': f'inline; filename="{result["filename"]}"'
    }
    from starlette.responses import Response
    return Response(
        content=result["pdf_bytes"],
        media_type="application/pdf",
        headers=headers
    )


# ==============================================================================
# 8. POST-ISSUANCE TRACKING (Steps 5.5 + 5.6)
# ==============================================================================

@router.patch("/lg-records/{lg_id}/record-delivery")
def record_delivery(
    lg_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """
    Step 5.5a: Record delivery of bank form to bank.
    Optionally requires delivery proof document (configurable per customer).
    Only END_USER (treasury officer) can execute this.
    """
    from app.constants import UserRole
    if current_user.role not in (UserRole.END_USER, UserRole.END_USER.value):
        raise HTTPException(status_code=403, detail="Only treasury end users can record delivery.")
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequestDocument
    from app.crud.crud import log_action
    from app.crud import crud_customer_configuration
    from app.constants import GlobalConfigKey

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="LG record not found.")

    if lg.delivery_date:
        raise HTTPException(status_code=400, detail="Delivery already recorded.")

    # Check if delivery proof is mandatory
    proof_required_config = crud_customer_configuration.get_customer_config_or_global_fallback(
        db, current_user.customer_id, GlobalConfigKey.DOC_MANDATORY_RECORD_DELIVERY
    )
    proof_required = (proof_required_config or {}).get("effective_value", "false").lower() == "true"

    if proof_required:
        # Check if a DELIVERY_PROOF document exists for the related request
        request_id = lg.request_id
        if request_id:
            proof_doc = db.query(IssuanceRequestDocument).filter(
                IssuanceRequestDocument.request_id == request_id,
                IssuanceRequestDocument.document_type == "DELIVERY_PROOF"
            ).first()
            if not proof_doc:
                raise HTTPException(
                    status_code=400,
                    detail="Delivery proof document is required before recording delivery. Please upload a scanned copy with bank receiving stamp."
                )

    from datetime import date
    lg.delivery_date = payload.get("delivery_date") or date.today().isoformat()
    lg.delivery_method = payload.get("delivery_method", "HAND_DELIVERY")
    lg.delivery_notes = payload.get("delivery_notes")
    lg.status = "DELIVERED_TO_BANK"

    log_action(db, current_user.user_id, "ISSUANCE_DELIVERY_RECORDED", "IssuedLGRecord", lg.id,
               {"delivery_method": lg.delivery_method, "delivery_date": str(lg.delivery_date)},
               current_user.customer_id)

    return {
        "message": "Delivery recorded successfully.",
        "id": lg.id,
        "status": lg.status,
        "delivery_date": str(lg.delivery_date),
        "delivery_method": lg.delivery_method
    }


@router.patch("/lg-records/{lg_id}/record-bank-reply")
def record_bank_reply(
    lg_id: int,
    payload: dict,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context)
):
    """
    Step 5.5b: Record bank's reply to issuance request.
    Reply types: LG_ISSUED, INQUIRY, REJECTED, NO_RESPONSE

    INQUIRY: Logs a dated note to bank_inquiry_log but does NOT finalize.
             The BANK_REPLY step stays open for a later final reply.
    REJECTED / NO_RESPONSE: Finalizes this LG attempt AND reopens
             the original IssuanceRequest for reprocessing.
    Only END_USER (treasury officer) can execute this.
    """
    from app.constants import UserRole
    if current_user.role not in (UserRole.END_USER, UserRole.END_USER.value):
        raise HTTPException(status_code=403, detail="Only treasury end users can record bank replies.")
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.crud.crud import log_action
    from datetime import date, datetime

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="LG record not found.")

    reply_type = payload.get("bank_reply_type")
    if reply_type not in ("LG_ISSUED", "INQUIRY", "REJECTED", "NO_RESPONSE"):
        raise HTTPException(status_code=400, detail="Invalid reply type. Must be: LG_ISSUED, INQUIRY, REJECTED, NO_RESPONSE")

    # ── INQUIRY: append note only, keep step open ──
    if reply_type == "INQUIRY":
        inquiry_entry = {
            "date": payload.get("bank_reply_date") or date.today().isoformat(),
            "notes": payload.get("bank_reply_notes", ""),
            "type": "INQUIRY",
            "logged_by_user_id": current_user.user_id,
            "created_at": datetime.utcnow().isoformat(),
        }
        current_log = list(lg.bank_inquiry_log or [])
        current_log.append(inquiry_entry)
        lg.bank_inquiry_log = current_log
        # Do NOT set bank_reply_type — keeps bank reply selector open
        # Status stays as whatever it was (PENDING_BANK_REPLY or similar)

        log_action(db, current_user.user_id, "ISSUANCE_BANK_INQUIRY_NOTED", "IssuedLGRecord", lg.id,
                   {"inquiry_notes": inquiry_entry["notes"], "inquiry_date": inquiry_entry["date"]},
                   current_user.customer_id)

        return {
            "message": "Bank inquiry noted. You can continue to record the final bank reply when ready.",
            "id": lg.id,
            "status": lg.status,
            "bank_reply_type": None,  # Not finalized
            "inquiry_count": len(current_log),
        }

    # ── LG_ISSUED: normal flow → pending verification ──
    lg.bank_reply_type = reply_type
    lg.bank_reply_date = payload.get("bank_reply_date") or date.today().isoformat()
    lg.bank_reply_notes = payload.get("bank_reply_notes")

    if reply_type == "LG_ISSUED":
        lg.bank_lg_number = payload.get("bank_lg_number")
        lg.bank_lg_issue_date = payload.get("bank_lg_issue_date")
        lg.bank_lg_expiry_date = payload.get("bank_lg_expiry_date")
        lg.bank_lg_amount = payload.get("bank_lg_amount") or None
        # D2: Populate issue_date from bank's confirmed issue date
        lg.issue_date = payload.get("bank_lg_issue_date") or date.today()
        lg.status = "PENDING_VERIFICATION"
        lg.verification_status = "PENDING"

    # ── REJECTED / NO_RESPONSE: close this LG, reopen request ──
    elif reply_type in ("REJECTED", "NO_RESPONSE"):
        lg.status = "BANK_REJECTED" if reply_type == "REJECTED" else "SLA_EXCEEDED"

        # Reopen the original IssuanceRequest for reprocessing
        if lg.request_id:
            request_obj = db.query(IssuanceRequest).get(lg.request_id)
            if request_obj:
                request_obj.status = "APPROVED"  # Back to processing queue
                request_obj.lg_record_id = None   # Unlink from this failed LG

    log_action(db, current_user.user_id, "ISSUANCE_BANK_REPLY_RECORDED", "IssuedLGRecord", lg.id,
               {"reply_type": reply_type, "bank_lg_number": lg.bank_lg_number},
               current_user.customer_id)

    # Notify requestor (all reply types except INQUIRY get email)
    if lg.request_id:
        request = db.query(IssuanceRequest).get(lg.request_id)
        if request and request.requestor_email:
            _send_requestor_status_notification(
                db, background_tasks, request, reply_type, lg
            )

    return {
        "message": f"Bank reply recorded: {reply_type}",
        "id": lg.id,
        "status": lg.status,
        "bank_reply_type": reply_type,
        "bank_lg_number": lg.bank_lg_number,
        "request_reopened": reply_type in ("REJECTED", "NO_RESPONSE"),
    }


@router.patch("/lg-records/{lg_id}/verify")
def verify_lg_copy(
    lg_id: int,
    payload: dict,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    Step 5.6: Verify issued LG copy against original request.
    Auto-compares bank-confirmed values with request values.
    """
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.crud.crud import log_action
    from datetime import datetime

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="LG record not found.")

    if lg.verification_status not in (None, "PENDING", "DISCREPANCY"):
        raise HTTPException(status_code=400, detail=f"Cannot verify — current status: {lg.verification_status}")

    request = db.query(IssuanceRequest).get(lg.request_id) if lg.request_id else None

    # Allow user to update bank values if provided
    if payload.get("bank_lg_number"):
        lg.bank_lg_number = payload["bank_lg_number"]
    if payload.get("bank_lg_amount"):
        lg.bank_lg_amount = payload["bank_lg_amount"]
    if payload.get("bank_lg_issue_date"):
        lg.bank_lg_issue_date = payload["bank_lg_issue_date"]
    if payload.get("bank_lg_expiry_date"):
        lg.bank_lg_expiry_date = payload["bank_lg_expiry_date"]

    # Compliance auto-check
    discrepancies = []
    if request:
        # Amount check
        if lg.bank_lg_amount is not None and request.amount is not None:
            from decimal import Decimal
            bank_amt = Decimal(str(lg.bank_lg_amount))
            req_amt = Decimal(str(request.amount))
            if bank_amt != req_amt:
                discrepancies.append({
                    "field": "amount",
                    "requested": str(req_amt),
                    "bank_confirmed": str(bank_amt),
                    "severity": "HIGH"
                })

        # Expiry date check
        if lg.bank_lg_expiry_date and request.requested_expiry_date:
            if str(lg.bank_lg_expiry_date) != str(request.requested_expiry_date):
                discrepancies.append({
                    "field": "expiry_date",
                    "requested": str(request.requested_expiry_date),
                    "bank_confirmed": str(lg.bank_lg_expiry_date),
                    "severity": "MEDIUM"
                })

        # Beneficiary name (fuzzy)
        bank_beneficiary = payload.get("bank_beneficiary_name", "")
        if bank_beneficiary and request.beneficiary_name:
            if bank_beneficiary.strip().lower() != request.beneficiary_name.strip().lower():
                discrepancies.append({
                    "field": "beneficiary_name",
                    "requested": request.beneficiary_name,
                    "bank_confirmed": bank_beneficiary,
                    "severity": "MEDIUM"
                })

        # Currency check
        bank_currency = payload.get("bank_currency_id") or payload.get("bank_currency")
        if bank_currency and request.currency_id is not None:
            # Support both currency_id (int) and currency code (string)
            if isinstance(bank_currency, int) or (isinstance(bank_currency, str) and bank_currency.isdigit()):
                currency_match = int(bank_currency) == request.currency_id
            else:
                # Lookup by currency code
                from app.models.models import Currency as CurrencyModel
                req_currency = db.query(CurrencyModel).filter(CurrencyModel.id == request.currency_id).first()
                currency_match = req_currency and req_currency.code.upper() == str(bank_currency).upper()
            if not currency_match:
                req_currency_obj = db.query(CurrencyModel).filter(CurrencyModel.id == request.currency_id).first() if 'req_currency' not in dir() else req_currency
                discrepancies.append({
                    "field": "currency",
                    "requested": req_currency_obj.code if req_currency_obj else str(request.currency_id),
                    "bank_confirmed": str(bank_currency),
                    "severity": "HIGH"
                })

        # LG Type check
        bank_lg_type = payload.get("bank_lg_type_id") or payload.get("bank_lg_type")
        if bank_lg_type and request.lg_type_id is not None:
            if isinstance(bank_lg_type, int) or (isinstance(bank_lg_type, str) and bank_lg_type.isdigit()):
                lg_type_match = int(bank_lg_type) == request.lg_type_id
            else:
                from app.models.models import LGType
                req_lg_type = db.query(LGType).filter(LGType.id == request.lg_type_id).first()
                lg_type_match = req_lg_type and req_lg_type.name.strip().lower() == str(bank_lg_type).strip().lower()
            if not lg_type_match:
                discrepancies.append({
                    "field": "lg_type",
                    "requested": str(request.lg_type_id),
                    "bank_confirmed": str(bank_lg_type),
                    "severity": "MEDIUM"
                })

        # Purpose check
        bank_purpose = payload.get("bank_lg_purpose", "")
        if bank_purpose and request.lg_purpose:
            if bank_purpose.strip().lower() != request.lg_purpose.strip().lower():
                discrepancies.append({
                    "field": "purpose",
                    "requested": request.lg_purpose,
                    "bank_confirmed": bank_purpose,
                    "severity": "MEDIUM"
                })

        # Operational Status check (particularly for Advance Payment LGs)
        bank_operational_status = payload.get("bank_operational_status", "")
        if bank_operational_status and request.operational_status:
            if bank_operational_status.strip().lower() != request.operational_status.strip().lower():
                discrepancies.append({
                    "field": "operational_status",
                    "requested": request.operational_status,
                    "bank_confirmed": bank_operational_status,
                    "severity": "MEDIUM"
                })

    # Build full comparison (ALL fields, matched + mismatched) for the admin review
    all_comparisons = []
    if request:
        from decimal import Decimal
        # Amount
        if lg.bank_lg_amount is not None and request.amount is not None:
            bank_amt = Decimal(str(lg.bank_lg_amount))
            req_amt = Decimal(str(request.amount))
            all_comparisons.append({"field": "Amount", "requested": str(req_amt), "bank_confirmed": str(bank_amt), "severity": "HIGH" if bank_amt != req_amt else "OK", "match": bank_amt == req_amt})
        # Expiry Date
        if request.requested_expiry_date:
            bank_exp = str(lg.bank_lg_expiry_date) if lg.bank_lg_expiry_date else "—"
            req_exp = str(request.requested_expiry_date)
            all_comparisons.append({"field": "Expiry Date", "requested": req_exp, "bank_confirmed": bank_exp, "severity": "MEDIUM" if bank_exp != req_exp else "OK", "match": bank_exp == req_exp})
        # Beneficiary
        bank_beneficiary = payload.get("bank_beneficiary_name", "")
        if request.beneficiary_name:
            match_b = not bank_beneficiary or bank_beneficiary.strip().lower() == request.beneficiary_name.strip().lower()
            all_comparisons.append({"field": "Beneficiary", "requested": request.beneficiary_name, "bank_confirmed": bank_beneficiary or "—", "severity": "MEDIUM" if not match_b else "OK", "match": match_b})
        # Currency
        bank_currency_val = payload.get("bank_currency_id") or payload.get("bank_currency")
        if request.currency_id:
            from app.models.models import Currency as CurrencyModel
            req_curr_obj = db.query(CurrencyModel).filter(CurrencyModel.id == request.currency_id).first()
            req_curr_name = req_curr_obj.iso_code if req_curr_obj else str(request.currency_id)
            match_c = not bank_currency_val or req_curr_name.upper() == str(bank_currency_val).upper()
            all_comparisons.append({"field": "Currency", "requested": req_curr_name, "bank_confirmed": str(bank_currency_val) if bank_currency_val else "—", "severity": "HIGH" if not match_c else "OK", "match": match_c})
        # LG Type
        bank_lg_type_val = payload.get("bank_lg_type_id") or payload.get("bank_lg_type")
        if request.lg_type_id:
            from sqlalchemy import text as sa_text
            lg_type_row = db.execute(sa_text("SELECT name FROM lg_types WHERE id = :id"), {"id": request.lg_type_id}).first()
            req_lg_type_name = lg_type_row[0] if lg_type_row else str(request.lg_type_id)
            match_t = not bank_lg_type_val or req_lg_type_name.strip().lower() == str(bank_lg_type_val).strip().lower()
            all_comparisons.append({"field": "LG Type", "requested": req_lg_type_name, "bank_confirmed": str(bank_lg_type_val) if bank_lg_type_val else "—", "severity": "MEDIUM" if not match_t else "OK", "match": match_t})
        # Purpose
        bank_purpose_val = payload.get("bank_lg_purpose", "")
        if request.lg_purpose:
            match_p = not bank_purpose_val or bank_purpose_val.strip().lower() == request.lg_purpose.strip().lower()
            all_comparisons.append({"field": "Purpose", "requested": request.lg_purpose, "bank_confirmed": bank_purpose_val or "—", "severity": "MEDIUM" if not match_p else "OK", "match": match_p})

    # Determine result
    force_accept = payload.get("force_accept", False)

    if not discrepancies:
        # D4: Enforce bank_lg_number before allowing MATCHED status
        if not lg.bank_lg_number and not payload.get("force_no_number", False):
            raise HTTPException(
                status_code=400,
                detail="Bank LG number is required before confirmation. "
                       "Set 'force_no_number' to true if the bank did not assign a number."
            )
        lg.verification_status = "MATCHED"
        lg.status = "CONFIRMED"
    elif force_accept:
        # Only corporate_admin or checker can force-accept discrepancies
        if current_user.role not in (UserRole.CORPORATE_ADMIN, UserRole.CORPORATE_ADMIN.value, UserRole.CHECKER, UserRole.CHECKER.value):
            raise HTTPException(
                status_code=403,
                detail="Only Corporate Admin or Checker can accept discrepancies. Please submit for review."
            )
        # D4: Enforce bank_lg_number before allowing ACCEPTED status
        if not lg.bank_lg_number and not payload.get("force_no_number", False):
            raise HTTPException(
                status_code=400,
                detail="Bank LG number is required before confirmation. "
                       "Set 'force_no_number' to true if the bank did not assign a number."
            )
        lg.verification_status = "ACCEPTED"
        lg.verification_notes = payload.get("verification_notes", "Discrepancies manually accepted")
        lg.status = "CONFIRMED"
    else:
        lg.verification_status = "DISCREPANCY"
        lg.verification_notes = json.dumps(all_comparisons)
        # Status stays PENDING_VERIFICATION — user must review

    lg.verified_by_user_id = current_user.user_id
    lg.verified_at = datetime.utcnow()

    log_action(db, current_user.user_id, "ISSUANCE_LG_VERIFIED", "IssuedLGRecord", lg.id,
               {"verification_status": lg.verification_status, "discrepancies": discrepancies},
               current_user.customer_id)

    # Notify requestor on confirmation
    if lg.status == "CONFIRMED" and request and request.requestor_email:
        _send_requestor_status_notification(
            db, background_tasks, request, "VERIFIED", lg
        )

    return {
        "message": f"Verification complete: {lg.verification_status}",
        "id": lg.id,
        "status": lg.status,
        "verification_status": lg.verification_status,
        "discrepancies": discrepancies
    }


@router.patch("/lg-records/{lg_id}/manual-pricing")
def update_manual_pricing(
    lg_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    D3: Update manual pricing on an IssuedLGRecord.
    Only allowed for LGs issued without a facility (sub_limit_id is NULL).
    """
    from app.models.models_issuance import IssuedLGRecord
    from app.crud.crud import log_action

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="LG record not found.")

    if lg.facility_sub_limit_id is not None:
        raise HTTPException(
            status_code=400,
            detail="Cannot set manual pricing on a facility-backed LG. Pricing is derived from the facility terms."
        )

    # Build clean pricing object
    pricing = {
        "commission_rate": payload.get("commission_rate"),
        "flat_fee": payload.get("flat_fee"),
        "margin_pct": payload.get("margin_pct"),
        "notes": payload.get("notes"),
    }
    # Remove None values for clean storage
    pricing = {k: v for k, v in pricing.items() if v is not None}

    lg.manual_pricing = pricing if pricing else None
    db.flush()

    log_action(db, current_user.user_id, "ISSUANCE_MANUAL_PRICING_UPDATED", "IssuedLGRecord", lg.id,
               {"manual_pricing": pricing}, current_user.customer_id)

    return {
        "message": "Manual pricing updated successfully.",
        "id": lg.id,
        "manual_pricing": lg.manual_pricing
    }


@router.post("/lg-records/{lg_id}/reject-discrepancy")
def reject_discrepancy(
    lg_id: int,
    payload: dict = Body({}),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    Corporate Admin rejects discrepancies — resets verification_status to PENDING
    so the end user can re-upload a corrected LG copy.
    Bank reply data is kept intact (the bank DID reply).
    """
    from app.models.models_issuance import IssuedLGRecord
    from app.crud.crud import log_action

    # Only corporate_admin or checker can reject
    from app.constants import UserRole as _UserRole
    if current_user.role not in (_UserRole.CORPORATE_ADMIN, _UserRole.CORPORATE_ADMIN.value, _UserRole.CHECKER, _UserRole.CHECKER.value):
        raise HTTPException(status_code=403, detail="Only corporate admins or checkers can reject discrepancies.")

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="LG record not found.")

    if lg.verification_status != "DISCREPANCY":
        raise HTTPException(status_code=400, detail=f"LG is not in DISCREPANCY status (current: {lg.verification_status})")

    notes = payload.get("notes", "Discrepancies rejected by corporate admin — re-upload required.")
    from datetime import datetime

    # Log rejection before resetting
    log_action(
        db, current_user.user_id, "reject_discrepancy", "issuance",
        lg.id,
        {"notes": notes, "previous_status": "DISCREPANCY",
         "previous_verification_notes": lg.verification_notes},
        current_user.customer_id,
    )

    # Append rejection note to verification_notes (preserve the discrepancy details)
    existing_notes = lg.verification_notes or ""
    rejection_stamp = f"\n--- REJECTED by Admin ({datetime.utcnow().strftime('%Y-%m-%d %H:%M')}) ---\n{notes}"
    lg.verification_notes = existing_notes + rejection_stamp

    # Set a DISCREPANCY_REJECTED status for admin history, then reset to PENDING for end user
    # We use PENDING so the verification step shows as actionable for the end user
    lg.verification_status = "PENDING"
    lg.status = "PENDING_VERIFICATION"
    # Clear verification tracking so user can re-verify
    lg.verified_by_user_id = None
    lg.verified_at = None
    # DO NOT clear bank_reply_type, bank_reply_date, bank_lg_number, bank_lg_amount, etc.
    # The bank DID reply — those values stay. Only the verification needs to be redone.

    db.commit()

    return {
        "message": "Discrepancy rejected — end user can now re-upload a corrected LG copy.",
        "id": lg.id,
        "verification_status": lg.verification_status,
        "status": lg.status,
    }

@router.get("/lg-records/{lg_id}/post-issuance-status")
def get_post_issuance_status(
    lg_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """
    Returns the full post-issuance timeline for a given LG record.
    Used by the PostIssuanceTracker frontend component.
    """
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest, IssuanceRequestDocument, IssuanceFacility
    from app.crud import crud_customer_configuration
    from app.constants import GlobalConfigKey
    from datetime import date, timedelta

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="LG record not found.")

    request = db.query(IssuanceRequest).get(lg.request_id) if lg.request_id else None

    # Get documents
    docs = []
    if lg.request_id:
        doc_records = db.query(IssuanceRequestDocument).filter(
            IssuanceRequestDocument.request_id == lg.request_id,
            IssuanceRequestDocument.document_type.in_(["DELIVERY_PROOF", "BANK_REPLY", "BANK_LG_COPY"])
        ).all()
        docs = [{"id": d.id, "type": d.document_type, "file_name": d.file_name,
                 "created_at": d.created_at.isoformat() if d.created_at else None} for d in doc_records]

    # Compute SLA info
    delivery_sla_config = crud_customer_configuration.get_customer_config_or_global_fallback(
        db, current_user.customer_id, GlobalConfigKey.REMINDER_TO_BANKS_DAYS_SINCE_ISSUANCE
    )
    delivery_sla_days = int((delivery_sla_config or {}).get("effective_value", 3))

    # Bank reply SLA: use facility SLA if available, else fallback to 5
    bank_reply_sla_days = 5
    if request and request.selected_sub_limit_id:
        from sqlalchemy.orm import joinedload
        sub_limit = db.query(IssuanceFacilitySubLimit).options(
            joinedload(IssuanceFacilitySubLimit.facility)
        ).get(request.selected_sub_limit_id)
        if sub_limit and sub_limit.facility and sub_limit.facility.sla_agreement_days:
            bank_reply_sla_days = sub_limit.facility.sla_agreement_days

    # Delivery proof requirement
    proof_required_config = crud_customer_configuration.get_customer_config_or_global_fallback(
        db, current_user.customer_id, GlobalConfigKey.DOC_MANDATORY_RECORD_DELIVERY
    )
    delivery_proof_required = (proof_required_config or {}).get("effective_value", "false").lower() == "true"

    # SLA breach checks
    today = date.today()
    delivery_sla_breached = (
        lg.status in ("ISSUED", "PENDING_CONFIRMATION") and
        not lg.delivery_date and
        lg.created_at and
        (today - lg.created_at.date()).days > delivery_sla_days
    )
    bank_reply_sla_breached = (
        lg.delivery_date and
        not lg.bank_reply_type and
        (today - lg.delivery_date).days > bank_reply_sla_days
    )

    # Build timeline steps
    steps = [
        {
            "step": "ISSUED",
            "label": "LG Issued",
            "status": "completed",
            "date": lg.created_at.isoformat() if lg.created_at else None,
            "details": {"lg_ref": lg.lg_ref_number, "method": lg.issuance_method}
        },
        {
            "step": "DELIVERY",
            "label": "Delivered to Bank",
            "status": "completed" if lg.delivery_date else ("sla_breach" if delivery_sla_breached else "pending"),
            "date": str(lg.delivery_date) if lg.delivery_date else None,
            "details": {
                "method": lg.delivery_method,
                "notes": lg.delivery_notes,
                "proof_required": delivery_proof_required,
                "sla_days": delivery_sla_days,
            }
        },
        {
            "step": "BANK_REPLY",
            "label": "Bank Reply",
            "status": (
                "completed" if lg.bank_reply_type else
                ("sla_breach" if bank_reply_sla_breached else
                 ("pending" if lg.delivery_date else "future"))
            ),
            "date": str(lg.bank_reply_date) if lg.bank_reply_date else None,
            "details": {
                "reply_type": lg.bank_reply_type,
                "notes": lg.bank_reply_notes,
                "bank_lg_number": lg.bank_lg_number,
                "bank_lg_amount": str(lg.bank_lg_amount) if lg.bank_lg_amount else None,
                "bank_lg_expiry_date": str(lg.bank_lg_expiry_date) if lg.bank_lg_expiry_date else None,
                "sla_days": bank_reply_sla_days,
                "inquiry_log": lg.bank_inquiry_log or [],
            }
        },
        {
            "step": "VERIFICATION",
            "label": "LG Copy Verification",
            "status": (
                "completed" if lg.verification_status in ("MATCHED", "ACCEPTED") else
                ("discrepancy" if lg.verification_status == "DISCREPANCY" else
                 ("pending" if lg.bank_reply_type == "LG_ISSUED" else "future"))
            ),
            "date": lg.verified_at.isoformat() if lg.verified_at else None,
            "details": {
                "verification_status": lg.verification_status,
                "notes": lg.verification_notes,
                "verified_by": lg.verified_by_user_id,
                "request_amount": str(request.amount) if request and request.amount else None,
                "request_expiry": str(request.requested_expiry_date) if request else None,
                "request_beneficiary": request.beneficiary_name if request else None,
            }
        },
        {
            "step": "HANDOVER",
            "label": "LG Handover",
            "status": (
                "completed" if lg.handover_date else
                ("pending" if lg.verification_status in ("MATCHED", "ACCEPTED") else "future")
            ),
            "date": str(lg.handover_date) if lg.handover_date else None,
            "details": {
                "recipient_name": lg.recipient_name,
                "recipient_email": lg.recipient_email,
                "recipient_department": lg.recipient_department,
                "recipient_job_title": lg.recipient_job_title,
                "recipient_phone": lg.recipient_phone,
                "recipient_employee_id": lg.recipient_employee_id,
                "recipient_manager_email": lg.recipient_manager_email,
                "notes": lg.handover_notes,
                # Pre-fill defaults from requestor
                "requestor_defaults": {
                    "name": request.requestor_name if request else None,
                    "email": request.requestor_email if request else None,
                    "department": request.department if request else None,
                    "job_title": request.job_title if request else None,
                    "phone": request.phone_number if request else None,
                    "employee_id": request.employee_id if request else None,
                    "manager_email": request.manager_email if request else None,
                    "second_line_manager_email": request.second_line_manager_email if request else None,
                } if request else None,
            }
        }
    ]

    # Get recipient field config
    recipient_field_config = {}
    form_config = db.query(CustomerFormConfiguration).filter(
        CustomerFormConfiguration.customer_id == current_user.customer_id
    ).first()
    if form_config and hasattr(form_config, 'recipient_field_configurations') and form_config.recipient_field_configurations:
        recipient_field_config = form_config.recipient_field_configurations

    # Get handover signed copy requirement
    handover_doc_config = crud_customer_configuration.get_customer_config_or_global_fallback(
        db, current_user.customer_id, GlobalConfigKey.DOC_MANDATORY_LG_HANDOVER
    )
    handover_signed_copy_required = (handover_doc_config or {}).get("effective_value", "false").lower() == "true"

    return {
        "lg_id": lg.id,
        "lg_ref": lg.lg_ref_number,
        "overall_status": lg.status,
        "steps": steps,
        "documents": docs,
        "recipient_field_config": recipient_field_config,
        "handover_signed_copy_required": handover_signed_copy_required,
    }


@router.patch("/lg-records/{lg_id}/record-handover")
@router.post("/lg-records/{lg_id}/record-handover")
async def record_handover(
    lg_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """
    Record LG handover to recipient. Default = requestor from original request.
    Accepts JSON (PATCH) or multipart FormData with optional signed_copy file (POST).
    Only END_USER (treasury officer) can execute this.
    """
    from app.constants import UserRole
    if current_user.role not in (UserRole.END_USER, UserRole.END_USER.value):
        raise HTTPException(status_code=403, detail="Only treasury end users can record handover.")
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.core.email_service import send_email, get_global_email_settings
    from app.crud import crud_customer_configuration
    from app.constants import GlobalConfigKey
    from datetime import date as date_type, datetime
    import os, json

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id,
    ).first()
    if not lg:
        raise HTTPException(404, "Issued LG record not found.")

    # Must be verified first
    if lg.verification_status not in ("MATCHED", "ACCEPTED"):
        raise HTTPException(400, "LG must be verified before handover. Complete verification first.")

    if lg.handover_date:
        raise HTTPException(400, "Handover already recorded for this LG.")

    # Parse payload from either JSON or FormData
    content_type = request.headers.get("content-type", "")
    signed_copy_file = None
    if "multipart/form-data" in content_type:
        form = await request.form()
        data_str = form.get("data", "{}")
        payload = json.loads(data_str)
        signed_copy_file = form.get("signed_copy")
    else:
        payload = await request.json()

    # Validate required fields
    recipient_name = payload.get("recipient_name")
    recipient_email = payload.get("recipient_email")
    if not recipient_name or not recipient_email:
        raise HTTPException(400, "Recipient name and email are required.")

    handover_date = payload.get("handover_date", str(date_type.today()))

    # Set handover fields
    lg.handover_date = handover_date
    lg.handover_notes = payload.get("handover_notes")
    lg.handover_by_user_id = current_user.user_id
    lg.recipient_name = recipient_name
    lg.recipient_email = recipient_email
    lg.recipient_department = payload.get("recipient_department")
    lg.recipient_job_title = payload.get("recipient_job_title")
    lg.recipient_phone = payload.get("recipient_phone")
    lg.recipient_employee_id = payload.get("recipient_employee_id")
    lg.recipient_manager_email = payload.get("recipient_manager_email")
    lg.recipient_second_line_manager_email = payload.get("recipient_second_line_manager_email")

    # Upload signed receiving copy if provided
    if signed_copy_file:
        try:
            from app.core.ai_integration import storage_client
            import uuid
            bucket_name = os.environ.get("GCS_BUCKET_NAME", "lg_custody_bucket")
            bucket = storage_client.bucket(bucket_name)
            file_ext = signed_copy_file.filename.rsplit(".", 1)[-1] if "." in signed_copy_file.filename else "pdf"
            gcs_path = f"handover_signed_copies/{lg.lg_ref_number}_{uuid.uuid4().hex[:8]}.{file_ext}"
            blob = bucket.blob(gcs_path)
            file_content = await signed_copy_file.read()
            blob.upload_from_string(file_content, content_type=signed_copy_file.content_type)
            lg.handover_signed_copy_path = f"gs://{bucket_name}/{gcs_path}"
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"Failed to upload signed copy: {e}")

    lg.status = "HANDED_OVER"
    db.commit()

    # --- Send email notifications ---
    orig_request = db.query(IssuanceRequest).get(lg.request_id) if lg.request_id else None
    email_settings = get_global_email_settings()
    frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")

    # Collect all email recipients
    email_recipients = set()
    email_recipients.add(recipient_email)  # Always notify recipient
    if orig_request:
        email_recipients.add(orig_request.requestor_email)
        if orig_request.manager_email:
            email_recipients.add(orig_request.manager_email)
    if lg.recipient_manager_email:
        email_recipients.add(lg.recipient_manager_email)
    # Add corporate admins
    from app.models import User
    from app.constants import UserRole
    admins = db.query(User).filter(
        User.customer_id == current_user.customer_id,
        User.role == UserRole.CORPORATE_ADMIN,
        User.is_deleted == False
    ).all()
    for admin in admins:
        email_recipients.add(admin.email)
    # Add delivering user
    delivering_user = db.query(User).get(current_user.user_id)
    if delivering_user:
        email_recipients.add(delivering_user.email)

    email_recipients = [e for e in email_recipients if e]  # Remove None/empty

    subject = f"📦 LG Handover Confirmed — {lg.lg_ref_number}"
    body = f"""
    <html>
    <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
        <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
            <h2 style="color: #059669; margin-top: 0;">📦 LG Handover Confirmed</h2>
            <div style="background: #f0fdf4; border-left: 4px solid #059669; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <table style="width: 100%; border-collapse: collapse;">
                    <tr><td style="padding: 4px 0; color: #666;">LG Ref:</td><td style="padding: 4px 0; font-weight: bold;">{lg.lg_ref_number}</td></tr>
                    <tr><td style="padding: 4px 0; color: #666;">Bank LG #:</td><td style="padding: 4px 0;">{lg.bank_lg_number or 'N/A'}</td></tr>
                    <tr><td style="padding: 4px 0; color: #666;">Delivered to:</td><td style="padding: 4px 0; font-weight: bold;">{recipient_name}</td></tr>
                    <tr><td style="padding: 4px 0; color: #666;">Recipient Email:</td><td style="padding: 4px 0;">{recipient_email}</td></tr>
                    <tr><td style="padding: 4px 0; color: #666;">Handover Date:</td><td style="padding: 4px 0;">{handover_date}</td></tr>
                </table>
            </div>
            <p>The Letter of Guarantee has been handed over to the recipient above.</p>
            <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
            <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
        </div>
    </body>
    </html>
    """

    try:
        from starlette.background import BackgroundTasks
        bg_tasks = BackgroundTasks()
        bg_tasks.add_task(send_email, db, email_recipients, subject, body, {}, email_settings)
    except Exception:
        pass  # Non-blocking email failure

    return {"message": "LG handover recorded successfully.", "status": lg.status}


# --- Helper: Send requestor status notification ---
def _send_requestor_status_notification(db, background_tasks, request, event_type, lg=None):
    """Send email to requestor about status change. Covers both approval and post-issuance events."""
    from app.core.email_service import send_email, get_global_email_settings
    import os

    email_settings = get_global_email_settings()
    frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")

    # Bank reply / post-issuance events (require lg)
    bank_events = {}
    if lg:
        bank_events = {
            "LG_ISSUED": {"emoji": "✅", "title": "Your LG Has Been Issued!", "color": "#10b981",
                           "body": f"The bank has issued your Letter of Guarantee.<br><strong>LG Number:</strong> {lg.bank_lg_number or 'Pending'}"},
            "INQUIRY": {"emoji": "❓", "title": "Bank Inquiry on Your LG Request", "color": "#f59e0b",
                         "body": f"The bank has requested additional information.<br><strong>Details:</strong> {lg.bank_reply_notes or 'Please contact your treasury team.'}"},
            "REJECTED": {"emoji": "❌", "title": "LG Issuance Request Declined", "color": "#ef4444",
                          "body": f"Unfortunately, the bank has declined this issuance request.<br><strong>Reason:</strong> {lg.bank_reply_notes or 'No reason provided.'}"},
            "NO_RESPONSE": {"emoji": "⏰", "title": "Bank SLA Exceeded", "color": "#6b7280",
                             "body": "The bank has not responded within the expected timeframe. Your treasury team is following up."},
            "VERIFIED": {"emoji": "🎉", "title": "LG Verified & Confirmed", "color": "#059669",
                          "body": f"Your Letter of Guarantee has been verified and confirmed.<br><strong>LG Number:</strong> {lg.bank_lg_number or 'N/A'}"},
        }

    # Approval lifecycle events (no lg needed)
    approval_events = {
        "APPROVED_STEP": {"emoji": "👍", "title": "Your LG Request Advanced", "color": "#3b82f6",
                           "body": "Your issuance request has passed an approval step and is moving to the next reviewer."},
        "APPROVED_INTERNAL": {"emoji": "✅", "title": "Your LG Request Has Been Fully Approved!", "color": "#10b981",
                               "body": "Great news! Your issuance request has been fully approved and is now ready for issuance execution by the treasury team."},
        "REQUEST_REJECTED": {"emoji": "❌", "title": "Your LG Request Has Been Rejected", "color": "#ef4444",
                              "body": f"Your issuance request has been rejected by an approver.<br><strong>Notes:</strong> {getattr(request, 'revision_notes', '') or 'Please contact the treasury team for details.'}"},
        "RETURNED_FOR_REVISION": {"emoji": "🔄", "title": "Your LG Request Needs Revision", "color": "#f59e0b",
                                   "body": f"An approver has returned your request for revision. Please review and re-submit.<br><strong>Notes:</strong> {getattr(request, 'revision_notes', '') or 'No specific notes provided.'}"},
    }

    all_events = {**bank_events, **approval_events}
    status_info = all_events.get(event_type)

    if not status_info:
        return

    subject = f"{status_info['emoji']} {status_info['title']} — {request.serial_number}"
    body = f"""
    <html>
    <body style="font-family: 'Segoe UI', sans-serif; color: #333; background-color: #f5f5f5; padding: 20px;">
        <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
            <h2 style="color: {status_info['color']}; margin-top: 0;">{status_info['emoji']} {status_info['title']}</h2>
            <div style="background: #f8fafc; border-left: 4px solid {status_info['color']}; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <table style="width: 100%; border-collapse: collapse;">
                    <tr><td style="padding: 4px 0; color: #666;">Request:</td><td style="padding: 4px 0; font-weight: bold;">{request.serial_number}</td></tr>
                    <tr><td style="padding: 4px 0; color: #666;">Beneficiary:</td><td style="padding: 4px 0;">{request.beneficiary_name}</td></tr>
                    <tr><td style="padding: 4px 0; color: #666;">Amount:</td><td style="padding: 4px 0; font-weight: bold;">{request.amount}</td></tr>
                </table>
            </div>
            <p>{status_info['body']}</p>
            <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;" />
            <p style="font-size: 12px; color: #999;">This is an automated notification from your Treasury LG Issuance system.</p>
        </div>
    </body>
    </html>
    """
    background_tasks.add_task(
        send_email, db, [request.requestor_email], subject, body, {}, email_settings
    )


# ==============================================================================
# 9. AI-POWERED LG COPY EXTRACTION
# ==============================================================================

@router.post("/lg-records/{lg_id}/extract-lg-copy")
async def extract_lg_copy(
    lg_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """
    Upload a scanned LG copy (PDF or image). AI extracts LG number, amount,
    dates, beneficiary, etc. Returns extracted fields + comparison with original request.
    """
    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id,
    ).first()
    if not lg:
        raise HTTPException(404, "Issued LG record not found")

    # Validate file
    allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/tiff", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(400, f"Unsupported file type: {file.content_type}. Accepted: PDF, JPEG, PNG, TIFF, WebP")

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:  # 10 MB limit
        raise HTTPException(400, "File too large (max 10 MB)")

    # Run AI extraction
    from app.core.ai_integration import process_lg_document_with_ai
    extracted_data, usage_meta = await process_lg_document_with_ai(
        file_bytes=file_bytes,
        mime_type=file.content_type,
        lg_number_hint=lg.lg_ref_number or f"lg_{lg_id}",
        db=db,
        current_user=current_user,
        file_name=file.filename or "lg_copy_scan",
    )

    if not extracted_data:
        raise HTTPException(422, "AI could not extract data from the uploaded document. Please ensure the scan is clear and readable.")

    # Save the uploaded LG copy as a document so admin can view it during review
    if lg.request_id:
        try:
            import os
            from app.models.models_issuance import IssuanceRequestDocument
            upload_dir = os.path.join("uploads", "issuance", str(current_user.customer_id), str(lg.request_id))
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = f"lg_copy_{lg_id}_{file.filename or 'scan'}"
            file_path = os.path.join(upload_dir, safe_name)
            with open(file_path, "wb") as f:
                f.write(file_bytes)
            doc = IssuanceRequestDocument(
                request_id=lg.request_id,
                document_type="BANK_LG_COPY",
                file_name=file.filename or "LG Copy Scan",
                file_path=file_path,
                uploaded_by=current_user.user_id,
            )
            db.add(doc)
            db.commit()
        except Exception as doc_err:
            import logging
            logging.getLogger(__name__).warning(f"Could not save LG copy document: {doc_err}")

    # Get original request for comparison
    request_obj = None
    if lg.request_id:
        request_obj = db.query(IssuanceRequest).filter(
            IssuanceRequest.id == lg.request_id,
        ).first()

    # Build comparison
    from difflib import SequenceMatcher

    comparison = {"fields": [], "has_discrepancy": False}
    request_amount = float(request_obj.amount) if request_obj and request_obj.amount else None
    request_expiry = str(request_obj.requested_expiry_date) if request_obj and request_obj.requested_expiry_date else None
    request_beneficiary = request_obj.beneficiary_name if request_obj else None

    extracted_amount = extracted_data.get("lgAmount")
    extracted_expiry = extracted_data.get("expiryDate")
    extracted_beneficiary = extracted_data.get("beneficiaryName")

    # Resolve FK fields for comparison
    request_currency = None
    request_lg_type = None
    request_address = request_obj.beneficiary_address if request_obj else None
    request_purpose = request_obj.lg_purpose if request_obj else None
    request_op_status = request_obj.operational_status if request_obj else None

    if request_obj:
        # Resolve currency from currency_id
        if request_obj.currency_id:
            from app.models import Currency
            currency_obj = db.query(Currency).get(request_obj.currency_id)
            request_currency = currency_obj.iso_code if currency_obj else None
        # Resolve LG type from lg_type_id
        if request_obj.lg_type_id:
            from app.models.models_issuance import IssuanceFacility  # lg_types table
            from sqlalchemy import text as sa_text
            lg_type_row = db.execute(sa_text("SELECT name FROM lg_types WHERE id = :id"), {"id": request_obj.lg_type_id}).first()
            request_lg_type = lg_type_row[0] if lg_type_row else None

    extracted_currency = extracted_data.get("currency", "")
    extracted_lg_type = extracted_data.get("lgType", "")
    extracted_address = extracted_data.get("beneficiaryAddress", "")
    extracted_purpose = extracted_data.get("purpose", "")
    extracted_op_status = extracted_data.get("operationalStatus", "")

    def _compare(label, requested, extracted_val, severity="HIGH"):
        match = True
        if requested is not None and extracted_val is not None:
            if isinstance(requested, (int, float)):
                match = abs(float(requested) - float(extracted_val)) < 0.01
            else:
                match = str(requested).strip().lower() == str(extracted_val).strip().lower()
        return {
            "field": label,
            "requested": str(requested) if requested else None,
            "extracted": str(extracted_val) if extracted_val else None,
            "match": match,
            "severity": severity,
        }

    def _fuzzy_compare(label, requested, extracted_val, threshold=0.90, severity="MEDIUM"):
        """Compare using fuzzy matching with SequenceMatcher. Match passes if ratio >= threshold."""
        match = True
        ratio = None
        if requested and extracted_val:
            ratio = SequenceMatcher(None, str(requested).strip().lower(), str(extracted_val).strip().lower()).ratio()
            match = ratio >= threshold
        return {
            "field": label,
            "requested": str(requested) if requested else None,
            "extracted": str(extracted_val) if extracted_val else None,
            "match": match,
            "match_pct": round(ratio * 100, 1) if ratio is not None else None,
            "severity": severity,
        }

    # Core comparisons (existing)
    comparison["fields"].append(_compare("Amount", request_amount, extracted_amount, "HIGH"))
    comparison["fields"].append(_compare("Expiry Date", request_expiry, extracted_expiry, "HIGH"))
    comparison["fields"].append(_compare("Beneficiary Name", request_beneficiary, extracted_beneficiary, "HIGH"))

    # New comparisons
    comparison["fields"].append(_compare("Currency", request_currency, extracted_currency, "HIGH"))
    comparison["fields"].append(_compare("LG Type", request_lg_type, extracted_lg_type, "HIGH"))
    comparison["fields"].append(_fuzzy_compare("Beneficiary Address", request_address, extracted_address, 0.90, "MEDIUM"))
    comparison["fields"].append(_fuzzy_compare("Purpose / Description", request_purpose, extracted_purpose, 0.90, "MEDIUM"))

    # Operational status — only compare for advance payment LGs
    if request_lg_type and "advance" in (request_lg_type or "").lower():
        comparison["fields"].append(_compare("Operational Status", request_op_status, extracted_op_status, "HIGH"))

    comparison["has_discrepancy"] = any(not f["match"] for f in comparison["fields"])

    return {
        "extracted": {
            "bank_lg_number": extracted_data.get("lgNumber", ""),
            "bank_lg_amount": extracted_amount,
            "bank_lg_issue_date": extracted_data.get("issuanceDate", ""),
            "bank_lg_expiry_date": extracted_expiry,
            "bank_beneficiary_name": extracted_beneficiary,
            "issuing_bank_name": extracted_data.get("issuingBankName", ""),
            "currency": extracted_currency,
            "lg_type": extracted_lg_type,
            "purpose": extracted_purpose,
            "beneficiary_address": extracted_address,
            "operational_status": extracted_op_status,
        },
        "comparison": comparison,
        "raw_extracted": extracted_data,
        "usage": usage_meta,
    }


# ==============================================================================
# 10. CORRECTION REQUEST LETTER
# ==============================================================================

@router.post("/lg-records/{lg_id}/generate-correction-letter")
async def generate_correction_letter(
    lg_id: int,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """
    Generate a formal correction request letter to the bank listing discrepancies
    found during LG verification. Returns a downloadable PDF.
    """
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.models import User, Customer
    from app.crud.base import log_action
    from datetime import date as date_type
    from starlette.responses import StreamingResponse
    import io

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id,
    ).first()
    if not lg:
        raise HTTPException(404, "Issued LG record not found.")

    payload = await request.json()
    discrepancies = payload.get("discrepancies", [])
    if not discrepancies:
        raise HTTPException(400, "No discrepancies provided.")

    # Get context
    request_obj = db.query(IssuanceRequest).get(lg.request_id) if lg.request_id else None
    customer = db.query(Customer).get(current_user.customer_id)
    requesting_user = db.query(User).get(current_user.user_id)

    # Get entity name
    entity_name = customer.name if customer else "Our Company"
    if request_obj and request_obj.issuing_entity_id:
        from app.models import CustomerEntity
        entity = db.query(CustomerEntity).get(request_obj.issuing_entity_id)
        if entity:
            entity_name = entity.entity_name

    # Get bank name
    bank_name = "The Bank"
    if request_obj:
        from sqlalchemy import text as sa_text
        if hasattr(request_obj, 'selected_sub_limit_id') and request_obj.selected_sub_limit_id:
            from app.models.models_issuance import IssuanceFacilitySubLimit, IssuanceFacility
            sub = db.query(IssuanceFacilitySubLimit).get(request_obj.selected_sub_limit_id)
            if sub:
                fac = db.query(IssuanceFacility).get(sub.facility_id)
                if fac and fac.bank_id:
                    from app.models import Bank
                    bank = db.query(Bank).get(fac.bank_id)
                    if bank:
                        bank_name = bank.name

    today_str = date_type.today().strftime("%B %d, %Y")

    # Build discrepancy rows
    disc_rows = ""
    for d in discrepancies:
        severity_badge = "&#128308;" if d.get("severity") == "HIGH" else "&#128992;"
        match_info = ""
        if d.get("match_pct") is not None:
            match_info = f' <span style="color:#888;font-size:11px;">({d["match_pct"]}% match)</span>'
        disc_rows += f"""
        <tr>
            <td style="padding:10px 12px;border-bottom:1px solid #e5e7eb;font-weight:600;">{severity_badge} {d.get("field", "")}</td>
            <td style="padding:10px 12px;border-bottom:1px solid #e5e7eb;color:#059669;">{d.get("requested", "N/A")}</td>
            <td style="padding:10px 12px;border-bottom:1px solid #e5e7eb;color:#dc2626;">{d.get("extracted", "N/A")}{match_info}</td>
        </tr>"""

    user_fullname = (request_obj.requestor_name if request_obj and request_obj.requestor_name
                      else requesting_user.email if requesting_user
                      else "Authorized Signatory")

    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4; margin: 2.5cm; }}
            body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #1e293b; line-height: 1.7; font-size: 13px; }}
            .header {{ border-bottom: 3px solid #1e40af; padding-bottom: 15px; margin-bottom: 30px; }}
            .header h1 {{ color: #1e40af; font-size: 18px; margin: 0; }}
            .header p {{ color: #64748b; font-size: 12px; margin: 3px 0; }}
            .meta-block {{ font-size: 12px; color: #475569; margin-bottom: 20px; }}
            .meta-block strong {{ color: #1e293b; }}
            h2 {{ color: #1e40af; font-size: 15px; border-bottom: 1px solid #e5e7eb; padding-bottom: 8px; }}
            table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 12px; }}
            th {{ background: #f1f5f9; padding: 10px 12px; text-align: left; font-weight: 700; color: #334155; border-bottom: 2px solid #cbd5e1; }}
            .ref {{ background: #eff6ff; padding: 12px 16px; border-radius: 8px; border-left: 4px solid #3b82f6; margin: 20px 0; }}
            .footer {{ margin-top: 50px; padding-top: 15px; border-top: 1px solid #e5e7eb; font-size: 11px; color: #94a3b8; }}
            .signature {{ margin-top: 60px; }}
            .signature .line {{ border-top: 1px solid #1e293b; width: 200px; margin-top: 40px; padding-top: 5px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>{entity_name}</h1>
            <p>Letter of Guarantee &mdash; Correction Request</p>
        </div>

        <div class="meta-block">
            <strong>To:</strong> {bank_name}<br>
            <strong>Date:</strong> {today_str}<br>
            <strong>LG Ref:</strong> {lg.lg_ref_number or "N/A"} &nbsp;|&nbsp;
            <strong>Bank LG No:</strong> {lg.bank_lg_number or "N/A"} &nbsp;|&nbsp;
            <strong>Request:</strong> {request_obj.serial_number if request_obj else "N/A"}
        </div>

        <div class="ref">
            <strong>Subject:</strong> Request for Correction of Letter of Guarantee &mdash; {lg.lg_ref_number or "N/A"}
        </div>

        <p>Dear Sir/Madam,</p>
        <p>
            We refer to the above-mentioned Letter of Guarantee issued by your esteemed bank.
            Upon reviewing the issued LG document, we have identified the following discrepancies
            between our original request and the LG as issued:
        </p>

        <h2>Discrepancy Details</h2>
        <table>
            <thead>
                <tr>
                    <th>Field</th>
                    <th>As Requested</th>
                    <th>As Issued</th>
                </tr>
            </thead>
            <tbody>
                {disc_rows}
            </tbody>
        </table>

        <p>
            We kindly request that you review the above discrepancies and issue a corrected
            Letter of Guarantee at your earliest convenience to reflect the originally requested terms.
        </p>
        <p>
            Please do not hesitate to contact us should you require any clarification or additional documentation.
        </p>

        <p>Thank you for your prompt attention to this matter.</p>

        <div class="signature">
            <p>Yours faithfully,</p>
            <div class="line">
                <strong>{user_fullname}</strong><br>
                <span style="color:#64748b;font-size:12px;">{entity_name}</span>
            </div>
        </div>

        <div class="footer">
            Generated by Treasury Management System on {today_str}. Document reference: CORR-{lg.lg_ref_number or lg_id}
        </div>
    </body>
    </html>
    """

    # Generate PDF
    try:
        from weasyprint import HTML as WeasyHTML
        pdf_bytes = WeasyHTML(string=html).write_pdf()
    except ImportError:
        # Fallback: return HTML directly if weasyprint not available
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html",
            headers={"Content-Disposition": f'inline; filename="correction_request_{lg.lg_ref_number or lg_id}.html"'}
        )

    # Log the action
    log_action(db, current_user.user_id, "ISSUANCE_CORRECTION_REQUESTED", "IssuedLGRecord", lg.id,
               {"discrepancies": discrepancies, "bank": bank_name},
               current_user.customer_id)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="correction_request_{lg.lg_ref_number or lg_id}.pdf"'}
    )


# ==============================================================================
# 20. ISSUANCE LG MAINTENANCE ACTIONS
# ==============================================================================

from app.services.issuance_maintenance_service import maintenance_service
from app.models.models_issuance import IssuanceMaintenanceAction
from pydantic import BaseModel as PydanticBase
from typing import Optional as Opt

class MaintenanceActionCreate(PydanticBase):
    action_type: str  # EXTEND, INCREASE_AMOUNT, CLOSE, LIQUIDATION, AMENDMENT, ACTIVATE
    action_data: dict  # {new_expiry_date, new_amount, amendment_text, ...}
    notes: Opt[str] = None

class DeliveryRecord(PydanticBase):
    delivery_method: str  # HAND_DELIVERY, COURIER, EMAIL
    delivery_notes: Opt[str] = None

class BankReplyRecord(PydanticBase):
    bank_reply_notes: Opt[str] = None

class RejectPayload(PydanticBase):
    reason: Opt[str] = None


@router.post("/issued-lgs/{issued_lg_id}/maintenance")
def create_maintenance_action(
    issued_lg_id: int,
    payload: MaintenanceActionCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(check_subscription_status),
):
    """Create a maintenance action on an issued LG.
    
    Open to: Treasury end users, Corporate admins, and the original requestor
    of the LG (verified via the linked IssuanceRequest.requestor_user_id).
    """
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.constants import UserRole

    # Verify the LG exists and belongs to the customer
    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == issued_lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id,
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="Issued LG record not found.")

    # Authorization: only end users (treasury officers) or the original requestor can create maintenance actions
    # Corporate admins supervise and approve — they do not initiate maintenance
    is_treasury = current_user.role in (UserRole.END_USER, UserRole.END_USER.value)
    is_requestor = False
    if lg.request_id:
        orig_request = db.query(IssuanceRequest).filter(
            IssuanceRequest.id == lg.request_id,
        ).first()
        if orig_request and orig_request.requestor_user_id == current_user.user_id:
            is_requestor = True

    if not is_treasury and not is_requestor:
        raise HTTPException(
            status_code=403,
            detail="Only treasury users, corporate admins, or the original requestor can raise maintenance actions."
        )
    action = maintenance_service.create_action(
        db, issued_lg_id, payload.action_type, payload.action_data,
        current_user.user_id, current_user.customer_id, payload.notes,
        initiation_source="INTERNAL_USER"
    )

    # --- Email + In-App Notification ---
    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(issued_lg_id)

    if action.status == "PENDING_APPROVAL" and action.pending_approver_users:
        from app.core.email_service import send_email, get_customer_email_settings
        from app.services.issuance_notifications import _get_user_emails, _base_url
        from app.services.notification_service import notify

        email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
        approver_ids = [int(uid) for uid in action.pending_approver_users]
        approver_emails = _get_user_emails(db, approver_ids)

        if approver_emails:
            subject = f"ACTION REQUIRED: LG {ref} — {payload.action_type} Request"
            body = f"""
            <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #1a56db;">🔔 LG Maintenance Request</h2>
                <p>A <strong>{payload.action_type}</strong> action on LG <strong>{ref}</strong> requires your approval.</p>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{_base_url()}/corporate-admin/approval-inbox" style="padding: 12px 30px; background: #1a56db; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">Review Request</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee;" />
                <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
            </div></body></html>
            """
            background_tasks.add_task(send_email, db, approver_emails, subject, body, {}, email_settings)

        # In-App Notification
        notify(
            db, user_ids=approver_ids, module="ISSUANCE",
            event_type=f"MAINTENANCE_{payload.action_type}_PENDING",
            title=f"LG {ref} — {payload.action_type} Pending",
            message=f"A {payload.action_type.lower().replace('_', ' ')} action on LG {ref} requires your approval.",
            link="/corporate-admin/approval-inbox",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

    return _serialize_action(action, lg)


@router.get("/issued-lgs/{issued_lg_id}/maintenance")
def list_maintenance_actions(
    issued_lg_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """List all maintenance actions for an issued LG."""
    actions = db.query(IssuanceMaintenanceAction).filter(
        IssuanceMaintenanceAction.issued_lg_id == issued_lg_id
    ).order_by(IssuanceMaintenanceAction.created_at.desc()).all()

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == issued_lg_id).first()
    return [_serialize_action(a, lg) for a in actions]


@router.get("/issued-lgs/{issued_lg_id}/documents")
def list_issued_lg_documents(
    issued_lg_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """Aggregate all documents related to an issued LG from multiple sources."""
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest, IssuanceRequestDocument, IssuanceMaintenanceAction

    lg = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.id == issued_lg_id,
        IssuedLGRecord.customer_id == current_user.customer_id,
    ).first()
    if not lg:
        raise HTTPException(status_code=404, detail="Issued LG not found")

    documents = []

    # Source 1: Request documents (contracts, special wording, formal requests)
    if lg.request_id:
        req_docs = db.query(IssuanceRequestDocument).filter(
            IssuanceRequestDocument.request_id == lg.request_id
        ).all()
        for doc in req_docs:
            documents.append({
                "id": f"req-{doc.id}",
                "document_id": doc.id,
                "request_id": lg.request_id,
                "file_name": doc.file_name,
                "document_type": doc.document_type,
                "source": "Request",
                "download_type": "request_doc",
                "created_at": doc.created_at.isoformat() if doc.created_at else None,
            })

    # Source 2: LG soft copy (scanned original)
    if lg.soft_copy_path:
        documents.append({
            "id": f"lg-soft-{lg.id}",
            "lg_id": lg.id,
            "file_name": lg.soft_copy_path.split("/")[-1] if "/" in (lg.soft_copy_path or "") else "LG Soft Copy",
            "document_type": "LG_SOFT_COPY",
            "source": "LG Record",
            "download_type": "lg_reprint",
            "created_at": lg.created_at.isoformat() if lg.created_at else None,
        })

    # Source 3: Maintenance action letters (generated PDFs)
    maint_actions = db.query(IssuanceMaintenanceAction).filter(
        IssuanceMaintenanceAction.issued_lg_id == issued_lg_id,
        IssuanceMaintenanceAction.letter_generated_path.isnot(None),
    ).all()
    for ma in maint_actions:
        path = ma.letter_generated_path
        documents.append({
            "id": f"maint-{ma.id}",
            "action_id": ma.id,
            "file_name": path.split("/")[-1] if "/" in (path or "") else f"{ma.action_type} Letter",
            "document_type": f"MAINTENANCE_{ma.action_type}",
            "source": f"Maintenance — {ma.action_type.replace('_', ' ').title()}",
            "download_type": "maintenance_letter",
            "serial": ma.letter_serial_number,
            "created_at": ma.created_at.isoformat() if ma.created_at else None,
        })

    return documents


@router.post("/maintenance/{action_id}/approve")
def approve_maintenance_action(
    action_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context),
):
    """Approve a maintenance action step."""
    action = maintenance_service.approve_action(db, action_id, current_user.user_id, current_user.customer_id)

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(action.issued_lg_id)

    # Notify initiator
    from app.services.notification_service import notify
    if action.initiated_by_user_id:
        status_label = "Fully Approved ✅" if action.status in ("APPROVED", "EXECUTED") else "Step Approved"
        notify(
            db, user_ids=[action.initiated_by_user_id], module="ISSUANCE",
            event_type=f"MAINTENANCE_{action.action_type}_APPROVED",
            title=f"LG {ref} — {action.action_type} {status_label}",
            message=f"Your {action.action_type.lower().replace('_', ' ')} request has been {status_label.lower()}.",
            link=f"/corporate-admin/issuance/issued-lgs",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

        # Email notification
        from app.core.email_service import send_email, get_customer_email_settings
        from app.services.issuance_notifications import _get_user_emails, _base_url, get_common_communication_emails
        email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
        initiator_emails = _get_user_emails(db, [action.initiated_by_user_id])
        cc_emails = get_common_communication_emails(db, current_user.customer_id)
        if initiator_emails:
            subject = f"LG {ref} — {action.action_type.replace('_', ' ')} {status_label}"
            body = f"""
            <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #16a34a;">✅ Maintenance Action Approved</h2>
                <p>Your <strong>{action.action_type.replace('_', ' ')}</strong> request on LG <strong>{ref}</strong> has been <strong>{status_label.lower()}</strong>.</p>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{_base_url()}/corporate-admin/issuance/issued-lgs" style="padding: 12px 30px; background: #16a34a; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">View LG</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee;" />
                <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
            </div></body></html>
            """
            background_tasks.add_task(send_email, db, initiator_emails, subject, body, {}, email_settings, cc_emails=cc_emails)

    return _serialize_action(action, lg)


@router.post("/maintenance/{action_id}/reject")
def reject_maintenance_action(
    action_id: int,
    payload: RejectPayload,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context),
):
    """Reject a maintenance action."""
    action = maintenance_service.reject_action(db, action_id, current_user.user_id, current_user.customer_id, payload.reason)

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(action.issued_lg_id)

    # Notify initiator
    from app.services.notification_service import notify
    if action.initiated_by_user_id:
        notify(
            db, user_ids=[action.initiated_by_user_id], module="ISSUANCE",
            event_type=f"MAINTENANCE_{action.action_type}_REJECTED",
            title=f"LG {ref} — {action.action_type} Rejected ❌",
            message=f"Your {action.action_type.lower().replace('_', ' ')} request was rejected. Reason: {payload.reason or 'Not specified'}",
            link=f"/corporate-admin/issuance/issued-lgs",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

        # Email notification
        from app.core.email_service import send_email, get_customer_email_settings
        from app.services.issuance_notifications import _get_user_emails, _base_url, get_common_communication_emails
        email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
        initiator_emails = _get_user_emails(db, [action.initiated_by_user_id])
        cc_emails = get_common_communication_emails(db, current_user.customer_id)
        if initiator_emails:
            reason_text = payload.reason or "Not specified"
            subject = f"LG {ref} — {action.action_type.replace('_', ' ')} Rejected ❌"
            body = f"""
            <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #dc2626;">❌ Maintenance Action Rejected</h2>
                <p>Your <strong>{action.action_type.replace('_', ' ')}</strong> request on LG <strong>{ref}</strong> has been rejected.</p>
                <div style="background: #fef2f2; border-left: 4px solid #dc2626; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <p style="margin: 0;"><strong>Reason:</strong> {reason_text}</p>
                </div>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{_base_url()}/corporate-admin/issuance/issued-lgs" style="padding: 12px 30px; background: #dc2626; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">View LG</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee;" />
                <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
            </div></body></html>
            """
            background_tasks.add_task(send_email, db, initiator_emails, subject, body, {}, email_settings, cc_emails=cc_emails)

    return _serialize_action(action, lg)


@router.post("/maintenance/{action_id}/delivery")
def record_maintenance_delivery(
    action_id: int,
    delivery_method: str = Form(...),
    delivery_notes: Opt[str] = Form(None),
    delivery_date: Opt[str] = Form(None),
    delivery_document: Opt[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Record letter delivery to bank — with date and optional proof document."""
    doc_bytes = None
    doc_mime = None
    if delivery_document and delivery_document.filename:
        doc_bytes = delivery_document.file.read()
        doc_mime = delivery_document.content_type

    action = maintenance_service.record_delivery(
        db, action_id, current_user.user_id,
        delivery_method, delivery_notes,
        delivery_date_str=delivery_date,
        delivery_document_bytes=doc_bytes,
        delivery_document_mime_type=doc_mime,
        customer_id=current_user.customer_id,
    )
    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    return _serialize_action(action, lg)


@router.post("/maintenance/{action_id}/bank-reply")
def record_maintenance_bank_reply(
    action_id: int,
    background_tasks: BackgroundTasks,
    bank_reply_notes: Opt[str] = Form(None),
    bank_reply_date: Opt[str] = Form(None),
    bank_reply_file: Opt[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Record bank reply and apply confirmed changes to LG.
    F3: Optionally accepts a bank reply document for AI verification."""
    # F3: Read file bytes if uploaded
    file_bytes = None
    mime_type = None
    if bank_reply_file and bank_reply_file.filename:
        file_bytes = bank_reply_file.file.read()
        mime_type = bank_reply_file.content_type

    action = maintenance_service.record_bank_reply(
        db, action_id, current_user.user_id, current_user.customer_id,
        bank_reply_notes, file_bytes, mime_type,
        bank_reply_date_str=bank_reply_date,
    )

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(action.issued_lg_id)

    # Notify initiator (only when changes were applied, not when awaiting confirmation)
    from app.services.notification_service import notify
    if action.initiated_by_user_id and action.instruction_status == "Confirmed by Bank":
        notify(
            db, user_ids=[action.initiated_by_user_id], module="ISSUANCE",
            event_type=f"MAINTENANCE_{action.action_type}_BANK_CONFIRMED",
            title=f"LG {ref} — Bank Confirmed {action.action_type}",
            message=f"The bank has confirmed the {action.action_type.lower().replace('_', ' ')} on LG {ref}.",
            link=f"/corporate-admin/issuance/issued-lgs",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

        # Email notification — bank confirmed
        from app.core.email_service import send_email, get_customer_email_settings
        from app.services.issuance_notifications import _get_user_emails, _base_url, get_common_communication_emails
        email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
        to_emails = _get_user_emails(db, [action.initiated_by_user_id])
        cc_emails = get_common_communication_emails(db, current_user.customer_id)

        is_liquidation = action.action_type in ("LIQUIDATION",)
        color = "#dc2626" if is_liquidation else "#16a34a"
        icon = "🚨" if is_liquidation else "🏦"
        alert = "HIGH ALERT: " if is_liquidation else ""

        if to_emails:
            subject = f"{alert}LG {ref} — Bank Confirmed {action.action_type.replace('_', ' ')}"
            body = f"""
            <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);{' border: 2px solid #dc2626;' if is_liquidation else ''}">
                <h2 style="color: {color};">{icon} Bank Reply Confirmed — {action.action_type.replace('_', ' ')}</h2>
                <p>The bank has confirmed the <strong>{action.action_type.replace('_', ' ')}</strong> action on LG <strong>{ref}</strong>. Changes have been applied.</p>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{_base_url()}/corporate-admin/issuance/issued-lgs" style="padding: 12px 30px; background: {color}; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">View LG</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee;" />
                <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
            </div></body></html>
            """
            background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings, cc_emails=cc_emails)

    return _serialize_action(action, lg)


@router.post("/maintenance/{action_id}/confirm-bank-reply")
def confirm_bank_reply(
    action_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Phase 2: User reviewed AI verification and chose to proceed.
    Applies the previously paused changes to the LG record."""
    action = maintenance_service.confirm_bank_reply(
        db, action_id, current_user.user_id, current_user.customer_id
    )

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(action.issued_lg_id)

    # Notify initiator that bank confirmed (Phase 2)
    from app.services.notification_service import notify
    if action.initiated_by_user_id:
        notify(
            db, user_ids=[action.initiated_by_user_id], module="ISSUANCE",
            event_type=f"MAINTENANCE_{action.action_type}_BANK_CONFIRMED",
            title=f"LG {ref} — Bank Confirmed {action.action_type}",
            message=f"The bank has confirmed the {action.action_type.lower().replace('_', ' ')} on LG {ref}. Changes applied.",
            link=f"/corporate-admin/issuance/issued-lgs",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

        # Email notification — bank confirmed after user review
        from app.core.email_service import send_email, get_customer_email_settings
        from app.services.issuance_notifications import _get_user_emails, _base_url, get_common_communication_emails
        email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
        to_emails = _get_user_emails(db, [action.initiated_by_user_id])
        cc_emails = get_common_communication_emails(db, current_user.customer_id)

        is_liquidation = action.action_type in ("LIQUIDATION",)
        color = "#dc2626" if is_liquidation else "#16a34a"
        icon = "🚨" if is_liquidation else "🏦"
        alert = "HIGH ALERT: " if is_liquidation else ""

        if to_emails:
            subject = f"{alert}LG {ref} — Bank Confirmed {action.action_type.replace('_', ' ')}"
            body = f"""
            <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);{' border: 2px solid #dc2626;' if is_liquidation else ''}">
                <h2 style="color: {color};">{icon} Bank Reply Confirmed — {action.action_type.replace('_', ' ')}</h2>
                <p>The bank has confirmed the <strong>{action.action_type.replace('_', ' ')}</strong> action on LG <strong>{ref}</strong>. Changes have been applied.</p>
                <p style="color: #b45309; font-weight: bold;">⚠️ Note: AI verification flagged discrepancies, but user approved proceeding.</p>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{_base_url()}/corporate-admin/issuance/issued-lgs" style="padding: 12px 30px; background: {color}; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">View LG</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee;" />
                <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
            </div></body></html>
            """
            background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings, cc_emails=cc_emails)

    return _serialize_action(action, lg)


@router.post("/maintenance/{action_id}/cancel-bank-reply")
def cancel_pending_bank_reply(
    action_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """User reviewed AI verification and chose NOT to proceed.
    Reverts the action so user can re-upload a corrected document."""
    action = maintenance_service.cancel_pending_bank_reply(
        db, action_id, current_user.user_id, current_user.customer_id
    )
    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    return _serialize_action(action, lg)


@router.post("/maintenance/{action_id}/cancel")
def cancel_maintenance_action(
    action_id: int,
    payload: RejectPayload,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Cancel a recently executed maintenance action within the cancellation window.
    Only the most recent action on an LG can be cancelled, and only while instruction_status = 'Instruction Issued'."""
    action = maintenance_service.cancel_action(
        db, action_id, current_user.user_id, current_user.customer_id,
        reason=payload.reason or ""
    )

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(action.issued_lg_id)

    # In-app notification to initiator
    from app.services.notification_service import notify
    if action.initiated_by_user_id:
        notify(
            db, user_ids=[action.initiated_by_user_id], module="ISSUANCE",
            event_type=f"MAINTENANCE_{action.action_type}_CANCELLED",
            title=f"LG {ref} — {action.action_type} Cancelled",
            message=f"The {action.action_type.lower().replace('_', ' ')} action on LG {ref} has been cancelled. Reason: {payload.reason or 'Not specified'}",
            link=f"/corporate-admin/issuance/issued-lgs",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

        # Email notification
        from app.core.email_service import send_email, get_customer_email_settings
        from app.services.issuance_notifications import _get_user_emails, _base_url, get_common_communication_emails
        email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
        to_emails = _get_user_emails(db, [action.initiated_by_user_id])
        cc_emails = get_common_communication_emails(db, current_user.customer_id)
        if to_emails:
            subject = f"LG {ref} — {action.action_type.replace('_', ' ')} Cancelled"
            body = f"""
            <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
            <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                <h2 style="color: #b45309;">⚠️ Maintenance Action Cancelled</h2>
                <p>The <strong>{action.action_type.replace('_', ' ')}</strong> action on LG <strong>{ref}</strong> has been cancelled.</p>
                <div style="background: #fffbeb; border-left: 4px solid #b45309; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <p style="margin: 0;"><strong>Reason:</strong> {payload.reason or 'Not specified'}</p>
                </div>
                <div style="text-align: center; margin: 25px 0;">
                    <a href="{_base_url()}/corporate-admin/issuance/issued-lgs" style="padding: 12px 30px; background: #b45309; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">View LG</a>
                </div>
                <hr style="border: none; border-top: 1px solid #eee;" />
                <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
            </div></body></html>
            """
            background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings, cc_emails=cc_emails)

    return _serialize_action(action, lg)


@router.post("/issued-lgs/{issued_lg_id}/bank-initiated-change")
def process_bank_initiated_change(
    issued_lg_id: int,
    bank_letter: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Upload a bank letter → AI extracts what changed → returns diff for user review.
    Does NOT apply changes yet."""
    file_bytes = bank_letter.file.read()
    mime_type = bank_letter.content_type

    result = maintenance_service.process_bank_initiated_change(
        db, issued_lg_id, file_bytes, mime_type,
        current_user.user_id, current_user.customer_id,
    )
    return result


@router.post("/maintenance/{action_id}/confirm-bank-change")
def confirm_bank_initiated_change(
    action_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """User reviewed AI-detected bank changes and confirms. Applies changes to LG record."""
    action = maintenance_service.confirm_bank_initiated_change(
        db, action_id, current_user.user_id, current_user.customer_id,
    )

    lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == action.issued_lg_id).first()
    ref = lg.lg_ref_number if lg else str(action.issued_lg_id)

    # In-app + Email notification
    from app.services.notification_service import notify
    from app.core.email_service import send_email, get_customer_email_settings
    from app.services.issuance_notifications import _get_user_emails, _base_url, get_common_communication_emails

    # Notify all end users + admins for this customer
    users = db.query(User).filter(
        User.customer_id == current_user.customer_id,
        User.is_deleted == False,
        User.role.in_(["corporate_admin", "end_user"])
    ).all()
    user_ids = [u.id for u in users]

    is_liquidation = action.action_type in ("LIQUIDATION",)
    icon = "🚨" if is_liquidation else "🏦"
    alert = "HIGH ALERT: " if is_liquidation else ""

    if user_ids:
        notify(
            db, user_ids=user_ids, module="ISSUANCE",
            event_type=f"BANK_INITIATED_{action.action_type}",
            title=f"{alert}LG {ref} — Bank-Initiated {action.action_type.replace('_', ' ')}",
            message=f"The bank has initiated a {action.action_type.lower().replace('_', ' ')} on LG {ref}. Changes have been applied.",
            link=f"/corporate-admin/issuance/issued-lgs",
            actor_user_id=current_user.user_id,
            reference_id=action.id
        )

    # Email notification
    email_settings, _ = get_customer_email_settings(db, current_user.customer_id)
    to_emails = [u.email for u in users if u.email]
    cc_emails = get_common_communication_emails(db, current_user.customer_id)

    color = "#dc2626" if is_liquidation else "#7c3aed"
    if to_emails:
        subject = f"{alert}LG {ref} — Bank-Initiated {action.action_type.replace('_', ' ')} Applied"
        body = f"""
        <html><body style="font-family: 'Segoe UI', sans-serif; color: #333; padding: 20px;">
        <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 12px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.08);{' border: 2px solid #dc2626;' if is_liquidation else ''}">
            <h2 style="color: {color};">{icon} Bank-Initiated Change — {action.action_type.replace('_', ' ')}</h2>
            <p>The bank has initiated a <strong>{action.action_type.replace('_', ' ')}</strong> change on LG <strong>{ref}</strong>. Changes have been reviewed and applied.</p>
            <div style="text-align: center; margin: 25px 0;">
                <a href="{_base_url()}/corporate-admin/issuance/issued-lgs" style="padding: 12px 30px; background: {color}; color: #fff; text-decoration: none; border-radius: 8px; font-weight: bold;">View LG</a>
            </div>
            <hr style="border: none; border-top: 1px solid #eee;" />
            <p style="font-size: 12px; color: #999;">Automated notification from Treasury LG Issuance system.</p>
        </div></body></html>
        """
        background_tasks.add_task(send_email, db, to_emails, subject, body, {}, email_settings, cc_emails=cc_emails)

    return _serialize_action(action, lg)


@router.get("/maintenance/pending")
def get_pending_maintenance_actions(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_approver_context),
):
    """Get all maintenance actions pending the current user's approval."""
    from sqlalchemy import cast, String
    actions = db.query(IssuanceMaintenanceAction).filter(
        IssuanceMaintenanceAction.status == "PENDING_APPROVAL"
    ).all()

    # Filter to those where current user is in pending_approver_users
    pending = []
    for a in actions:
        approvers = [int(uid) for uid in (a.pending_approver_users or [])]
        if current_user.user_id in approvers:
            lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == a.issued_lg_id).first()
            pending.append(_serialize_action(a, lg))

    return pending


def _serialize_action(action: IssuanceMaintenanceAction, lg=None) -> dict:
    """Serialize maintenance action to dict for API response."""
    return {
        "id": action.id,
        "issued_lg_id": action.issued_lg_id,
        "lg_ref_number": lg.lg_ref_number if lg else None,
        "lg_beneficiary": lg.beneficiary_name if lg else None,
        "lg_current_amount": str(lg.current_amount) if lg else None,
        "lg_expiry_date": str(lg.expiry_date) if lg else None,
        "lg_status": lg.status if lg else None,
        "action_type": action.action_type,
        "status": action.status,
        "action_data": action.action_data,
        "pending_approver_users": action.pending_approver_users,
        "current_step_number": action.current_step_number,
        "approval_history": action.approval_history,
        "instruction_status": action.instruction_status,
        "letter_serial_number": action.letter_serial_number,
        "is_printed": action.is_printed,
        "delivery_date": str(action.delivery_date) if action.delivery_date else None,
        "delivery_method": action.delivery_method,
        "delivery_notes": action.delivery_notes,
        "delivery_document_path": action.delivery_document_path,
        "bank_reply_date": str(action.bank_reply_date) if action.bank_reply_date else None,
        "bank_reply_notes": action.bank_reply_notes,
        "bank_reply_document_path": action.bank_reply_document_path,
        "initiation_source": action.initiation_source,
        "initiated_by_user_id": action.initiated_by_user_id,
        "executed_by_user_id": action.executed_by_user_id,
        "notes": action.notes,
        "created_at": str(action.created_at) if action.created_at else None,
        "updated_at": str(action.updated_at) if action.updated_at else None,
    }


# ==============================================================================
# MAINTENANCE DOCUMENT ACCESS
# ==============================================================================

@router.get("/maintenance/{action_id}/document/{doc_type}")
async def get_maintenance_document_url(
    action_id: int,
    doc_type: str,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Generate a signed URL for maintenance action documents.
    doc_type: 'delivery', 'bank_reply', or 'bank_initiated'
    """
    action = db.query(IssuanceMaintenanceAction).filter(
        IssuanceMaintenanceAction.id == action_id
    ).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")

    gcs_path = None
    if doc_type == "delivery":
        gcs_path = action.delivery_document_path
    elif doc_type == "bank_reply":
        gcs_path = action.bank_reply_document_path
    elif doc_type == "bank_initiated":
        data = action.action_data or {}
        gcs_path = data.get("bank_document_gcs")
    else:
        raise HTTPException(status_code=400, detail="Invalid document type. Use 'delivery', 'bank_reply', or 'bank_initiated'.")

    if not gcs_path:
        raise HTTPException(status_code=404, detail=f"No {doc_type.replace('_', ' ')} document found for this action")

    from app.core.ai_integration import generate_signed_gcs_url
    signed_url = await generate_signed_gcs_url(gcs_path, expiration=3600)
    if not signed_url:
        raise HTTPException(status_code=500, detail="Could not generate download URL")

    return {"download_url": signed_url, "doc_type": doc_type}


# ==============================================================================
# LG POSITION RECONCILIATION
# ==============================================================================

from app.services.reconciliation_service import reconciliation_service
from app.models.models_issuance import (
    ReconciliationSession as ReconSession,
    ReconciliationBankRow as ReconBankRow,
    ReconciliationResult as ReconResultModel,
    BankColumnMapping,
)


@router.post("/reconciliation/sessions")
async def create_reconciliation_session(
    bank_id: int = Form(...),
    position_date: str = Form(...),
    notes: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Create a reconciliation session and upload the bank position report."""
    from datetime import date as date_type
    try:
        pd = date_type.fromisoformat(position_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid position_date format. Use YYYY-MM-DD.")

    session = reconciliation_service.create_session(
        db, customer_id=current_user.customer_id, bank_id=bank_id,
        position_date=pd, user_id=current_user.user_id,
        file_name=file.filename, notes=notes,
    )

    # Parse the file
    file_bytes = await file.read()
    session = await reconciliation_service.parse_file(
        db, session.id, file_bytes, file.filename,
        customer_id=current_user.customer_id, user_id=current_user.user_id,
    )

    return _serialize_recon_session(session, db)


@router.get("/reconciliation/sessions")
def list_reconciliation_sessions(
    bank_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """List reconciliation sessions for the customer."""
    q = db.query(ReconSession).filter(
        ReconSession.customer_id == current_user.customer_id,
    )
    if bank_id:
        q = q.filter(ReconSession.bank_id == bank_id)
    sessions = q.order_by(ReconSession.created_at.desc()).all()
    return [_serialize_recon_session(s, db, brief=True) for s in sessions]


@router.get("/reconciliation/sessions/{session_id}")
def get_reconciliation_session(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Get detailed reconciliation session."""
    session = db.query(ReconSession).filter(
        ReconSession.id == session_id,
        ReconSession.customer_id == current_user.customer_id,
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return _serialize_recon_session(session, db)


@router.post("/reconciliation/sessions/{session_id}/match")
def run_reconciliation_matching(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Run the matching engine on a parsed session."""
    session = reconciliation_service.run_matching(
        db, session_id,
        customer_id=current_user.customer_id,
        user_id=current_user.user_id,
    )
    return _serialize_recon_session(session, db)


@router.get("/reconciliation/sessions/{session_id}/results")
def get_reconciliation_results(
    session_id: int,
    severity: Optional[str] = Query(None),
    mismatch_type: Optional[str] = Query(None),
    resolved: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Get reconciliation results with optional filters."""
    # Verify session access
    session = db.query(ReconSession).filter(
        ReconSession.id == session_id,
        ReconSession.customer_id == current_user.customer_id,
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    q = db.query(ReconResultModel).filter(
        ReconResultModel.session_id == session_id
    )
    if severity:
        q = q.filter(ReconResultModel.severity == severity)
    if mismatch_type:
        q = q.filter(ReconResultModel.mismatch_type == mismatch_type)
    if resolved is not None:
        if resolved:
            q = q.filter(ReconResultModel.user_resolution.isnot(None))
        else:
            q = q.filter(ReconResultModel.user_resolution.is_(None))

    results = q.order_by(
        # HIGH first, then MEDIUM, LOW, INFO
        func.array_position(func.cast(['HIGH', 'MEDIUM', 'LOW', 'INFO'], type_=None),
                            ReconResultModel.severity) if False else ReconResultModel.id
    ).all()

    # Sort by severity manually
    severity_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "INFO": 3}
    results.sort(key=lambda r: severity_order.get(r.severity, 99))

    return [_serialize_recon_result(r, db) for r in results]


@router.get("/reconciliation/sessions/{session_id}/bank-rows")
def get_reconciliation_bank_rows(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Get all parsed bank rows for a session."""
    session = db.query(ReconSession).filter(
        ReconSession.id == session_id,
        ReconSession.customer_id == current_user.customer_id,
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    rows = db.query(ReconBankRow).filter(
        ReconBankRow.session_id == session_id
    ).all()

    return [{
        "id": r.id,
        "bank_lg_number": r.bank_lg_number,
        "beneficiary_name": r.beneficiary_name,
        "amount": str(r.amount) if r.amount else None,
        "currency_code": r.currency_code,
        "issue_date": str(r.issue_date) if r.issue_date else None,
        "expiry_date": str(r.expiry_date) if r.expiry_date else None,
        "match_status": r.match_status,
        "matched_lg_id": r.matched_lg_id,
        "variances": r.variances,
    } for r in rows]


@router.post("/reconciliation/results/{result_id}/resolve")
def resolve_reconciliation_result(
    result_id: int,
    payload: Dict = Body(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Resolve a reconciliation mismatch: ADJUSTED, DISPUTE, or IGNORE."""
    result = reconciliation_service.resolve_result(
        db, result_id,
        resolution=payload.get("resolution", ""),
        notes=payload.get("notes"),
        user_id=current_user.user_id,
        customer_id=current_user.customer_id,
    )
    return _serialize_recon_result(result, db)


@router.post("/reconciliation/results/{result_id}/approve")
def approve_reconciliation_adjustment(
    result_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context),
):
    """Corporate admin approves a reconciliation adjustment — updates the LG record."""
    result = reconciliation_service.approve_adjustment(
        db, result_id,
        admin_user_id=current_user.user_id,
        customer_id=current_user.customer_id,
    )
    return _serialize_recon_result(result, db)


@router.post("/reconciliation/results/{result_id}/reject-approval")
def reject_reconciliation_adjustment(
    result_id: int,
    payload: Dict = Body(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context),
):
    """Corporate admin rejects a reconciliation adjustment."""
    result = reconciliation_service.reject_adjustment(
        db, result_id,
        admin_user_id=current_user.user_id,
        customer_id=current_user.customer_id,
        reason=payload.get("reason"),
    )
    return _serialize_recon_result(result, db)


@router.post("/reconciliation/sessions/{session_id}/complete")
def complete_reconciliation_session(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """Mark a reconciliation session as complete (all items must be resolved)."""
    session = reconciliation_service.complete_session(
        db, session_id,
        user_id=current_user.user_id,
        customer_id=current_user.customer_id,
    )
    return _serialize_recon_session(session, db)


@router.get("/reconciliation/pending-approvals")
def get_pending_reconciliation_approvals(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context),
):
    """Get all reconciliation results pending corporate admin approval."""
    results = db.query(ReconResultModel).join(
        ReconSession,
        ReconResultModel.session_id == ReconSession.id,
    ).filter(
        ReconSession.customer_id == current_user.customer_id,
        ReconResultModel.approval_status == "PENDING_APPROVAL",
    ).all()
    return [_serialize_recon_result(r, db) for r in results]


# ── Serialization helpers ──

def _serialize_recon_session(session, db: Session, brief: bool = False) -> dict:
    bank = db.query(Bank).filter(Bank.id == session.bank_id).first()
    result = {
        "id": session.id,
        "bank_id": session.bank_id,
        "bank_name": bank.name if bank else None,
        "position_date": str(session.position_date),
        "status": session.status,
        "file_format": session.file_format,
        "original_file_name": session.original_file_name,
        "total_bank_records": session.total_bank_records,
        "matched_count": session.matched_count,
        "mismatched_count": session.mismatched_count,
        "bank_only_count": session.bank_only_count,
        "system_only_count": session.system_only_count,
        "notes": session.notes,
        "created_at": str(session.created_at) if session.created_at else None,
    }
    if not brief:
        result.update({
            "parsing_method": session.parsing_method,
            "bank_reported_total": str(session.bank_reported_total) if session.bank_reported_total else None,
            "bank_reported_count": session.bank_reported_count,
            "completeness_status": session.completeness_status,
            "completeness_note": session.completeness_note,
            "error_message": session.error_message,
            "reviewed_at": str(session.reviewed_at) if session.reviewed_at else None,
        })
    return result


def _serialize_recon_result(result, db: Session) -> dict:
    lg = None
    if result.issued_lg_id:
        lg = db.query(IssuedLGRecord).filter(IssuedLGRecord.id == result.issued_lg_id).first()
    bank_row = None
    if result.bank_row_id:
        bank_row = db.query(ReconBankRow).filter(
            ReconBankRow.id == result.bank_row_id
        ).first()

    return {
        "id": result.id,
        "session_id": result.session_id,
        "mismatch_type": result.mismatch_type,
        "severity": result.severity,
        "field_name": result.field_name,
        "bank_value": result.bank_value,
        "system_value": result.system_value,
        "user_resolution": result.user_resolution,
        "resolution_notes": result.resolution_notes,
        "approval_status": result.approval_status,
        "record_updated": result.record_updated,
        # Context
        "lg_ref_number": lg.lg_ref_number if lg else None,
        "lg_bank_number": lg.bank_lg_number if lg else None,
        "lg_beneficiary": lg.beneficiary_name if lg else None,
        "lg_status": lg.status if lg else None,
        "bank_row_lg_number": bank_row.bank_lg_number if bank_row else None,
        "bank_row_beneficiary": bank_row.beneficiary_name if bank_row else None,
        "bank_row_amount": str(bank_row.amount) if bank_row and bank_row.amount else None,
    }


# ==============================================================================
# ADMIN DUAL-CONTROL ENDPOINTS
# ==============================================================================

@router.get("/admin/change-requests", response_model=List[AdminChangeRequestOut])
def list_change_requests(
    status_filter: Optional[str] = Query(None, description="PENDING, APPROVED, REJECTED"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """List all admin change requests for this customer."""
    query = db.query(AdminChangeRequest).filter(
        AdminChangeRequest.customer_id == current_user.customer_id,
    )
    if status_filter:
        query = query.filter(AdminChangeRequest.status == status_filter.upper())
    requests = query.order_by(AdminChangeRequest.created_at.desc()).all()
    
    results = []
    for req in requests:
        out = AdminChangeRequestOut.model_validate(req)
        out.requested_by_email = req.requested_by.email if req.requested_by else None
        out.approved_by_email = req.approved_by.email if req.approved_by else None
        results.append(out)
    return results


@router.post("/admin/change-requests", response_model=AdminChangeRequestOut, status_code=201)
def create_change_request(
    payload: AdminChangeRequestCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Submit a new admin change request requiring dual-control approval."""
    new_req = AdminChangeRequest(
        customer_id=current_user.customer_id,
        requested_by_user_id=current_user.user_id,
        change_type=payload.change_type,
        change_payload=payload.change_payload,
        status="PENDING"
    )
    db.add(new_req)
    db.commit()
    db.refresh(new_req)
    
    out = AdminChangeRequestOut.model_validate(new_req)
    out.requested_by_email = new_req.requested_by.email if new_req.requested_by else None
    return out


# ---------------------------------------------------------------------------
# K1: Governance — Auto‑Apply Logic & Helpers
# ---------------------------------------------------------------------------

def _apply_admin_change(db: Session, change_req: AdminChangeRequest):
    """Auto-apply an approved AdminChangeRequest based on its change_type."""
    from datetime import datetime as _dt
    payload = change_req.change_payload or {}
    ct = change_req.change_type

    if ct == "FORM_CONFIG_UPDATE":
        config = db.query(CustomerFormConfiguration).filter(
            CustomerFormConfiguration.customer_id == change_req.customer_id
        ).first()
        if not config:
            config = CustomerFormConfiguration(customer_id=change_req.customer_id)
            db.add(config)
        new_val = payload.get("new_value", {})
        if "field_configurations" in new_val:
            config.field_configurations = new_val["field_configurations"]
        if "custom_field_1_config" in new_val:
            config.custom_field_1_config = new_val["custom_field_1_config"]
        if "custom_field_2_config" in new_val:
            config.custom_field_2_config = new_val["custom_field_2_config"]
        if "mandatory_document_types" in new_val:
            config.mandatory_document_types = new_val["mandatory_document_types"]
        if "reference_types" in new_val:
            config.reference_types = new_val["reference_types"]
        if "document_config" in new_val:
            config.document_config = new_val["document_config"]

    elif ct == "APPROVAL_MATRIX_UPDATE":
        # Bulk-replace workflow policies
        db.query(IssuanceWorkflowPolicy).filter(
            IssuanceWorkflowPolicy.customer_id == change_req.customer_id
        ).delete()
        amount_types = {"AMOUNT_OVER", "AMOUNT_RANGE"}
        for idx, p in enumerate(payload.get("new_value", [])):
            db_obj = IssuanceWorkflowPolicy(
                customer_id=change_req.customer_id,
                step_sequence=idx + 1,
                condition_type=p.get("condition_type", "ALWAYS"),
                condition_value=str(p["condition_value"]) if p.get("condition_value") else None,
                currency_id=p.get("currency_id") if p.get("condition_type") in amount_types else None,
                approver_type=p.get("approver_type", "ROLE"),
                approver_values=p.get("approver_values", []),
                required_signatures=p.get("required_signatures", 1),
                is_active=True
            )
            db.add(db_obj)

    elif ct == "DEPARTMENT_CREATE":
        from app.crud.crud_org import crud_department
        from app.schemas.all_schemas import DepartmentCreate
        dept_data = DepartmentCreate(**payload.get("new_value", {}))
        crud_department.create_dept(db, dept_data, change_req.customer_id, change_req.requested_by_user_id)

    elif ct == "DEPARTMENT_UPDATE":
        from app.crud.crud_org import crud_department
        from app.schemas.all_schemas import DepartmentUpdate
        dept = crud_department.get(db, id=payload.get("entity_id"))
        if dept and dept.customer_id == change_req.customer_id:
            update_data = DepartmentUpdate(**payload.get("new_value", {}))
            crud_department.update_dept(db, dept, update_data, change_req.requested_by_user_id)

    elif ct == "GROUP_CREATE":
        from app.crud.crud_org import crud_approval_group
        from app.schemas.all_schemas import ApprovalGroupCreate
        grp_data = ApprovalGroupCreate(**payload.get("new_value", {}))
        crud_approval_group.create_group(db, grp_data, change_req.customer_id, change_req.requested_by_user_id)

    elif ct == "GROUP_UPDATE":
        from app.crud.crud_org import crud_approval_group
        from app.schemas.all_schemas import ApprovalGroupUpdate
        grp = crud_approval_group.get(db, id=payload.get("entity_id"))
        if grp and grp.customer_id == change_req.customer_id:
            update_data = ApprovalGroupUpdate(**payload.get("new_value", {}))
            crud_approval_group.update_group(db, grp, update_data, change_req.requested_by_user_id)

    db.flush()


def _create_governed_change(
    db: Session, customer_id: int, user_id: int,
    change_type: str, change_payload: dict
) -> tuple:
    """
    Create an AdminChangeRequest for dual-control.
    If only 1 corp admin exists, auto-approve and apply immediately
    (still creates the record for audit).
    Returns (change_req, auto_approved: bool).
    """
    from app.models.models import User
    from app.constants import UserRole
    from datetime import datetime as _dt

    new_req = AdminChangeRequest(
        customer_id=customer_id,
        requested_by_user_id=user_id,
        change_type=change_type,
        change_payload=change_payload,
        status="PENDING"
    )

    # Single-admin exception: auto-approve
    corp_admin_count = db.query(User).filter(
        User.customer_id == customer_id,
        User.role == UserRole.CORPORATE_ADMIN,
        User.is_deleted == False
    ).count()

    auto_approved = corp_admin_count <= 1
    if auto_approved:
        new_req.status = "APPROVED"
        new_req.approved_by_user_id = user_id
        new_req.applied_at = _dt.utcnow()

    db.add(new_req)
    db.flush()

    if auto_approved:
        _apply_admin_change(db, new_req)

    db.commit()
    db.refresh(new_req)
    return new_req, auto_approved


@router.post("/admin/change-requests/{request_id}/action", response_model=AdminChangeRequestOut)
def action_change_request(
    request_id: int,
    action_payload: AdminChangeRequestAction,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Approve or reject an admin change request (dual-control)."""
    change_req = db.query(AdminChangeRequest).filter(
        AdminChangeRequest.id == request_id,
        AdminChangeRequest.customer_id == current_user.customer_id
    ).first()
    if not change_req:
        raise HTTPException(status_code=404, detail="Change request not found")
    if change_req.status != "PENDING":
        raise HTTPException(status_code=400, detail=f"Request is already {change_req.status}")
    
    # Dual-control: approver must be different from requester
    if change_req.requested_by_user_id == current_user.user_id:
        raise HTTPException(status_code=403, detail="Cannot approve/reject your own change request")
    
    action = action_payload.action.upper()
    if action == "APPROVE":
        change_req.status = "APPROVED"
        change_req.approved_by_user_id = current_user.user_id
        from datetime import datetime
        change_req.applied_at = datetime.utcnow()
        # K1: Apply the change automatically on approval
        _apply_admin_change(db, change_req)
    elif action == "REJECT":
        change_req.status = "REJECTED"
        change_req.approved_by_user_id = current_user.user_id
        change_req.rejection_reason = action_payload.rejection_reason
    else:
        raise HTTPException(status_code=400, detail="Action must be APPROVE or REJECT")
    
    db.commit()
    db.refresh(change_req)
    
    out = AdminChangeRequestOut.model_validate(change_req)
    out.requested_by_email = change_req.requested_by.email if change_req.requested_by else None
    out.approved_by_email = change_req.approved_by.email if change_req.approved_by else None
    return out


# ==============================================================================
# BANK FORM ISSUE REPORTING ENDPOINTS
# ==============================================================================

@router.get("/bank-form-issues", response_model=List[BankFormIssueReportOut])
def list_bank_form_issues(
    status_filter: Optional[str] = Query(None),
    bank_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """List bank form issue reports for this customer."""
    query = db.query(BankFormIssueReport).filter(
        BankFormIssueReport.customer_id == current_user.customer_id,
    )
    if status_filter:
        query = query.filter(BankFormIssueReport.status == status_filter.upper())
    if bank_id:
        query = query.filter(BankFormIssueReport.bank_id == bank_id)
    reports = query.order_by(BankFormIssueReport.created_at.desc()).all()
    
    results = []
    for report in reports:
        out = BankFormIssueReportOut.model_validate(report)
        out.reported_by_email = report.reported_by.email if report.reported_by else None
        out.bank_name = report.bank.name if report.bank else None
        results.append(out)
    return results


@router.post("/bank-form-issues", response_model=BankFormIssueReportOut, status_code=201)
def create_bank_form_issue(
    payload: BankFormIssueReportCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context)
):
    """Report an issue with a bank form."""
    new_report = BankFormIssueReport(
        customer_id=current_user.customer_id,
        reported_by_user_id=current_user.user_id,
        bank_id=payload.bank_id,
        form_config_id=payload.form_config_id,
        issue_type=payload.issue_type,
        description=payload.description,
        field_name=payload.field_name,
        severity=payload.severity,
        status="OPEN"
    )
    db.add(new_report)
    db.commit()
    db.refresh(new_report)
    
    out = BankFormIssueReportOut.model_validate(new_report)
    out.reported_by_email = new_report.reported_by.email if new_report.reported_by else None
    out.bank_name = new_report.bank.name if new_report.bank else None
    return out


@router.patch("/bank-form-issues/{issue_id}", response_model=BankFormIssueReportOut)
def update_bank_form_issue(
    issue_id: int,
    payload: BankFormIssueReportUpdate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_corporate_admin_context)
):
    """Update a bank form issue report (admin only)."""
    report = db.query(BankFormIssueReport).filter(
        BankFormIssueReport.id == issue_id,
        BankFormIssueReport.customer_id == current_user.customer_id,
    ).first()
    if not report:
        raise HTTPException(status_code=404, detail="Issue report not found")
    
    if payload.status:
        report.status = payload.status
        if payload.status in ("RESOLVED", "CLOSED"):
            from datetime import datetime
            report.resolved_at = datetime.utcnow()
    if payload.resolution_notes is not None:
        report.resolution_notes = payload.resolution_notes
    if payload.severity:
        report.severity = payload.severity
    
    db.commit()
    db.refresh(report)
    
    out = BankFormIssueReportOut.model_validate(report)
    out.reported_by_email = report.reported_by.email if report.reported_by else None
    out.bank_name = report.bank.name if report.bank else None
    return out


# ==============================================================================
# 3.2 BANK FORM GAP DETECTION
# ==============================================================================

@router.post("/bank-forms/{form_id}/gap-analysis/{request_id}")
def bank_form_gap_analysis(
    form_id: int,
    request_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    3.2: After auto-fill, compare request fields vs mapped form fields.
    Alert user about unmapped critical details and offer supplementary letter option.
    """
    from sqlalchemy.orm import selectinload
    from app.models.models_issuance import BankFormTemplate

    form_template = db.query(BankFormTemplate).filter(
        BankFormTemplate.id == form_id,
        BankFormTemplate.is_deleted == False,
    ).first()
    if not form_template:
        raise HTTPException(404, "Bank form template not found.")
    if not form_template.field_mapping:
        raise HTTPException(400, "Form has no field mapping. Run AI analysis first.")

    # Load request with relationships
    request = db.query(IssuanceRequest).options(
        selectinload(IssuanceRequest.currency),
        selectinload(IssuanceRequest.lg_type),
        selectinload(IssuanceRequest.issuing_entity),
        selectinload(IssuanceRequest.customer),
        selectinload(IssuanceRequest.project),
    ).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id,
    ).first()
    if not request:
        raise HTTPException(404, "Issuance request not found.")

    # Build the data dict the same way fill_bank_form does
    from app.core.pdf_form_filler import build_request_data_dict
    request_data = build_request_data_dict(request, db, bank_id=form_template.bank_id)

    # Analyze gaps: which form fields couldn't get filled?
    field_mapping = form_template.field_mapping  # {form_field_name: request_data_key}
    filled_fields = []
    empty_fields = []
    unmapped_request_fields = []

    # Critical request fields that should ideally be on the form
    critical_fields = {
        "beneficiary_name", "amount", "currency", "expiry_date", "issue_date",
        "lg_purpose", "beneficiary_address", "reference_number",
        "applicable_rules", "issuance_country",
    }

    # Check what the form maps vs what data we have
    mapped_request_keys = set()
    for form_field, data_key in field_mapping.items():
        mapped_request_keys.add(data_key)
        value = request_data.get(data_key)
        if value and str(value).strip():
            filled_fields.append({
                "form_field": form_field,
                "data_key": data_key,
                "value": str(value)[:100],
            })
        else:
            empty_fields.append({
                "form_field": form_field,
                "data_key": data_key,
                "reason": "Request data is empty or missing for this field",
            })

    # Find critical request fields NOT mapped to any form field
    for crit_field in critical_fields:
        if crit_field not in mapped_request_keys:
            value = request_data.get(crit_field)
            if value and str(value).strip():
                unmapped_request_fields.append({
                    "field": crit_field,
                    "value": str(value)[:100],
                    "severity": "HIGH",
                    "suggestion": f"This critical field '{crit_field}' has data but no corresponding form field.",
                })

    has_gaps = bool(empty_fields or unmapped_request_fields)

    # Build supplementary letter suggestion if gaps exist
    supplementary_letter = None
    if unmapped_request_fields:
        letter_lines = [f"- {f['field']}: {f['value']}" for f in unmapped_request_fields]
        supplementary_letter = {
            "suggested": True,
            "reason": f"{len(unmapped_request_fields)} critical field(s) have data but no form field.",
            "content_preview": "\n".join(letter_lines),
        }

    return {
        "form_id": form_id,
        "form_name": form_template.name,
        "request_id": request_id,
        "serial_number": request.serial_number,
        "has_gaps": has_gaps,
        "summary": {
            "total_form_fields": len(field_mapping),
            "filled": len(filled_fields),
            "empty": len(empty_fields),
            "unmapped_critical": len(unmapped_request_fields),
        },
        "filled_fields": filled_fields,
        "empty_fields": empty_fields,
        "unmapped_critical_fields": unmapped_request_fields,
        "supplementary_letter": supplementary_letter,
    }


# ==============================================================================
# 3.3 RECONCILIATION HEADER DRIFT DETECTION
# ==============================================================================

from app.services.reconciliation_service import reconciliation_service as recon_service

@router.post("/reconciliation/check-headers")
def check_reconciliation_headers(
    bank_id: int = Body(...),
    headers: List[str] = Body(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """
    3.3: Compare uploaded file headers against cached bank column mapping.
    Detect column drift (new/missing columns) and warn user before parsing.
    """
    return recon_service.detect_header_drift(
        db, bank_id, current_user.customer_id, headers
    )


@router.post("/reconciliation/re-analyze-mapping")
def re_analyze_reconciliation_mapping(
    bank_id: int = Body(...),
    headers: List[str] = Body(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_treasury_context),
):
    """
    3.3: Clear cached column mapping and re-run keyword analysis on new headers.
    Called after user confirms re-analysis when header drift is detected.
    """
    return recon_service.re_analyze_mapping(
        db, bank_id, current_user.customer_id, headers
    )


# ==============================================================================
# 4.2 TREASURY DASHBOARD STATS
# ==============================================================================

@router.get("/dashboard-stats")
def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    4.2: Treasury dashboard KPI endpoint.
    Returns aggregated stats: pending requests, approvals, SLA breaches,
    expiring LGs, facility utilization, and recent activity.
    """
    from datetime import datetime, timedelta, date, timezone
    from sqlalchemy import func as sqla_func

    cust_id = current_user.customer_id
    today = date.today()
    d7 = today + timedelta(days=7)
    d30 = today + timedelta(days=30)

    # --- Pending Requests ---
    pending_requests = db.query(sqla_func.count(IssuanceRequest.id)).filter(
        IssuanceRequest.customer_id == cust_id,
        IssuanceRequest.status.in_(["SUBMITTED", "PENDING_APPROVAL", "RETURNED_FOR_REVISION"])
    ).scalar() or 0

    # --- Pending Approvals (for this user) ---
    pending_approvals_q = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == cust_id,
        IssuanceRequest.status == "PENDING_APPROVAL",
    ).all()
    pending_approvals = sum(
        1 for r in pending_approvals_q
        if r.pending_approver_users and str(current_user.user_id) in [str(u) for u in (r.pending_approver_users or [])]
    )

    # --- Pending Bank Replies ---
    pending_bank = db.query(sqla_func.count(IssuedLGRecord.id)).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "PENDING_CONFIRMATION",
    ).scalar() or 0

    # --- SLA Breaches (requests pending > sla_agreement_days on their facility) ---
    sla_breaches = 0
    pending_reqs = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == cust_id,
        IssuanceRequest.status.in_(["PENDING_APPROVAL", "SUBMITTED", "APPROVED_INTERNAL", "FACILITY_RESERVED"])
    ).all()
    for req in pending_reqs:
        if req.created_at:
            age_days = (datetime.now(timezone.utc) - req.created_at).days
            if age_days > 7:  # Default SLA threshold
                sla_breaches += 1

    # --- Expiring LGs ---
    expiring_7d = db.query(sqla_func.count(IssuedLGRecord.id)).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
        IssuedLGRecord.expiry_date != None,
        IssuedLGRecord.expiry_date <= d7,
        IssuedLGRecord.expiry_date >= today,
    ).scalar() or 0

    expiring_30d = db.query(sqla_func.count(IssuedLGRecord.id)).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
        IssuedLGRecord.expiry_date != None,
        IssuedLGRecord.expiry_date <= d30,
        IssuedLGRecord.expiry_date >= today,
    ).scalar() or 0

    # --- Active LGs totals ---
    active_stats = db.query(
        sqla_func.count(IssuedLGRecord.id),
        sqla_func.coalesce(sqla_func.sum(IssuedLGRecord.current_amount), 0)
    ).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
    ).first()

    total_active_lgs = active_stats[0] or 0
    total_active_amount = float(active_stats[1] or 0)

    # --- Facility Utilization per Bank ---
    facilities = db.query(IssuanceFacility).filter(
        IssuanceFacility.customer_id == cust_id,
        IssuanceFacility.status == "ACTIVE",
        IssuanceFacility.is_deleted == False,
    ).all()

    facility_utilization = []
    for fac in facilities:
        total_limit = float(fac.total_limit_amount or 0)
        if total_limit <= 0:
            continue
        # Sum utilized amount from active LGs under this facility's sub-limits
        sub_limit_ids = [sl.id for sl in (fac.sub_limits or [])]
        utilized = 0
        if sub_limit_ids:
            utilized = float(db.query(
                sqla_func.coalesce(sqla_func.sum(IssuedLGRecord.current_amount), 0)
            ).filter(
                IssuedLGRecord.facility_sub_limit_id.in_(sub_limit_ids),
                IssuedLGRecord.status.in_(["ACTIVE", "PENDING_CONFIRMATION"]),
            ).scalar() or 0)
            # Add initial utilization from sub-limits
            for sl in fac.sub_limits:
                utilized += float(sl.initial_utilization or 0)

        used_pct = round((utilized / total_limit) * 100, 1) if total_limit > 0 else 0
        bank_name = fac.bank.name if fac.bank else f"Bank #{fac.bank_id}"
        facility_utilization.append({
            "facility_id": fac.id,
            "facility_name": fac.facility_name,
            "bank": bank_name,
            "total_limit": total_limit,
            "utilized": utilized,
            "available": max(total_limit - utilized, 0),
            "used_pct": used_pct,
            "currency": fac.currency.iso_code if fac.currency else "N/A",
        })

    # --- Expiring LGs list (for table) ---
    expiring_lgs_list = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
        IssuedLGRecord.expiry_date != None,
        IssuedLGRecord.expiry_date <= d30,
        IssuedLGRecord.expiry_date >= today,
    ).order_by(IssuedLGRecord.expiry_date.asc()).limit(10).all()

    expiring_lgs = []
    for lg in expiring_lgs_list:
        days_remaining = (lg.expiry_date - today).days
        expiring_lgs.append({
            "id": lg.id,
            "ref": lg.lg_ref_number,
            "beneficiary": lg.beneficiary_name,
            "amount": float(lg.current_amount or 0),
            "currency": lg.currency.iso_code if lg.currency else "",
            "expiry_date": lg.expiry_date.isoformat() if lg.expiry_date else None,
            "days_remaining": days_remaining,
            "bank": lg.bank.name if lg.bank else "",
        })

    # --- Recent Activity (latest 10 audit trail entries) ---
    recent_lgs = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == cust_id,
    ).order_by(IssuedLGRecord.created_at.desc()).limit(5).all()

    recent_actions = db.query(IssuanceMaintenanceAction).join(
        IssuedLGRecord, IssuedLGRecord.id == IssuanceMaintenanceAction.issued_lg_id
    ).filter(
        IssuedLGRecord.customer_id == cust_id,
    ).order_by(IssuanceMaintenanceAction.created_at.desc()).limit(5).all()

    recent_activity = []
    for lg in recent_lgs:
        recent_activity.append({
            "type": "ISSUED",
            "ref": lg.lg_ref_number,
            "description": f"LG issued to {lg.beneficiary_name}",
            "timestamp": lg.created_at.isoformat() if lg.created_at else None,
        })
    for action in recent_actions:
        lg = action.issued_lg
        recent_activity.append({
            "type": f"MAINTENANCE_{action.action_type}",
            "ref": lg.lg_ref_number if lg else f"LG#{action.issued_lg_id}",
            "description": f"{action.action_type.replace('_', ' ').title()} — {action.status}",
            "timestamp": action.created_at.isoformat() if action.created_at else None,
        })
    # Sort combined by timestamp desc
    recent_activity.sort(key=lambda x: x.get("timestamp") or "", reverse=True)
    recent_activity = recent_activity[:10]

    return {
        "pending_requests": pending_requests,
        "pending_approvals": pending_approvals,
        "pending_bank_replies": pending_bank,
        "sla_breaches": sla_breaches,
        "expiring_lgs_7d": expiring_7d,
        "expiring_lgs_30d": expiring_30d,
        "total_active_lgs": total_active_lgs,
        "total_active_amount": total_active_amount,
        "facility_utilization": facility_utilization,
        "expiring_lgs": expiring_lgs,
        "recent_activity": recent_activity,
    }


# ==============================================================================
# H2: Supporting Document Analysis During Request
# ==============================================================================

@router.post("/requests/{request_id}/analyze-document")
async def analyze_supporting_document_endpoint(
    request_id: int,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    H2: Upload a supporting document (Contract, PO, Formal Request) for AI analysis.
    Cross-references extracted fields against the issuance request.
    ADVISORY only — highlights potential gaps, never blocks the user.
    """
    from app.core.ai_integration import analyze_supporting_document, AI_DOC_MAX_SIZE_BYTES

    # Validate request
    request_obj = db.query(IssuanceRequest).filter(
        IssuanceRequest.id == request_id,
        IssuanceRequest.customer_id == current_user.customer_id,
    ).first()
    if not request_obj:
        raise HTTPException(404, "Issuance request not found")

    pdf_bytes = await file.read()

    # File size guard (friendly message, not an error)
    if len(pdf_bytes) > AI_DOC_MAX_SIZE_BYTES:
        return {
            "status": "TOO_LARGE",
            "message": f"Document is too large for AI analysis ({len(pdf_bytes) / (1024*1024):.1f} MB). Maximum is 5 MB.",
            "comparison": None,
        }

    # Run AI extraction
    ai_result = await analyze_supporting_document(
        pdf_bytes, doc_type.upper(), file.filename,
        db=db, customer_id=current_user.customer_id, user_id=current_user.user_id,
    )

    if ai_result["status"] != "OK":
        return ai_result

    extracted = ai_result["extracted_fields"]

    # Build advisory comparison against request fields
    from decimal import Decimal
    comparison = []

    def _compare(field, request_val, doc_val, label):
        if doc_val is None:
            return
        match = False
        if request_val is not None:
            if isinstance(request_val, (int, float, Decimal)):
                try:
                    match = abs(float(request_val) - float(doc_val)) < 0.01
                except (ValueError, TypeError):
                    match = False
            else:
                match = str(request_val).strip().lower() == str(doc_val).strip().lower()
        comparison.append({
            "field": field,
            "label": label,
            "request_value": str(request_val) if request_val is not None else None,
            "document_value": str(doc_val),
            "match": match,
            "severity": "info" if match else ("warning" if request_val is not None else "suggestion"),
        })

    # Cross-reference fields based on doc type — ONLY compare fields found in document
    doc_type_upper = doc_type.upper()

    # Amount comparison (doc_type agnostic)
    amount_field = {"CONTRACT": "contract_value", "PURCHASE_ORDER": "po_value", "FORMAL_REQUEST": "requested_amount"}.get(doc_type_upper, "contract_value")
    _compare("amount", request_obj.amount, extracted.get(amount_field), f"Document Amount vs Request Amount")

    # Beneficiary name
    ben_field = {"CONTRACT": "parties_involved", "PURCHASE_ORDER": "vendor_name", "FORMAL_REQUEST": "requested_beneficiary"}.get(doc_type_upper, "beneficiary_name")
    _compare("beneficiary_name", request_obj.beneficiary_name, 
             extracted.get(ben_field) or extracted.get("beneficiary_name"), "Beneficiary Name")

    # Beneficiary address
    _compare("beneficiary_address", request_obj.beneficiary_address,
             extracted.get("beneficiary_address"), "Beneficiary Address")

    # Reference number
    ref_field = {"PURCHASE_ORDER": "po_number"}.get(doc_type_upper, "reference_number")
    _compare("reference_number", getattr(request_obj, 'reference_number', None), 
             extracted.get(ref_field) or extracted.get("reference_number"), "Reference Number")

    # LG Type — bidirectional fuzzy word match
    if extracted.get("lg_type_hint") and hasattr(request_obj, 'lg_type') and request_obj.lg_type:
        lg_type_name = getattr(request_obj.lg_type, 'name', '')
        doc_lg_type = extracted.get("lg_type_hint", '')
        req_words = set(w.lower() for w in lg_type_name.split() if len(w) > 2)
        doc_words = set(w.lower() for w in doc_lg_type.split() if len(w) > 2)
        if req_words and doc_words:
            overlap = len(req_words & doc_words)
            union = len(req_words | doc_words)
            lg_match = (overlap / union) >= 0.5 if union > 0 else False
        else:
            lg_match = lg_type_name.lower() in doc_lg_type.lower() or doc_lg_type.lower() in lg_type_name.lower()
        comparison.append({
            "field": "lg_type",
            "label": "LG Type",
            "request_value": lg_type_name,
            "document_value": doc_lg_type,
            "match": lg_match,
            "severity": "info" if lg_match else "warning",
        })

    # Currency
    if extracted.get("currency_code") and hasattr(request_obj, 'currency') and request_obj.currency:
        currency_match = getattr(request_obj.currency, 'iso_code', '') == extracted.get("currency_code")
        comparison.append({
            "field": "currency",
            "label": "Currency",
            "request_value": getattr(request_obj.currency, 'iso_code', None),
            "document_value": extracted.get("currency_code"),
            "match": currency_match,
            "severity": "info" if currency_match else "warning",
        })

    # Payable currency
    if extracted.get("payable_currency") and hasattr(request_obj, 'payable_currency') and request_obj.payable_currency:
        pay_match = getattr(request_obj.payable_currency, 'iso_code', '') == extracted.get("payable_currency")
        comparison.append({
            "field": "payable_currency",
            "label": "Payable Currency",
            "request_value": getattr(request_obj.payable_currency, 'iso_code', None),
            "document_value": extracted.get("payable_currency"),
            "match": pay_match,
            "severity": "info" if pay_match else "warning",
        })

    # Maturity / Expiry Date — smart comparison with duration parsing
    doc_maturity = extracted.get("maturity_date")
    req_expiry = request_obj.requested_expiry_date
    if doc_maturity and req_expiry:
        import re
        from datetime import date, timedelta
        expiry_match = False
        doc_display = doc_maturity
        req_display = str(req_expiry)
        duration_pattern = re.search(r'(\d+)\s*(month|year|day|week)s?', doc_maturity, re.IGNORECASE)
        if duration_pattern:
            num = int(duration_pattern.group(1))
            unit = duration_pattern.group(2).lower()
            today = date.today()
            if unit in ('month',):
                expected_expiry = today + timedelta(days=num * 30)
            elif unit in ('year',):
                expected_expiry = today + timedelta(days=num * 365)
            elif unit in ('week',):
                expected_expiry = today + timedelta(weeks=num)
            else:
                expected_expiry = today + timedelta(days=num)
            delta = abs((req_expiry - expected_expiry).days)
            expiry_match = delta <= 15
            doc_display = f"{doc_maturity} (≈ {expected_expiry.isoformat()}, {delta}d diff)"
        else:
            try:
                from dateutil import parser as dateparser
                doc_date = dateparser.parse(doc_maturity, dayfirst=True).date()
                delta = abs((req_expiry - doc_date).days)
                expiry_match = delta <= 7
                doc_display = f"{doc_maturity} ({doc_date.isoformat()}, {delta}d diff)"
            except Exception:
                expiry_match = str(req_expiry) in doc_maturity or doc_maturity in str(req_expiry)
        comparison.append({"field": "requested_expiry_date", "label": "Maturity / Expiry Date",
                           "request_value": req_display, "document_value": doc_display,
                           "match": expiry_match,
                           "severity": "info" if expiry_match else "warning"})

    # Purpose
    if extracted.get("purpose") and request_obj.lg_purpose:
        # Fuzzy: check if significant words overlap
        req_words = set(w.lower() for w in (request_obj.lg_purpose or '').split() if len(w) > 3)
        doc_words = set(w.lower() for w in extracted.get("purpose", '').split() if len(w) > 3)
        overlap = len(req_words & doc_words)
        purpose_match = overlap >= min(2, len(req_words)) if req_words else False
        comparison.append({
            "field": "purpose",
            "label": "LG Purpose",
            "request_value": request_obj.lg_purpose,
            "document_value": extracted.get("purpose"),
            "match": purpose_match,
            "severity": "info" if purpose_match else "suggestion",
        })

    # Special conditions (advisory only)
    if extracted.get("special_conditions"):
        comparison.append({
            "field": "special_conditions",
            "label": "Special Conditions Found in Document",
            "request_value": request_obj.other_conditions or "None specified",
            "document_value": "; ".join(extracted["special_conditions"]),
            "match": True,  # Always informational
            "severity": "info",
        })

    result = {
        "status": "OK",
        "message": None,
        "doc_type": doc_type_upper,
        "summary": extracted.get("summary"),
        "comparison": comparison,
        "all_extracted_fields": extracted,
        "mismatches": len([c for c in comparison if not c["match"]]),
        "total_fields_compared": len(comparison),
    }

    # Persist the verification result on the matching document
    from app.models.models_issuance import IssuanceRequestDocument
    doc_record = db.query(IssuanceRequestDocument).filter(
        IssuanceRequestDocument.request_id == request_id,
        IssuanceRequestDocument.document_type == doc_type_upper,
        IssuanceRequestDocument.is_deleted == False
    ).order_by(IssuanceRequestDocument.created_at.desc()).first()
    if doc_record and hasattr(doc_record, 'ai_verification_result'):
        doc_record.ai_verification_result = result
        db.commit()

    return result
