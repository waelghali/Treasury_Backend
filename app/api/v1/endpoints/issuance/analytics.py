from typing import List, Any, Optional, Dict
from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, status, Query, BackgroundTasks, Body, Request, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from datetime import date
import io
import json
import logging

logger = logging.getLogger(__name__)

from app.database import get_db
from app.core.security import get_current_corporate_admin_context, get_current_approver_context, get_current_treasury_context, get_issuance_read_context, check_subscription_status, TokenData
from app.core.document_generator import generate_pdf_from_html
from app.core.encryption import encrypt_data 

# Models
from app.models.models import Bank, Currency 
from app.models.models_issuance import IssuedLGRecord, IssuanceRequest, IssuanceFacilitySubLimit, IssuanceFacility, IssuanceWorkflowPolicy, CustomerFormConfiguration, IssuanceRequestSnapshot, IssuanceRequestVersion, AdminChangeRequest, BankFormIssueReport
# NOTE: Ensure you created app/models/models_reconciliation.py first!
from app.models.models_reconciliation import BankPositionBatch, BankPositionRow 
# Schemas
from app.schemas.all_schemas import BankOut, CurrencyOut 
from app.schemas.schemas_issuance import (
    IssuanceRequestCreate, IssuanceRequestOut, IssuanceRequestUpdate, IssuanceRequestDraftCreate,
    CustomerFormConfigurationCreateUpdate, IssuanceRequestVersionOut,
    IssuanceFacilityCreate, IssuanceFacilityOut, SuitableFacilityOut, 
    IssuanceRequestContentUpdate, IssuanceFacilityUpdate,
    IssuedLGRecordOut, IssuedLGRecordDetailOut,
    IssuanceExecuteRequest, IssuanceCancelRequest,
    ReconciliationRequest, ReconciliationResult,
    IssuanceWorkflowPolicyCreate, IssuanceWorkflowPolicyOut,
    BankIssuanceOptionOut, BankIssuanceOptionCreateUpdate,
    AdminChangeRequestCreate, AdminChangeRequestOut, AdminChangeRequestAction,
    BankFormIssueReportCreate, BankFormIssueReportOut, BankFormIssueReportUpdate
)
from app.services.issuance_service import issuance_service

# CRUD
from app.crud.crud_issuance import crud_issuance_request
from app.crud.crud_facility import crud_facility
from app.crud.crud_bank_methods import crud_bank_methods
from fastapi.responses import StreamingResponse

router = APIRouter()

from .base import *
from .base import _read_bank_form_pdf_bytes, _send_edit_notifications, _detect_coverage_gaps, _make_doc_filename, _get_lg_copy_docs, _send_requestor_status_notification, _serialize_action, _serialize_recon_session, _serialize_recon_result, _apply_admin_change, _create_governed_change

@router.get("/issued-lgs/export")
def export_issued_lgs(
    export_type: str = Query("summary", description="summary | detailed | full_audit"),
    status_filter: Optional[str] = Query(None, description="Filter by status, e.g. ACTIVE"),
    search: Optional[str] = Query(None, description="Search LG ref or beneficiary"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """Export Issued LGs to XLSX. Types: summary, detailed, full_audit."""
    from app.models.models_issuance import IssuedLGRecord, IssuanceRequest
    from app.models import Bank, Currency, User
    from fastapi.responses import StreamingResponse
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed. Run: pip install openpyxl")

    query = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == current_user.customer_id
    )
    if status_filter:
        query = query.filter(IssuedLGRecord.status == status_filter)
    records = query.order_by(IssuedLGRecord.created_at.desc()).all()

    # Optional search filter (in-memory for simplicity)
    if search:
        s = search.lower()
        records = [r for r in records if
                   (r.lg_ref_number and s in r.lg_ref_number.lower()) or
                   (r.beneficiary_name and s in r.beneficiary_name.lower())]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issued LGs"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc'),
    )

    # Define columns based on export type
    # Status label mapping for export
    status_labels = {
        "INTERNAL_PROCESSING": "Processing",
        "DELIVERED_TO_BANK": "At Bank",
        "BANK_INQUIRY": "Bank Inquiry",
        "BANK_REJECTED": "Rejected by Bank",
        "LG_ISSUED": "LG Issued",
        "ACTIVE": "Active",
        "EXPIRED": "Expired",
        "CANCELLED": "Cancelled",
        "PENDING_CLOSE": "Closing",
        "CLOSED": "Closed",
        "LIQUIDATED": "Liquidated",
        "SLA_EXCEEDED": "SLA Breach",
    }

    if export_type == "summary":
        headers = ["Serial", "LG Ref", "Status", "Amount", "Currency", "Expiry Date", "Bank", "Beneficiary"]
    elif export_type == "detailed":
        headers = [
            "Serial", "LG Ref", "Status", "Amount", "Currency",
            "Issue Date", "Expiry Date", "Bank", "Beneficiary",
            "Bank LG Number", "Facility", "Sub-Limit", "Method",
            "Requestor", "Department", "LG Type", "LG Purpose",
            "Reference Type", "Reference #", "Validity Flag",
            "Delivery Date", "Delivery Method",
            "Bank Reply Type", "Bank Reply Date",
            "Verification Status", "Verified At",
            "Handover Date", "Recipient",
            "Custody Holder", "Issued By", "Created At",
        ]
    else:  # full_audit
        headers = [
            "Serial", "LG Ref", "Status", "Amount", "Currency",
            "Issue Date", "Expiry Date", "Bank", "Beneficiary",
            "Bank LG Number", "Facility", "Sub-Limit", "Method",
            "Requestor", "Department", "LG Type", "LG Purpose",
            "Reference Type", "Reference #", "Validity Flag",
            "Delivery Date", "Delivery Method",
            "Bank Reply Type", "Bank Reply Date",
            "Verification Status", "Verified At",
            "Handover Date", "Recipient",
            "Custody Holder", "Issued By", "Created At",
            "Action History", "Approval Chain",
        ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_idx, r in enumerate(records, 2):
        bank_name = r.bank.name if r.bank else "N/A"
        if not r.bank_id and r.sub_limit and r.sub_limit.facility and r.sub_limit.facility.bank:
            bank_name = r.sub_limit.facility.bank.name
        currency_code = r.currency.iso_code if r.currency else "N/A"
        facility_name = r.sub_limit.facility.facility_name if r.sub_limit and r.sub_limit.facility else ""
        sub_limit_name = r.sub_limit.limit_name if r.sub_limit else ""
        owner_name = ""
        if getattr(r, 'current_owner_user_id', None):
            owner = db.query(User).filter(User.id == r.current_owner_user_id).first()
            owner_name = owner.email if owner else ""

        # Friendly status label
        display_status = status_labels.get(r.status, r.status or "")

        # Build row data from request
        ref_type = ""
        ref_number = ""
        requestor_name = ""
        department = ""
        lg_type = ""
        lg_purpose = ""
        issued_by_name = ""
        approval_chain_str = ""
        if r.request_id:
            req = db.query(IssuanceRequest).filter(IssuanceRequest.id == r.request_id).first()
            if req:
                ref_type = req.reference_type or ""
                ref_number = req.reference_number or ""
                lg_type = (req.lg_type.name if req.lg_type else "") or ""
                lg_purpose = req.lg_purpose or ""
                department = req.department or ""
                # Requestor — use stored name, fall back to user email
                requestor_name = req.requestor_name or ""
                if not requestor_name and req.requestor_user_id:
                    requestor = db.query(User).filter(User.id == req.requestor_user_id).first()
                    requestor_name = requestor.email if requestor else ""
                # Issued by
                if req.issued_by_user_id:
                    issuer = db.query(User).filter(User.id == req.issued_by_user_id).first()
                    issued_by_name = issuer.email if issuer else ""
                # Approval chain
                if export_type == "full_audit" and req.approval_chain_audit:
                    import json as json_mod
                    approval_chain_str = json_mod.dumps(req.approval_chain_audit, default=str)

        # Tracking fields
        delivery_date = str(getattr(r, 'delivery_date', '') or '')
        delivery_method = getattr(r, 'delivery_method', '') or ''
        bank_reply_type = getattr(r, 'bank_reply_type', '') or ''
        bank_reply_date = str(getattr(r, 'bank_reply_date', '') or '')
        verification_status = getattr(r, 'verification_status', '') or ''
        verified_at = str(getattr(r, 'verified_at', '') or '')
        handover_date = str(getattr(r, 'handover_date', '') or '')
        recipient_name = getattr(r, 'recipient_name', '') or ''
        internal_serial = getattr(r, 'internal_serial', '') or ''
        bank_lg_number = getattr(r, 'bank_lg_number', '') or ''

        if export_type == "summary":
            row_data = [
                internal_serial, r.lg_ref_number, display_status,
                float(r.current_amount), currency_code,
                str(r.expiry_date) if r.expiry_date else "", bank_name, r.beneficiary_name,
            ]
        elif export_type == "detailed":
            row_data = [
                internal_serial, r.lg_ref_number, display_status,
                float(r.current_amount), currency_code,
                str(r.issue_date) if r.issue_date else "", str(r.expiry_date) if r.expiry_date else "",
                bank_name, r.beneficiary_name, bank_lg_number,
                facility_name, sub_limit_name, r.issuance_method or "",
                requestor_name, department, lg_type, lg_purpose,
                ref_type, ref_number, getattr(r, 'reference_validity_flag', "") or "",
                delivery_date, delivery_method,
                bank_reply_type, bank_reply_date,
                verification_status, verified_at,
                handover_date, recipient_name,
                r.custody_holder or "", issued_by_name,
                r.created_at.isoformat() if r.created_at else "",
            ]
        else:  # full_audit
            import json
            history = json.dumps(r.action_history or r.custody_transfer_log or [], default=str)
            row_data = [
                internal_serial, r.lg_ref_number, display_status,
                float(r.current_amount), currency_code,
                str(r.issue_date) if r.issue_date else "", str(r.expiry_date) if r.expiry_date else "",
                bank_name, r.beneficiary_name, bank_lg_number,
                facility_name, sub_limit_name, r.issuance_method or "",
                requestor_name, department, lg_type, lg_purpose,
                ref_type, ref_number, getattr(r, 'reference_validity_flag', "") or "",
                delivery_date, delivery_method,
                bank_reply_type, bank_reply_date,
                verification_status, verified_at,
                handover_date, recipient_name,
                r.custody_holder or "", issued_by_name,
                r.created_at.isoformat() if r.created_at else "",
                history, approval_chain_str,
            ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"issued_lgs_{export_type}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/dashboard-stats")
def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_issuance_read_context),
):
    """
    4.2: Treasury dashboard KPI endpoint.
    Returns aggregated stats: pending requests, approvals, SLA breaches,
    expiring LGs, facility utilization, and recent activity.
    """
    from datetime import datetime, timedelta, date, timezone
    from sqlalchemy import func as sqla_func

    cust_id = current_user.customer_id
    today = date.today()
    d7 = today + timedelta(days=7)
    d30 = today + timedelta(days=30)

    # --- Pending Requests ---
    pending_requests = db.query(sqla_func.count(IssuanceRequest.id)).filter(
        IssuanceRequest.customer_id == cust_id,
        IssuanceRequest.status.in_(["SUBMITTED", "PENDING_APPROVAL", "REVISION_REQUIRED"])
    ).scalar() or 0

    # --- Pending Approvals (for this user) ---
    pending_approvals_q = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == cust_id,
        IssuanceRequest.status == "PENDING_APPROVAL",
    ).all()
    pending_approvals = sum(
        1 for r in pending_approvals_q
        if r.pending_approver_users and str(current_user.user_id) in [str(u) for u in (r.pending_approver_users or [])]
    )

    # --- Pending Bank Replies ---
    pending_bank = db.query(sqla_func.count(IssuedLGRecord.id)).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "INTERNAL_PROCESSING",
    ).scalar() or 0

    # --- SLA Breaches (requests pending > sla_agreement_days on their facility) ---
    sla_breaches = 0
    pending_reqs = db.query(IssuanceRequest).filter(
        IssuanceRequest.customer_id == cust_id,
        IssuanceRequest.status.in_(["PENDING_APPROVAL", "SUBMITTED", "APPROVED_INTERNAL", "FACILITY_RESERVED"])
    ).all()
    for req in pending_reqs:
        if req.created_at:
            age_days = (datetime.now(timezone.utc) - req.created_at).days
            if age_days > 7:  # Default SLA threshold
                sla_breaches += 1

    # --- Expiring LGs ---
    expiring_7d = db.query(sqla_func.count(IssuedLGRecord.id)).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
        IssuedLGRecord.expiry_date != None,
        IssuedLGRecord.expiry_date <= d7,
        IssuedLGRecord.expiry_date >= today,
    ).scalar() or 0

    expiring_30d = db.query(sqla_func.count(IssuedLGRecord.id)).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
        IssuedLGRecord.expiry_date != None,
        IssuedLGRecord.expiry_date <= d30,
        IssuedLGRecord.expiry_date >= today,
    ).scalar() or 0

    # --- Active LGs totals ---
    active_stats = db.query(
        sqla_func.count(IssuedLGRecord.id),
        sqla_func.coalesce(sqla_func.sum(IssuedLGRecord.current_amount), 0)
    ).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
    ).first()

    total_active_lgs = active_stats[0] or 0
    total_active_amount = float(active_stats[1] or 0)

    # --- Facility Utilization per Bank ---
    facilities = db.query(IssuanceFacility).filter(
        IssuanceFacility.customer_id == cust_id,
        IssuanceFacility.status == "ACTIVE",
        IssuanceFacility.is_deleted == False,
    ).all()

    facility_utilization = []
    for fac in facilities:
        total_limit = float(fac.total_limit_amount or 0)
        if total_limit <= 0:
            continue
        # Sum utilized amount from active LGs under this facility's sub-limits
        sub_limit_ids = [sl.id for sl in (fac.sub_limits or [])]
        utilized = 0
        if sub_limit_ids:
            utilized = float(db.query(
                sqla_func.coalesce(sqla_func.sum(IssuedLGRecord.current_amount), 0)
            ).filter(
                IssuedLGRecord.facility_sub_limit_id.in_(sub_limit_ids),
                IssuedLGRecord.status.in_(["ACTIVE", "INTERNAL_PROCESSING"]),
            ).scalar() or 0)
            # Add initial utilization from sub-limits
            for sl in fac.sub_limits:
                utilized += float(sl.initial_utilization or 0)

        used_pct = round((utilized / total_limit) * 100, 1) if total_limit > 0 else 0
        bank_name = fac.bank.name if fac.bank else f"Bank #{fac.bank_id}"
        facility_utilization.append({
            "facility_id": fac.id,
            "facility_name": fac.facility_name,
            "bank": bank_name,
            "total_limit": total_limit,
            "utilized": utilized,
            "available": max(total_limit - utilized, 0),
            "used_pct": used_pct,
            "currency": fac.currency.iso_code if fac.currency else "N/A",
        })

    # --- Expiring LGs list (for table) ---
    expiring_lgs_list = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == cust_id,
        IssuedLGRecord.status == "ACTIVE",
        IssuedLGRecord.expiry_date != None,
        IssuedLGRecord.expiry_date <= d30,
        IssuedLGRecord.expiry_date >= today,
    ).order_by(IssuedLGRecord.expiry_date.asc()).limit(10).all()

    expiring_lgs = []
    for lg in expiring_lgs_list:
        days_remaining = (lg.expiry_date - today).days
        expiring_lgs.append({
            "id": lg.id,
            "ref": lg.lg_ref_number,
            "beneficiary": lg.beneficiary_name,
            "amount": float(lg.current_amount or 0),
            "currency": lg.currency.iso_code if lg.currency else "",
            "expiry_date": lg.expiry_date.isoformat() if lg.expiry_date else None,
            "days_remaining": days_remaining,
            "bank": lg.bank.name if lg.bank else "",
        })

    # --- Recent Activity (latest 10 audit trail entries) ---
    recent_lgs = db.query(IssuedLGRecord).filter(
        IssuedLGRecord.customer_id == cust_id,
    ).order_by(IssuedLGRecord.created_at.desc()).limit(5).all()

    recent_actions = db.query(IssuanceMaintenanceAction).join(
        IssuedLGRecord, IssuedLGRecord.id == IssuanceMaintenanceAction.issued_lg_id
    ).filter(
        IssuedLGRecord.customer_id == cust_id,
    ).order_by(IssuanceMaintenanceAction.created_at.desc()).limit(5).all()

    recent_activity = []
    for lg in recent_lgs:
        recent_activity.append({
            "type": "ISSUED",
            "ref": lg.lg_ref_number,
            "description": f"LG issued to {lg.beneficiary_name}",
            "timestamp": lg.created_at.isoformat() if lg.created_at else None,
        })
    for action in recent_actions:
        lg = action.issued_lg
        recent_activity.append({
            "type": f"MAINTENANCE_{action.action_type}",
            "ref": lg.lg_ref_number if lg else f"LG#{action.issued_lg_id}",
            "description": f"{action.action_type.replace('_', ' ').title()} — {action.status}",
            "timestamp": action.created_at.isoformat() if action.created_at else None,
        })
    # Sort combined by timestamp desc
    recent_activity.sort(key=lambda x: x.get("timestamp") or "", reverse=True)
    recent_activity = recent_activity[:10]

    return {
        "pending_requests": pending_requests,
        "pending_approvals": pending_approvals,
        "pending_bank_replies": pending_bank,
        "sla_breaches": sla_breaches,
        "expiring_lgs_7d": expiring_7d,
        "expiring_lgs_30d": expiring_30d,
        "total_active_lgs": total_active_lgs,
        "total_active_amount": total_active_amount,
        "facility_utilization": facility_utilization,
        "expiring_lgs": expiring_lgs,
        "recent_activity": recent_activity,
    }


# ==============================================================================
# H2: Supporting Document Analysis During Request
# ==============================================================================

