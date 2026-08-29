# scripts/generate_all_test_assets.py
import sys, os
sys.path.insert(0, os.path.abspath("."))
import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Output directories
BASE_DIR = os.path.join(os.getcwd(), "test_assets")
SETUP_DIR = os.path.join(BASE_DIR, "01_setup")
ISSUANCE_DIR = os.path.join(BASE_DIR, "02_issuance_requests")
CUSTODY_DIR = os.path.join(BASE_DIR, "03_custody_lg_scans")
RECON_DIR = os.path.join(BASE_DIR, "04_reconciliation_statements")
GUIDES_DIR = os.path.join(BASE_DIR, "05_testing_guides")

for d in [SETUP_DIR, ISSUANCE_DIR, CUSTODY_DIR, RECON_DIR, GUIDES_DIR]:
    os.makedirs(d, exist_ok=True)

styles = getSampleStyleSheet()

# Typography Styles
title_style = ParagraphStyle(
    "DocTitle",
    parent=styles["Heading1"],
    fontName="Helvetica-Bold",
    fontSize=14,
    leading=18,
    textColor=colors.HexColor("#1e3a8a"),
    alignment=TA_CENTER,
    spaceAfter=3,
)

subtitle_style = ParagraphStyle(
    "DocSubtitle",
    parent=styles["Normal"],
    fontName="Helvetica-Bold",
    fontSize=9,
    leading=12,
    textColor=colors.HexColor("#475569"),
    alignment=TA_CENTER,
    spaceAfter=10,
)

h2_style = ParagraphStyle(
    "SectionH2",
    parent=styles["Heading2"],
    fontName="Helvetica-Bold",
    fontSize=10.5,
    leading=14,
    textColor=colors.HexColor("#0f172a"),
    spaceBefore=7,
    spaceAfter=4,
)

body_style = ParagraphStyle(
    "BodyDark",
    parent=styles["Normal"],
    fontName="Helvetica",
    fontSize=8.5,
    leading=12,
    textColor=colors.HexColor("#1e293b"),
    alignment=TA_JUSTIFY,
    spaceAfter=5,
)

body_bold = ParagraphStyle(
    "BodyBold",
    parent=body_style,
    fontName="Helvetica-Bold",
)

cell_style = ParagraphStyle(
    "CellText",
    parent=styles["Normal"],
    fontName="Helvetica",
    fontSize=8,
    leading=10.5,
    textColor=colors.HexColor("#1e293b"),
)

cell_bold = ParagraphStyle(
    "CellTextBold",
    parent=cell_style,
    fontName="Helvetica-Bold",
)

cell_header = ParagraphStyle(
    "CellHeader",
    parent=cell_style,
    fontName="Helvetica-Bold",
    textColor=colors.white,
    alignment=TA_CENTER,
)

notice_style = ParagraphStyle(
    "NoticeText",
    parent=styles["Normal"],
    fontName="Helvetica",
    fontSize=8,
    leading=11.5,
    textColor=colors.HexColor("#1e3a8a"),
)

def add_page_decorations(canvas, doc, header_title=""):
    canvas.saveState()
    # Header on page 2+
    if doc.page > 1 and header_title:
        canvas.setFont("Helvetica-Bold", 7.5)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(36, 812, header_title)
        canvas.drawRightString(559, 812, "CONFIDENTIAL — EVALUATION ASSET")
        canvas.setStrokeColor(colors.HexColor("#cbd5e1"))
        canvas.setLineWidth(0.5)
        canvas.line(36, 806, 559, 806)
    
    # Footer on all pages
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawString(36, 22, "Grow Treasury Management Platform — Staging Evaluation Playbook")
    canvas.drawRightString(559, 22, f"Page {doc.page}")
    canvas.setStrokeColor(colors.HexColor("#e2e8f0"))
    canvas.setLineWidth(0.5)
    canvas.line(36, 32, 559, 32)
    canvas.restoreState()

# ==============================================================================
# 1. SETUP PHASE: ENBD Facility Sanction Letter
# ==============================================================================
def create_enbd_facility_contract():
    filepath = os.path.join(SETUP_DIR, "ENBD_Credit_Facility_Agreement_Sanction_Letter.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story = []

    story.append(Paragraph("EMIRATES NBD EGYPT S.A.E.", title_style))
    story.append(Paragraph("CORPORATE BANKING DIVISION — CREDIT SANCTION ADVICE", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1e3a8a"), spaceAfter=8))

    meta_data = [
        [Paragraph("<b>Facility Ref No:</b>", cell_style), Paragraph("ENBD/CORP/LG/2026/0942", cell_style),
         Paragraph("<b>Sanction Date:</b>", cell_style), Paragraph(datetime.date.today().strftime("%d-%b-%Y"), cell_style)],
        [Paragraph("<b>Borrower / Customer:</b>", cell_style), Paragraph("Apex Global / Horizon / Delta Modern", cell_style),
         Paragraph("<b>Facility Type:</b>", cell_style), Paragraph("Revolving LG Credit Facility", cell_style)],
        [Paragraph("<b>Approved Total Limit:</b>", cell_style), Paragraph("<b>EGP 60,000,000.00</b>", cell_bold),
         Paragraph("<b>Facility Expiry:</b>", cell_style), Paragraph((datetime.date.today() + datetime.timedelta(days=365)).strftime("%d-%b-%Y"), cell_style)],
        [Paragraph("<b>Agreed SLA Commitment:</b>", cell_style), Paragraph("<b>1.0 Business Day (Fast-Track)</b>", cell_bold),
         Paragraph("<b>Facility Default Margin:</b>", cell_style), Paragraph("5.00% Cash Margin", cell_style)],
    ]
    t = Table(meta_data, colWidths=[120, 140, 115, 148])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 8))

    story.append(Paragraph("1. APPROVED SUB-LIMIT ALLOCATION", h2_style))
    sub_data = [
        [Paragraph("Sub-Limit Name", cell_header), 
         Paragraph("Approved Cap", cell_header), 
         Paragraph("Max Per LG", cell_header), 
         Paragraph("Max Tenor", cell_header), 
         Paragraph("Commission Rate", cell_header), 
         Paragraph("Cash Margin", cell_header)],
        
        [Paragraph("<b>Bid Bonds (Tender)</b>", cell_style), 
         Paragraph("EGP 15,000,000", cell_style), 
         Paragraph("EGP 5,000,000", cell_style), 
         Paragraph("180 Days", cell_style), 
         Paragraph("0.75% per quarter (Min 250 EGP)", cell_style), 
         Paragraph("0.00%", cell_style)],
        
        [Paragraph("<b>Performance Guarantees</b>", cell_style), 
         Paragraph("EGP 30,000,000", cell_style), 
         Paragraph("EGP 15,000,000", cell_style), 
         Paragraph("730 Days", cell_style), 
         Paragraph("1.10% per quarter (Min 500 EGP)", cell_style), 
         Paragraph("5.00%", cell_style)],
        
        [Paragraph("<b>Advance Payment LGs</b>", cell_style), 
         Paragraph("EGP 15,000,000", cell_style), 
         Paragraph("EGP 10,000,000", cell_style), 
         Paragraph("365 Days", cell_style), 
         Paragraph("1.25% per quarter (Min 500 EGP)", cell_style), 
         Paragraph("10.00%", cell_style)],
    ]
    st = Table(sub_data, colWidths=[120, 75, 75, 55, 140, 58])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e3a8a")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ('TOPPADDING', (0,0), (-1,-1), 3.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3.5),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(st)
    story.append(Spacer(1, 8))

    story.append(Paragraph("2. SPECIAL COVENANTS & FAST-TRACK ISSUANCE", h2_style))
    story.append(Paragraph(
        "<b>Agreed Issuance SLA:</b> Emirates NBD commits to execute and dispatch all compliant Letters of Guarantee within <b>1.0 business day (24 hours)</b> from receipt of the automated electronic bank form. If issuance exceeds 2.0 business days without formal discrepancy notice, standard processing fees shall be waived.",
        body_style
    ))
    story.append(Paragraph(
        "<b>Form Fillability:</b> LG applications submitted under this facility must utilize Emirates NBD's standardized Corporate LG Application Form (Template Ref: ENBD-LG-V2). Automated digital overlays generated through Grow Treasury platform are pre-approved by operations.",
        body_style
    ))
    story.append(Paragraph(
        "<b>Operating Bank Account:</b> All cash margin allocations and quarterly commission debits will be processed directly from Customer Operating Account No. <b>EG44-0010-0094-1100-2938-4729</b>.",
        body_style
    ))
    story.append(Spacer(1, 10))

    sig_data = [
        [Paragraph("<b>For Emirates NBD Egypt S.A.E.</b>", cell_bold), Paragraph("<b>Accepted on Behalf of Borrower</b>", cell_bold)],
        [Spacer(1, 15), Spacer(1, 15)],
        [Paragraph("____________________________<br/>Senior Director - Corporate Credit", cell_style),
         Paragraph("____________________________<br/>Managing Director / Treasury Director", cell_style)]
    ]
    sig_t = Table(sig_data, colWidths=[260, 263])
    sig_t.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(sig_t)

    doc.build(story, onFirstPage=lambda c, d: add_page_decorations(c, d, "BANK SANCTION ADVICE - EMIRATES NBD"),
                     onLaterPages=lambda c, d: add_page_decorations(c, d, "BANK SANCTION ADVICE - EMIRATES NBD"))
    print(f"[CREATED] {filepath}")

# ==============================================================================
# 2. ISSUANCE REQUEST DOCUMENTS
# ==============================================================================
def create_issuance_request_docs():
    # Request 1: NAT Monorail Award Letter (Standard Issuance)
    f1 = os.path.join(ISSUANCE_DIR, "Request_1_Monorail_Package_B_Award_Letter.pdf")
    doc1 = SimpleDocTemplate(f1, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story1 = []
    story1.append(Paragraph("NATIONAL AUTHORITY FOR TUNNELS (NAT)", title_style))
    story1.append(Paragraph("MINISTRY OF TRANSPORT — ARAB REPUBLIC OF EGYPT", subtitle_style))
    story1.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0284c7"), spaceAfter=8))

    story1.append(Paragraph("OFFICIAL NOTICE OF AWARD — LETTER OF ACCEPTANCE", h2_style))
    story1.append(Paragraph(
        "<b>Tender Reference:</b> NAT-MONO-2026-PKG-B<br/>"
        "<b>Project Title:</b> New Capital Monorail - Track Electrification & Signaling Package B<br/>"
        "<b>Awarded Contractor:</b> Joint Venture Consortium / Contractor<br/>"
        "<b>Date of Notice:</b> " + datetime.date.today().strftime("%d-%B-%Y"),
        body_style
    ))
    story1.append(Spacer(1, 5))
    story1.append(Paragraph(
        "We are pleased to inform you that your tender submitted for the above-referenced contract has been accepted by the Supreme Tendering Committee for the total contract sum of <b>EGP 70,000,000.00 (Seventy Million Egyptian Pounds)</b>.",
        body_style
    ))
    story1.append(Paragraph("CONDITIONS PRECEDENT TO CONTRACT EXECUTION:", h2_style))
    story1.append(Paragraph(
        "In accordance with Section IV (General Conditions of Tender), you are hereby requested to provide the following within <b>10 calendar days</b> from the date of this letter:<br/><br/>"
        "1. <b>Final Performance Guarantee (LG):</b> An unconditional, irrevocable Letter of Guarantee representing <b>5% of the total contract value</b> amounting to <b>EGP 3,500,000.00 (Three Million Five Hundred Thousand Egyptian Pounds)</b>, valid for a minimum period of <b>12 months (365 days)</b> with automatic renewal until Final Acceptance.<br/>"
        "2. <b>Beneficiary Name:</b> National Authority for Tunnels (NAT), Ramses Square, Cairo, Egypt.<br/>"
        "3. <b>Advance Payment Guarantee (Optional):</b> Upon contract signing, an Advance Payment LG of <b>EGP 7,000,000.00 (10%)</b> will be required for mobilization.",
        body_style
    ))
    story1.append(Spacer(1, 10))
    story1.append(Paragraph("Eng. Tarek Hassan<br/>Chairman of Board of Directors, NAT", body_bold))
    doc1.build(story1, onFirstPage=lambda c, d: add_page_decorations(c, d, "OFFICIAL NOTICE OF AWARD - NAT"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "OFFICIAL NOTICE OF AWARD - NAT"))
    print(f"[CREATED] {f1}")

    # Request 2: Monorail Near-Match / Potential Duplicate
    f2 = os.path.join(ISSUANCE_DIR, "Request_2_Monorail_Package_B_Near_Match_Duplicate.pdf")
    doc2 = SimpleDocTemplate(f2, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story2 = []
    story2.append(Paragraph("NATIONAL AUTHORITY FOR TUNNELS (NAT)", title_style))
    story2.append(Paragraph("REVISED SPECIFICATION & SUB-SECTION AMENDMENT", subtitle_style))
    story2.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#dc2626"), spaceAfter=8))

    story2.append(Paragraph("TENDER NOTICE / DUPLICATE CHECK TRIGGER", h2_style))
    story2.append(Paragraph(
        "<b>Tender Reference:</b> NAT-MONO-2026-PKG-B (Sub-Section 2)<br/>"
        "<b>Beneficiary:</b> National Authority for Tunnels (NAT)<br/>"
        "<b>Contract Value:</b> EGP 70,000,000.00<br/>"
        "<b>Required Guarantee:</b> Performance Guarantee for <b>EGP 3,500,000.00</b>",
        body_style
    ))
    story2.append(Paragraph(
        "<i>Testing Note: Submitting this second request in the portal will trigger the platform's AI Near-Match & Duplicate Detection engine, alerting the Approver and Processing Officer that a request with identical Beneficiary ('NAT') and matching Amount ('3,500,000.00 EGP') is already in flight.</i>",
        body_style
    ))
    doc2.build(story2, onFirstPage=lambda c, d: add_page_decorations(c, d, "DUPLICATE VERIFICATION TEST - NAT"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "DUPLICATE VERIFICATION TEST - NAT"))
    print(f"[CREATED] {f2}")

    # Request 3: ENBD Form Auto-Fill Tender Specifications
    f3 = os.path.join(ISSUANCE_DIR, "Request_3_ENBD_Issuance_Tender_Specifications.pdf")
    doc3 = SimpleDocTemplate(f3, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story3 = []
    story3.append(Paragraph("MINISTRY OF ELECTRICITY & RENEWABLE ENERGY", title_style))
    story3.append(Paragraph("NEW & RENEWABLE ENERGY AUTHORITY (NREA)", subtitle_style))
    story3.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#059669"), spaceAfter=8))

    story3.append(Paragraph("BENBAN SOLAR PARK PHASE 3 — TENDER SPECIFICATIONS", h2_style))
    story3.append(Paragraph(
        "<b>Project:</b> Benban 50MW Solar Substation Grid Connection<br/>"
        "<b>Beneficiary:</b> New & Renewable Energy Authority (NREA)<br/>"
        "<b>Required Guarantee Type:</b> Bid Bond / Tender Guarantee<br/>"
        "<b>Guarantee Amount:</b> <b>EGP 1,250,000.00 (One Million Two Hundred Fifty Thousand EGP)</b><br/>"
        "<b>Tenor:</b> 180 Days from Tender Submission Date<br/>"
        "<b>Target Bank:</b> Emirates NBD Egypt (ENBD)",
        body_style
    ))
    story3.append(Paragraph(
        "<i>Testing Guidance: Run this request through the 3-Step Issuance Wizard choosing Emirates NBD (ENBD). The system will auto-populate the ENBD Corporate Application Form with exact coordinates for Arabic/English fields and provide a live overlay preview for download.</i>",
        body_style
    ))
    doc3.build(story3, onFirstPage=lambda c, d: add_page_decorations(c, d, "TENDER SPECIFICATIONS - NREA"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "TENDER SPECIFICATIONS - NREA"))
    print(f"[CREATED] {f3}")

# ==============================================================================
# 3. CUSTODY MODULE: 4 Sample LG Scans
# ==============================================================================
def create_custody_lg_scans():
    # LG 1: Special Liquidation Wording Trigger
    f1 = os.path.join(CUSTODY_DIR, "LG_01_Special_Liquidation_Wording_Scan.pdf")
    doc1 = SimpleDocTemplate(f1, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story1 = []
    story1.append(Paragraph("COMMERCIAL INTERNATIONAL BANK (EGYPT) S.A.E.", title_style))
    story1.append(Paragraph("TRADE FINANCE & GUARANTEES DEPARTMENT — CAIRO", subtitle_style))
    story1.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1e3a8a"), spaceAfter=8))

    story1.append(Paragraph("LETTER OF GUARANTEE — ADVANCE PAYMENT", h2_style))
    meta = [
        [Paragraph("<b>LG Number:</b>", cell_style), Paragraph("<b>LG/CIB/2025/ADV-88910</b>", cell_bold),
         Paragraph("<b>Issue Date:</b>", cell_style), Paragraph("15-Sep-2025", cell_style)],
        [Paragraph("<b>Beneficiary:</b>", cell_style), Paragraph("Apex Global / Horizon / Delta Modern", cell_style),
         Paragraph("<b>Expiry Date:</b>", cell_style), Paragraph("30-Jun-2026", cell_style)],
        [Paragraph("<b>Applicant:</b>", cell_style), Paragraph("Siemens Energy Transmission SAE", cell_style),
         Paragraph("<b>Amount:</b>", cell_style), Paragraph("<b>EGP 4,200,000.00</b>", cell_bold)],
    ]
    t = Table(meta, colWidths=[110, 150, 110, 153])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]))
    story1.append(t)
    story1.append(Spacer(1, 8))

    story1.append(Paragraph("MANDATORY SPECIAL LIQUIDATION & CLAIM CONDITIONS", h2_style))
    story1.append(Paragraph(
        "We hereby open in your favor this irrevocable Letter of Guarantee for an amount of EGP 4,200,000.00. Notwithstanding any standard terms, any demand or liquidation under this guarantee is <b>STRICTLY SUBJECT TO THE FOLLOWING SPECIAL CONDITIONS:</b>",
        body_style
    ))
    story1.append(Paragraph(
        "1. <b>Original Guarantee Presentation:</b> Payment will be made only upon physical surrender of this original Letter of Guarantee to our Trade Operations Counter at CIB Smart Village.<br/>"
        "2. <b>Independent Engineer Certificate:</b> The claim demand must be accompanied by a formal default certificate signed by an accredited FIDIC Engineer certifying that the Applicant failed to perform delivery milestones.<br/>"
        "3. <b>Bank Authentication:</b> The signature on the liquidation demand letter must be authenticated through your commercial bank with full SWIFT/telex confirmation.",
        body_style
    ))
    story1.append(Paragraph(
        "<i>Testing Focus: When this LG is uploaded into the Custody Module or triggered for liquidation, the AI engine will parse these terms and alert the Treasury Officer with: 'Special Liquidation Wording Detected: Requires Original Physical Presentation + FIDIC Engineer Certificate.'</i>",
        body_style
    ))
    doc1.build(story1, onFirstPage=lambda c, d: add_page_decorations(c, d, "INWARD GUARANTEE - CIB EGYPT"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "INWARD GUARANTEE - CIB EGYPT"))
    print(f"[CREATED] {f1}")

    # LG 2: Near Maturity - Forced Renewal
    f2 = os.path.join(CUSTODY_DIR, "LG_02_Near_Maturity_Forced_Renewal_Scan.pdf")
    doc2 = SimpleDocTemplate(f2, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story2 = []
    near_exp = (datetime.date.today() + datetime.timedelta(days=7)).strftime("%d-%b-%Y")
    story2.append(Paragraph("NATIONAL BANK OF EGYPT (NBE)", title_style))
    story2.append(Paragraph("PERFORMANCE GUARANTEE — EXPIRING IN 7 DAYS", subtitle_style))
    story2.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#b45309"), spaceAfter=8))

    story2.append(Paragraph(
        f"<b>LG Number:</b> NBE/PERF/2025/1109<br/>"
        f"<b>Beneficiary:</b> Your Company (Customer)<br/>"
        f"<b>Applicant:</b> Petrojet Contracting SAE<br/>"
        f"<b>Amount:</b> EGP 2,800,000.00<br/>"
        f"<b>Maturity / Expiry Date:</b> <font color='#b45309'><b>{near_exp} (URGENT - 7 DAYS REMAINING)</b></font>",
        body_style
    ))
    story2.append(Paragraph(
        "<i>Testing Focus: Uploading or recording this LG in Custody will trigger the 'Near Maturity Action Center'. The user can test: (a) Issuing a Forced Renewal Demand Letter to NBE, (b) Claiming Liquidation, or (c) Extending validity.</i>",
        body_style
    ))
    doc2.build(story2, onFirstPage=lambda c, d: add_page_decorations(c, d, "NBE PERFORMANCE GUARANTEE"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "NBE PERFORMANCE GUARANTEE"))
    print(f"[CREATED] {f2}")

    # LG 3: Auto-Renewal Notice Scan
    f3 = os.path.join(CUSTODY_DIR, "LG_03_Auto_Renewal_Notice_Scan.pdf")
    doc3 = SimpleDocTemplate(f3, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story3 = []
    story3.append(Paragraph("QNB ALAHLI S.A.E.", title_style))
    story3.append(Paragraph("RENEWABLE PERFORMANCE GUARANTEE — 30 DAY EVERGREEN CLAUSE", subtitle_style))
    story3.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1e3a8a"), spaceAfter=8))

    story3.append(Paragraph(
        "<b>LG Number:</b> QNBAA/LG/2025/5541<br/>"
        "<b>Amount:</b> EGP 1,900,000.00<br/>"
        "<b>Applicant:</b> Elsewedy Electric T&D<br/>"
        "<b>Auto-Renewal Clause:</b> This guarantee shall be automatically extended for successive periods of 12 months unless QNB Alahli gives notice of non-renewal at least 30 days prior to the then-current expiry date.",
        body_style
    ))
    story3.append(Paragraph(
        "<i>Testing Focus: Demonstrates how the system automatically flags evergreen guarantees, calculates renewal windows, and runs auto-renewal reminders without manual intervention.</i>",
        body_style
    ))
    doc3.build(story3, onFirstPage=lambda c, d: add_page_decorations(c, d, "QNB EVERGREEN GUARANTEE"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "QNB EVERGREEN GUARANTEE"))
    print(f"[CREATED] {f3}")

    # LG 4: Bank Handover & Delivery Reply Scan
    f4 = os.path.join(CUSTODY_DIR, "LG_04_Bank_Handover_Receipt_Reply_Scan.pdf")
    doc4 = SimpleDocTemplate(f4, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
    story4 = []
    story4.append(Paragraph("COMMERCIAL INTERNATIONAL BANK (CIB)", title_style))
    story4.append(Paragraph("OFFICIAL HANDOVER ACKNOWLEDGMENT & DISPATCH RECEIPT", subtitle_style))
    story4.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#047857"), spaceAfter=8))

    story4.append(Paragraph(
        "<b>Transaction Reference:</b> CIB/ACK/2026/099<br/>"
        "<b>Related LG:</b> LG/CIB/2025/ADV-88910<br/>"
        "<b>Handover Status:</b> Physical Original Delivered to Authorized Courier (Aramex Tracking: ARX-99281-EG)<br/>"
        "<b>Receipt Confirmed by:</b> Eng. Karim Mostafa (Authorized Treasury Officer)",
        body_style
    ))
    story4.append(Paragraph(
        "<i>Testing Focus: Use this document to test the 'Record Delivery / Bank Reply' action in the Custody tracker, marking the physical handover cycle as completed with registered receipt timestamp.</i>",
        body_style
    ))
    doc4.build(story4, onFirstPage=lambda c, d: add_page_decorations(c, d, "CIB HANDOVER ACKNOWLEDGMENT"),
                       onLaterPages=lambda c, d: add_page_decorations(c, d, "CIB HANDOVER ACKNOWLEDGMENT"))
    print(f"[CREATED] {f4}")

# ==============================================================================
# 4. RECONCILIATION: 2 Bank Position Excel Statements
# ==============================================================================
def create_reconciliation_excel_statements():
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "ENBD_Position_Report"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    border = Border(left=Side(style='thin', color='CBD5E1'),
                    right=Side(style='thin', color='CBD5E1'),
                    top=Side(style='thin', color='CBD5E1'),
                    bottom=Side(style='thin', color='CBD5E1'))

    ws1.append(["EMIRATES NBD EGYPT - MONTHLY CORPORATE LG POSITION STATEMENT"])
    ws1.append([f"Report As Of Date: {datetime.date.today().strftime('%d-%b-%Y')} | Account: EG44-0010-0094-1100-2938-4729"])
    ws1.append([])

    cols = ["Bank LG Ref", "Internal Customer Ref", "LG Type", "Beneficiary", "Currency", "Current Amount", "Issue Date", "Expiry Date", "Cash Margin Held", "Status"]
    ws1.append(cols)

    rows1 = [
        ["ENBD-LG-2026-001", "NAT-MONO-PKG-B", "Performance Guarantee", "National Authority for Tunnels", "EGP", 3500000.00, "2026-01-15", "2027-01-15", 175000.00, "ACTIVE"],
        ["ENBD-LG-2026-002", "NREA-BENBAN-01", "Bid Bond", "New & Renewable Energy Authority", "EGP", 1250000.00, "2026-02-01", "2026-08-01", 0.00, "ACTIVE"],
        ["ENBD-LG-2026-003", "ALEX-PORT-ADV", "Advance Payment", "Alexandria Port Authority", "EGP", 5000000.00, "2025-11-20", "2026-11-20", 500000.00, "ACTIVE"]
    ]
    for r in rows1:
        ws1.append(r)

    for cell in ws1[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws1.iter_rows(min_row=4, max_row=7, min_col=1, max_col=10):
        for cell in row:
            cell.border = border

    p1 = os.path.join(RECON_DIR, "ENBD_Monthly_Bank_Position_Statement_Perfect_Match.xlsx")
    wb1.save(p1)
    print(f"[CREATED] {p1}")

    # Statement 2: Discrepancies & Mismatches (CIB)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "CIB_Position_Report"

    ws2.append(["COMMERCIAL INTERNATIONAL BANK (CIB) - MONTHLY LG POSITION STATEMENT"])
    ws2.append([f"Report As Of Date: {datetime.date.today().strftime('%d-%b-%Y')} | Discrepancy Reconciliation Test File"])
    ws2.append([])

    ws2.append(cols)
    rows2 = [
        ["CIB-LG-2026-901", "NAT-MONO-PKG-B", "Performance Guarantee", "National Authority for Tunnels", "EGP", 3750000.00, "2026-01-15", "2027-01-15", 375000.00, "ACTIVE"],
        ["CIB-LG-2025-441", "SIEMENS-SUB-01", "Advance Payment", "Apex Global Engineering", "EGP", 4200000.00, "2025-09-15", "2026-06-30", 420000.00, "ACTIVE"],
        ["CIB-LG-2024-110", "UNKNOWN-REF-99", "Bid Bond", "Egyptian Electricity Holding", "EGP", 850000.00, "2024-05-10", "2024-11-10", 0.00, "EXPIRED_UNRELEASED"]
    ]
    for r in rows2:
        ws2.append(r)

    for cell in ws2[4]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws2.iter_rows(min_row=4, max_row=7, min_col=1, max_col=10):
        for cell in row:
            cell.border = border

    p2 = os.path.join(RECON_DIR, "CIB_Monthly_Bank_Position_Statement_Discrepancies.xlsx")
    wb2.save(p2)
    print(f"[CREATED] {p2}")

# ==============================================================================
# 5. PERSONALIZED PDF TESTING GUIDES
# ==============================================================================
def create_personalized_guides():
    customers_info = [
        {
            "filename": "Apex_Global_Tester_Guide.pdf",
            "company_name": "Apex Global Engineering & Contracting SAE",
            "admin_email": "apex.admin@globex.com",
            "approver_email": "apex.approver@globex.com",
            "officer_email": "apex.officer@globex.com",
            "smart_inbox": "apex.treasury.pilot@gmail.com",
            "smart_inbox_app_pw_clean": "evhhityvnerffqhd",
            "smart_inbox_app_pw_display": "evhh ityv nerf fqhd",
            "account_num": "EG44-0010-0094-1100-2938-4729",
            "project_name": "New Capital Monorail - Package B",
            "project_ref": "PRJ-2026-MONO-01",
        },
        {
            "filename": "Horizon_Infrastructure_Tester_Guide.pdf",
            "company_name": "Horizon Infrastructure & Power SAE",
            "admin_email": "horizon.admin@globex.com",
            "approver_email": "horizon.approver@globex.com",
            "officer_email": "horizon.officer@globex.com",
            "smart_inbox": "horizon.treasury.pilot@gmail.com",
            "smart_inbox_app_pw_clean": "wuftpwgjesrxppwd",
            "smart_inbox_app_pw_display": "wuft pwgj esrx ppwd",
            "account_num": "EG88-0010-0044-2200-8812-9901",
            "project_name": "Benban Solar Park - Substation Grid 3",
            "project_ref": "PRJ-2026-SOLAR-03",
        },
        {
            "filename": "Delta_Modern_Owner_Guide.pdf",
            "company_name": "Delta Modern Contracting & Infrastructure SAE",
            "admin_email": "delta.admin@globex.com",
            "approver_email": "delta.approver@globex.com",
            "officer_email": "delta.officer@globex.com",
            "smart_inbox": "delta.treasury.pilot@gmail.com",
            "smart_inbox_app_pw_clean": "bshllpqtvszfpuim",
            "smart_inbox_app_pw_display": "bshl lpqt vszf puim",
            "account_num": "EG12-0010-0010-3300-5197-7401",
            "project_name": "North Coast Water Desalination Facility",
            "project_ref": "PRJ-2026-DESAL-01",
        }
    ]

    for c in customers_info:
        filepath = os.path.join(GUIDES_DIR, c["filename"])
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=40)
        story = []

        # Header Title
        story.append(Paragraph("GROW TREASURY MANAGEMENT PLATFORM", title_style))
        story.append(Paragraph(f"EVALUATION & TESTING PLAYBOOK — {c['company_name'].upper()}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#1e3a8a"), spaceAfter=8))

        # Important Notice Box (Neutralized emails so PDF viewers don't auto-hyperlink)
        notice_data = [[
            Paragraph(
                "<b>IMPORTANT ACCESS & SECURITY INSTRUCTIONS:</b><br/>"
                "• <b>Staging Portal URL:</b> https://staging.growbusinessdevelopment.com<br/>"
                "• <b>Direct In-App Sign In:</b> Always log in directly via the portal URL. All evaluation accounts have <b>MFA bypassed</b> (zero OTP prompts).<br/>"
                "• <b>Tenant Isolation:</b> Your account is completely isolated to <i>" + c['company_name'] + "</i> with zero visibility into other companies.",
                notice_style
            )
        ]]
        nt = Table(notice_data, colWidths=[523])
        nt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#eff6ff")),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#3b82f6")),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 7),
            ('RIGHTPADDING', (0,0), (-1,-1), 7),
        ]))
        story.append(nt)
        story.append(Spacer(1, 7))

        # Credentials Table
        story.append(Paragraph("1. YOUR PERSONA CREDENTIALS", h2_style))
        cred_data = [
            [Paragraph("Persona / Role", cell_header),
             Paragraph("In-App Login Email", cell_header),
             Paragraph("Password", cell_header),
             Paragraph("Primary Responsibilities", cell_header)],
            
            [Paragraph("<b>Corporate Admin</b>", cell_style),
             Paragraph(c["admin_email"], cell_style),
             Paragraph("<b>DemoPass2026!</b>", cell_style),
             Paragraph("Company setup: Bank accounts, projects, approval rules, facilities.", cell_style)],
            
            [Paragraph("<b>Approver (CFO / Controller)</b>", cell_style),
             Paragraph(c["approver_email"], cell_style),
             Paragraph("<b>DemoPass2026!</b>", cell_style),
             Paragraph("Review, dual-authorize, or reject pending issuance requests.", cell_style)],
            
            [Paragraph("<b>Processing Officer (Treasury)</b>", cell_style),
             Paragraph(c["officer_email"], cell_style),
             Paragraph("<b>DemoPass2026!</b>", cell_style),
             Paragraph("3-Step Issuance Wizard, ENBD form auto-filling, and custody actions.", cell_style)],
        ]
        ct = Table(cred_data, colWidths=[115, 125, 85, 198])
        ct.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1e3a8a")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(ct)
        story.append(Spacer(1, 7))

        # Smart Inbox Details
        story.append(Paragraph("2. DEDICATED SMART INBOX CONFIGURATION", h2_style))
        inbox_data = [
            [Paragraph("<b>Smart Mailbox Address:</b>", cell_style), 
             Paragraph(c["smart_inbox"], cell_bold)],
            [Paragraph("<b>IMAP Server / Port:</b>", cell_style), 
             Paragraph("imap.gmail.com | Port: 993 (SSL Enabled)", cell_style)],
            [Paragraph("<b>IMAP App Password:</b>", cell_style), 
             Paragraph(f"<b>{c['smart_inbox_app_pw_clean']}</b> &nbsp; <i>(or formatted as: {c['smart_inbox_app_pw_display']}) — <b>Enter without spaces</b></i>", cell_style)],
            [Paragraph("<b>Simulated Bank Contact Email:</b>", cell_style), 
             Paragraph("bankdesk.simulation@gmail.com", cell_style)],
        ]
        it = Table(inbox_data, colWidths=[140, 383])
        it.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
            ('TOPPADDING', (0,0), (-1,-1), 2.5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(it)
        story.append(Spacer(1, 7))

        # Glossary of Unusual Fields
        story.append(Paragraph("3. GLOSSARY: UNUSUAL & SPECIALIZED PLATFORM FIELDS", h2_style))
        glossary_text = (
            "• <b>Internal Owner Contact:</b> Project engineer or commercial manager assigned to the guarantee. Used to route expiry warnings and auto-assign requests.<br/>"
            "• <b>Agreed SLA (Days):</b> Bank's contracted turnaround time (e.g. 1.0 day for ENBD). The system tracks live issuance duration and flags SLA breaches.<br/>"
            "• <b>Facility Default Margin %:</b> Cash collateral automatically debited from your operating account upon issuance.<br/>"
            "• <b>Dedicated Sub-Limit Reference:</b> Locks a sub-limit specifically to a single project, preventing other divisions from consuming the cap."
        )
        story.append(Paragraph(glossary_text, body_style))
        story.append(Spacer(1, 5))

        # Step by Step Missions
        story.append(Paragraph("4. STEP-BY-STEP TESTING MISSIONS", h2_style))
        
        # Mission 1: Structured prerequisite setup
        m1 = (
            "<b>PHASE 1: COMPLETE INITIAL COMPANY SETUP (As Corporate Admin)</b><br/>"
            "<i>Configure your company's accounts, projects, form settings, and banking lines in sequence:</i><br/>"
            "<b>Step A (Bank Account Setup):</b> Go to <b>LG Issuance &gt; Bank Accounts</b>. Click <b>Add Bank Account</b> &gt; Select <b>Emirates NBD</b>, Account No: <code>" + c["account_num"] + "</code>, Currency: <b>EGP</b>, Type: <b>Operating Account</b>.<br/>"
            "<b>Step B (Corporate Projects &amp; Contracts):</b> Go to <b>LG Issuance &gt; Bank Facilities</b>. Expand the top collapsible panel <b>'Projects &amp; Contracts'</b> &gt; Click <b>+ Add Project</b> &gt; Name: <code>" + c["project_name"] + "</code>, Reference: <code>" + c["project_ref"] + "</code>.<br/>"
            "<b>Step C (Issuance Request Form Configuration):</b> Go to <b>LG Issuance &gt; Form Configuration</b>. As Corporate Admin, customize the employee request form (enable/disable optional fields, set mandatory documents, and configure reference categories like Contract, Tender, PO).<br/>"
            "<b>Step D (User Governance &amp; Approval Controls):</b><br/>"
            "• <i>User Management:</i> Go to <b>Administration &gt; User Management</b> to inspect your 3 pre-provisioned team accounts or invite new colleagues with Maker / Checker roles.<br/>"
            "• <i>Approval Center &amp; Policies:</i> Approval workflows are driven by Maker-Checker roles. Approvers review items in the <b>Approval Center</b> (Shield icon in sidebar). Governance policies (e.g. Max pending days, auto-escalations) are configured in <b>Administration &gt; Settings &gt; Operational Governance</b>.<br/>"
            "<b>Step E (Facility &amp; Sub-Limits Creation):</b> Go to <b>LG Issuance &gt; Bank Facilities &gt; Create Facility</b>.<br/>"
            "• In <b>Basic Info:</b> Select Bank: <b>Emirates NBD</b>, Name: <code>ENBD Main Revolving Facility</code>, Total Limit: <code>60,000,000 EGP</code>, Tenor: <code>12 Months</code>, Link to Account: <code>" + c["account_num"] + "</code>.<br/>"
            "• In <b>Risk &amp; Governance:</b> Set Agreed SLA: <code>1.0 Day</code>, Margin: <code>5%</code>.<br/>"
            "• In <b>Sub-Limits &amp; Pricing:</b> Add 3 Sub-Limits: <i>Bid Bonds (15M EGP @ 0.75% / qtr)</i>, <i>Performance Guarantees (30M EGP @ 1.10% / qtr)</i>, <i>Advance Payment (15M EGP @ 1.25% / qtr)</i>."
        )
        story.append(Paragraph(m1, body_style))
        story.append(Spacer(1, 5))

        m2 = (
            "<b>PHASE 2: END-TO-END LG ISSUANCE &amp; DUPLICATE DETECTION</b><br/>"
            "1. <b>Initiate Request:</b> Go to <b>Issuance Requests &gt; New Request</b>. Select Project: <code>" + c["project_name"] + "</code>, Beneficiary: <code>National Authority for Tunnels (NAT)</code>, Amount: <code>3,500,000 EGP</code>, Type: <b>Performance Guarantee</b>. Attach <code>Request_1_Monorail_Package_B_Award_Letter.pdf</code>.<br/>"
            "2. <b>Dual Approval:</b> Log in as Approver (<code>" + c["approver_email"] + "</code>) &gt; Open <b>Pending Approvals</b> &gt; Review parameters and click <b>Approve</b>.<br/>"
            "3. <b>3-Step Issuance Wizard:</b> Log in as Processing Officer (<code>" + c["officer_email"] + "</code>) &gt; Open approved request:<br/>"
            "   • <b>Step 1 (Recommendation):</b> Review automated scoring selecting Emirates NBD based on pricing and remaining limit.<br/>"
            "   • <b>Step 2 (ENBD Auto-Fill):</b> View auto-populated ENBD Application Form with exact coordinates. Download completed PDF.<br/>"
            "   • <b>Step 3 (Dispatch):</b> Mark as dispatched to bank. Live SLA tracking timer initiates.<br/>"
            "4. <b>Duplicate Detection Test:</b> Submit <code>Request_2_Monorail_Package_B_Near_Match_Duplicate.pdf</code>. Notice the yellow AI warning identifying the duplicate beneficiary and amount!"
        )
        story.append(Paragraph(m2, body_style))
        story.append(Spacer(1, 5))

        m3 = (
            "<b>PHASE 3: CUSTODY, SPECIAL LIQUIDATION &amp; FORCED RENEWAL</b><br/>"
            "1. Go to <b>Custody Module &gt; Record New LG Received</b>.<br/>"
            "2. Upload <code>LG_01_Special_Liquidation_Wording_Scan.pdf</code>. Observe the AI banner detecting: <i>'Special Liquidation Wording: Requires Original Physical Presentation + FIDIC Engineer Certificate'</i>.<br/>"
            "3. Upload <code>LG_02_Near_Maturity_Forced_Renewal_Scan.pdf</code> (Expiring in 7 days) &gt; Test <b>Forced Renewal Demand</b> in the action center.<br/>"
            "4. Test <b>Record Delivery &amp; Bank Handover</b> using <code>LG_04_Bank_Handover_Receipt_Reply_Scan.pdf</code>.<br/>"
            "5. Test <b>Transaction Cancellation / Rollback</b> on any active amendment."
        )
        story.append(Paragraph(m3, body_style))
        story.append(Spacer(1, 5))

        m4 = (
            "<b>PHASE 4: RECONCILIATION &amp; SMART INBOX POLLING</b><br/>"
            "1. <b>Send Test Bank Position:</b> Send an email from your personal inbox or <code>bankdesk.simulation@gmail.com</code> to <code>" + c["smart_inbox"] + "</code> attaching <code>ENBD_Monthly_Bank_Position_Statement_Perfect_Match.xlsx</code>.<br/>"
            "2. <b>Smart Ingestion:</b> In the platform as Officer, go to <b>Smart Inbox &gt; Click 'Fetch Now'</b>. View automated AI classification.<br/>"
            "3. <b>Discrepancy Resolution:</b> Go to <b>Bank Reconciliation</b> &gt; Upload <code>CIB_Monthly_Bank_Position_Statement_Discrepancies.xlsx</code> and review how the system flags amount mismatches and expired unreleased LGs."
        )
        story.append(Paragraph(m4, body_style))

        doc.build(story, onFirstPage=lambda c, d: add_page_decorations(c, d, "GROW TREASURY - TESTER PLAYBOOK"),
                           onLaterPages=lambda c, d: add_page_decorations(c, d, "GROW TREASURY - TESTER PLAYBOOK"))
        print(f"[CREATED GUIDE] {filepath}")

def main():
    print("=" * 60)
    print("RE-GENERATING ALL TEST ASSETS & COMPREHENSIVE TESTING GUIDES")
    print("=" * 60)
    create_enbd_facility_contract()
    create_issuance_request_docs()
    create_custody_lg_scans()
    create_reconciliation_excel_statements()
    create_personalized_guides()
    print("\n" + "=" * 60)
    print("ALL TEST ASSETS & GUIDES SUCCESSFULLY UPDATED!")
    print("=" * 60)

if __name__ == "__main__":
    main()
